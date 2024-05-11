#!/usr/bin/python
# -*- coding: utf-8 -*-

#A few attributions to community in StackOverflow that helped with the little tricks on GUI building: including selection of table and copying to excel: https://stackoverflow.com/questions/40225270/copy-paste-multiple-items-from-qtableview-in-pyqt4 by user learncode, path expansion by https://stackoverflow.com/questions/56712979/expand-item-in-qtreeview-with-qfilesystemmodel by eyllanesc.
import sys,fileinput,os,time,copy,itertools,fnmatch,shutil,random,multiprocessing,xlsxwriter,tempfile,io,warnings
warnings.filterwarnings(action='ignore',module='.*paramiko.*')
import subprocess32 as subprocess
from datetime import datetime
from copy import deepcopy
import gzip
import paramiko,base64                      
from Bio import SeqIO
from Bio import Application
from Bio import Entrez
from Bio.Seq import Seq
#from Bio.Alphabet import IUPAC
from collections import Counter
from itertools import zip_longest
from PyQt5 import QtWidgets,QtCore,QtGui
from PyQt5.QtWidgets import QPushButton
from pyqtgraph.Qt import QtGui, QtCore
import pyqtgraph as pg
import re
import csv
import datetime
import edlib
import seqpy
import openpyxl,uuid
from numpy import *
import numpy as np
client = paramiko.SSHClient()
class livedetector(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
    notifyProgress=QtCore.pyqtSignal(int)
    notifyMessage=QtCore.pyqtSignal(str)
    def __init__(self,indir,parent=None):
        super(livedetector, self).__init__(parent)
        self.livedir=indir
    def run(self):
        while True:
            interval = 1
            old_f = os.listdir(self.livedir)

            old_f=[x for x in old_f if ".fastq" in x]
            time.sleep(interval)

            new_f = os.listdir(self.livedir)
            new_f=[x for x in new_f if ".fastq" in x]

        #    print old_f,new_f
            self.new_files = list(set(new_f) - set(old_f))
            if len(self.new_files)>0:
                time.sleep(5)
                if self.new_files[0].endswith(".gz"):
                    self.notifyProgress.emit(1)
                elif self.new_files[0].endswith(".fastq"):
                    self.notifyProgress.emit(0)
            #    print "changed",self.new_files
                break

class liveremotedetector(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
    notifyProgress=QtCore.pyqtSignal(int)
    notifyMessage=QtCore.pyqtSignal(str)
    def __init__(self,indir,hostname,username,password,parent=None):
        super(liveremotedetector, self).__init__(parent)
        self.remotepath=indir
        self.hostname=hostname
        self.username=username
        self.password=password
    def run(self):
        def readremotedir(remotepath):
            stdin, stdout, stderr = client.exec_command('ls '+remotepath)
            newlist=[]
            for line in stdout:
                newlist.append(line.encode(encoding="utf-8").strip())
            return newlist
        while True:
            try:
                client = paramiko.SSHClient()
                client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                client.connect(self.hostname, username=self.username, password=self.password)
            #    print "remote transfer running"
                old_f = readremotedir(self.remotepath)
                interval = 1 
                time.sleep(interval)
                new_f = readremotedir(self.remotepath)
           #     print old_f,new_f
                self.new_files = list(set(new_f) - set(old_f))
                client.close()
                if len(self.new_files)>0:
                    self.notifyProgress.emit(1)
               #     print "changed",self.new_files
                    break
            except Exception as e:
                print (e)

class liveremotetransfer(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
    notifyProgress=QtCore.pyqtSignal(int)
    notifyMessage=QtCore.pyqtSignal(str)
    def __init__(self,indir,localpath,inlist,hostname,username,password,parent=None):
        super(liveremotetransfer, self).__init__(parent)
        self.remotepath=indir
        self.localpath=localpath
        self.inlist=inlist
        self.hostname=hostname
        self.username=username
        self.password=password
    def run(self):
#        client=self.client
 #       print "transferring",self.remotepath,self.localpath,self.inlist,self.client
        def transferfile(inlist,client):
            ftp_client=client.open_sftp()
            for f in inlist:
                ftp_client.get(self.remotepath+"/"+f,os.path.join(self.localpath,f))
            ftp_client.close()
        try:
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(self.hostname, username=self.username, password=self.password)
            transferfile(self.inlist,client)
            client.close()
        except Exception as e:
            print (e, "couldn't transfer", self.inlist)
      #  transferfile(self.inlist)
        self.taskFinished.emit(1)

class livefast5detector(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
    notifyProgress=QtCore.pyqtSignal(int)
    notifyMessage=QtCore.pyqtSignal(str)
    def __init__(self,indir,parent=None):
        super(livefast5detector, self).__init__(parent)
        self.fast5dir=indir
    def run(self):
        while True:
            interval = 1
            old_f = os.listdir(self.fast5dir)
            time.sleep(interval)

            new_f = os.listdir(self.fast5dir)

         #   print "olf_f",old_f,new_f
            self.new_files = list(set(new_f) - set(old_f))
            if len(self.new_files)>0:
                with open("inlistfast5.txt",'w') as outfile:
                    outfile.write("\n".join(self.new_files))
                self.notifyProgress.emit(1)
         #       print "new fast5",self.new_files
                break


class livepod5detector(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
    notifyProgress=QtCore.pyqtSignal(int)
    notifyMessage=QtCore.pyqtSignal(str)
    def __init__(self,indir,parent=None):
        super(livepod5detector, self).__init__(parent)
        self.pod5dir=indir
    def run(self):
        while True:
            interval = 1
            old_f = os.listdir(self.pod5dir)
            old_f=[x for x in old_f if x.endswith(".pod5")]
            time.sleep(interval)

            new_f = os.listdir(self.pod5dir)
            new_f=[x for x in new_f if x.endswith(".pod5")]
        #    print("olf_f",old_f,new_f)
            self.new_files = list(set(new_f) - set(old_f))
            if len(self.new_files)>0:
                with open("inlistpod5.txt",'w') as outfile:
                    outfile.write("\n".join(self.new_files))
                self.notifyProgress.emit(1)
             #   print("new pod5",self.new_files)
                break
class livesamplecount(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
    notifyProgress=QtCore.pyqtSignal(int)
    notifyMessage=QtCore.pyqtSignal(str)
    def __init__(self,indir,sampleids,mincoverage,largecoverage,largecoveragestep,maxcoverage,parent=None):
        super(livesamplecount, self).__init__(parent)
        self.livedir=indir
        self.localsampleids=deepcopy(sampleids)
        self.mincoverage=mincoverage
        self.largecoverage=largecoverage
        self.largecoveragestep=largecoveragestep
        self.maxcoverage=maxcoverage
        self.flag=True
   
    def update_sampleids(self,sampleids2):
        if self.flag==True:
            oldsampleids=self.localsampleids
         #   print "length of old-sampleids is",len(oldsampleids),len(sampleids2)
            self.todolist=[]
            for k in sampleids2:
                d=sampleids2[k]-oldsampleids[k]
                if d>=1:
                    if self.maxcoverage==0:
                        if sampleids2[k]>=self.mincoverage:
                            if sampleids2[k]<=self.largecoverage:
                                self.todolist.append(k)
                            else:
                                if d>=self.largecoveragestep:
                                    self.todolist.append(k)
                    else:
                        if sampleids2[k]>=self.mincoverage:
                            if sampleids2[k]<=self.largecoverage:
                                self.todolist.append(k)
                            else:
                                if d>=self.largecoveragestep:
                                    if sampleids2[k]<=self.maxcoverage:
                                        self.todolist.append(k)
            
            if len(self.todolist)>0:
            #    print 'sampledetected',self.todolist
                self.flag=False
                self.localsampleids=deepcopy(sampleids2)
                self.taskFinished.emit(1)
            else:
             #   print 'livetector running no sample detected'
                self.flag=True
    def run(self):
        while True:
            interval = 1 
            time.sleep(interval)
            if self.flag==True:
             #   print "consensus not running"
                self.notifyProgress.emit(1)
            else:
                pass
             #   print "consensus running"
       #     if self.flag==True:
      #          print "running live sample count"
                
        #        self.flag=False
   #         else:
#                break

class liveguppyrun(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
    notifyProgress=QtCore.pyqtSignal(int)
    notifyMessage=QtCore.pyqtSignal(str)
    def __init__(self,guppyparams,outpath,guppyoutpath,iternum,parent=None):
        super(liveguppyrun, self).__init__(parent)
        self.outpath=outpath
        self.guppyparams=guppyparams
        self.guppyoutpath=guppyoutpath
        self.iternum=iternum
    def run(self):
        time.sleep(30)
        stdout=subprocess.check_output(self.guppyparams,shell=True)
        with open(os.path.join(self.outpath,"guppylog"), "wb") as handle:
            handle.write(stdout)
        fastqlist = os.listdir(os.path.join(self.guppyoutpath,"temp"))

        fastqlist=[x for x in fastqlist if x.endswith(".fastq")]
        for fname in fastqlist:
            os.rename(os.path.join(self.guppyoutpath,"temp",fname), os.path.join(self.guppyoutpath,str(self.iternum)+"_"+fname))
        self.taskFinished.emit(1)


class livedoradorun(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
    notifyProgress=QtCore.pyqtSignal(int)
    notifyMessage=QtCore.pyqtSignal(str)
    def __init__(self,doradoparams,outpath,doradooutpath,iternum,inlist,inpath,outpathfastq,parent=None):
        super(livedoradorun, self).__init__(parent)
        self.outpath=outpath
        self.doradoparams=doradoparams
        self.doradooutpath=doradooutpath
        self.iternum=iternum
        self.inlist=inlist
        self.doradoinpath=inpath
        self.outpathfastq=outpathfastq
        
    def run(self):
        time.sleep(30)
        os.mkdir(os.path.join(self.doradoinpath,"temp"+str(self.iternum)))
        for f in self.inlist:
            shutil.copyfile(os.path.join(self.doradoinpath,f), os.path.join(self.doradoinpath,"temp"+str(self.iternum),f))
        self.doradoparams=self.doradoparams.replace("ignorethis","temp"+str(self.iternum))
        stdout=subprocess.check_output(self.doradoparams,shell=True)
        with open(os.path.join(self.doradooutpath,str(self.iternum)+".fastq"), "wb") as handle:
            handle.write(stdout)
        fastqlist = os.listdir(os.path.join(self.doradooutpath,"temp"))

        fastqlist=[x for x in fastqlist if x.endswith(".fastq")]
        for fname in fastqlist:
            os.rename(os.path.join(self.doradooutpath,"temp",fname), os.path.join(self.doradooutpath,str(self.iternum)+"_"+fname))
        self.taskFinished.emit(1)

       
class runconsensuslive(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(list)
    notifyProgress=QtCore.pyqtSignal(int)
    notifyMessage=QtCore.pyqtSignal(int)
    notifyMessage2=QtCore.pyqtSignal(int)
    def __init__(self,outpath,todolist,consensesfreqfixed,consensesfreqrange,consensesfreqstep,gencode,plen,ndone,maxcoverage,subsetmax,sampleids,parent=None):
        super(runconsensuslive, self).__init__(parent)
        self.outpath=outpath
        self.todolist=todolist
        self.fixthresh=consensesfreqfixed
        self.rangefreq=consensesfreqrange
        self.stepsize=consensesfreqstep
        self.ingencode=gencode
        self.plen=plen
        self.ndone=ndone
        self.maxcoverage=maxcoverage
        self.subsetmax=subsetmax
        self.sampleids=sampleids
    def run(self):
        fixthresh=self.fixthresh
        rangefreq=self.rangefreq
        stepsize=self.stepsize
        ingencode=self.ingencode
        maxcoverage=self.maxcoverage
        plen=self.plen
        def consensus(indict,perc_thresh,abs_thresh):
            if len(indict.keys())>=abs_thresh:
                poslist=[]
                n=0
                while n<len(list(indict.values())[0]):
                    newlist=[]
                    for i,each in enumerate(indict.keys()):
                        try:
                            newlist.append(indict[each][n])
                        except IndexError:
                            break
                    poslist.append(newlist)
                    n+=1
                sequence=[]
                countpos=1
                for character in poslist:
                    charcounter=Counter(character)
                    baseset={}
                    for k,v in charcounter.items():
                        percbp=float(v)/float(len(character))
                        if percbp>perc_thresh:
                            baseset[k]=v
                    if len(baseset)==0:
                        bp='N'
                    if len(baseset)==1:
                        bp=list(baseset.keys())[0]
                    if len(baseset)>1:
                        try:
                            del baseset["-"]
                            if len(baseset)==1:
                                bp=list(baseset.keys())[0]
                            else:
                                bp='N'
                        except KeyError:
                            bp="N"
                    sequence.append(bp)
                    countpos+=1
                return ''.join(sequence)
            else:
                return ''
        def subset_bylength (infile,outfile,n,plen):
            samplesize=0
            with open(infile) as fulldata:
                with open(outfile,'w') as subsetdata:
                    l=fulldata.readlines()
                    tseqdict={}
                    tlendict={}
                    for i,j in enumerate(l):
                        if ">" in j:
                            samplesize+=1
                            tlendict[j]=abs(plen-len(l[i+1].strip()))
                            tseqdict[j]=l[i+1]
                    d=sorted(tlendict.items(), key=lambda x: x[1])
                    v=0
                    ntosubset=min(len(tseqdict),n)
                    while v<ntosubset:
                        subsetdata.write(d[v][0]+tseqdict[d[v][0]])
                        v+=1
            return samplesize
        def translate_corframe(seq,gencode):
            translatedset=[]
            each=seq
            corframe=get_cor_frame(each.replace("-",""),gencode)
            if corframe==1:
                translation=Seq(each.replace('-','')).translate(table=gencode,to_stop=True).__str__()
                explen=len(each.replace('-',''))
            if corframe==2:
                translation=Seq(each[1:].replace('-','')).translate(table=gencode,to_stop=True).__str__()
                explen=len(each[1:].replace('-',''))
            if corframe==3:
                translation=Seq(each[2:].replace('-','')).translate(table=gencode,to_stop=True).__str__()
                explen=len(each[2:].replace('-',''))
            if corframe==4:
                translation=Seq(each.replace('-','')).reverse_complement().translate(table=gencode,to_stop=True).__str__()
                explen=len(Seq(each.replace('-','')).__str__())
            if corframe==5:
                translation=Seq(each[:-1].replace('-','')).reverse_complement().translate(table=gencode,to_stop=True).__str__()
                explen=len(Seq(each[:-1].replace('-','')).__str__())
            if corframe==6:
                translation=Seq(each[:-2].replace('-','')).reverse_complement().translate(table=gencode,to_stop=True).__str__()
                explen=len(Seq(each[:-2].replace('-','')).__str__())

            if len(translation)<int(explen/3):
                return "0"
            elif len(translation)==int(explen/3):
                return "1"

        def get_cor_frame(seq,gencode):
            seqset=[Seq(seq),Seq(seq[1:]),Seq(seq[2:]),Seq(seq).reverse_complement(),Seq(seq[:-1]).reverse_complement(),Seq(seq[:-2]).reverse_complement()]
            aminoset=[]
            for each in seqset:
                a=each.translate(table=gencode,to_stop=True)
                aminoset.append(a.__str__())
            maxlen,corframe=0,0
            for i,each in enumerate(aminoset):
                if len(each)>maxlen:
                    maxlen=len(each)
                    corframe=i+1
            return corframe 
                                    
        def callconsensus(i,perc_thresh,abs_thresh,name):
            with open(i) as infile:
                l=infile.readlines()
                seqdict,poslist={},[]
                for i,j in enumerate(l):
                    if ">" in j:
                        poslist.append(i)
                                         
                                
                for i,j in enumerate(poslist):
                    ambcounts=0
                    k01=l[j].strip().split('>')[1]
                    if i!=len(poslist)-1:
                        k3=l[j+1:poslist[i+1]]
                    if i==len(poslist)-1:
                        k3=l[j+1:]
                    k4=''.join(k3).replace('\n','')
                    seqdict[k01]=k4.replace("E","A").replace("F","G").replace("Q","C").replace("P","T")
                conseq=consensus(seqdict,perc_thresh,abs_thresh)
                conseq=conseq.replace("-",'')
                flag=False
                if len(conseq)!=0:
                    transcheck=translate_corframe(conseq,ingencode)
                    coverage=len(seqdict.keys())
                    if len(conseq)==plen:
                        if conseq.count("N")==0:
                            if translate_corframe(conseq,ingencode)=="1":
                                flag=True
                                
                else:
                    transcheck="NA"
                    coverage="NA"
                return transcheck,conseq,flag,coverage

        transcheck={}
        conseqs={}
        flags={}
        coverages={}
        sampleids={}
        parstring=''
        with open("parfile") as parfile:
            l=parfile.readlines()
            parstring=l[0].strip()

        self.todolist=[x+"_all.fa" for x in self.todolist]
        for c,name in enumerate(self.todolist):
            if self.subsetmax==True:
                if self.sampleids[name.split("_all.fa")[0].split(".")[0]]<=maxcoverage:
                    indir2="Demultiplexed"
                else:
                    indir="Demultiplexed"
                    indir2="Subsets"
            else:
                indir2="Demultiplexed"
            try:
                if self.subsetmax==False:
                    cmd='disttbfast.exe '+parstring+' -i '+os.path.join(self.outpath,indir2,name)
                else:
                    if self.sampleids[name.split("_all.fa")[0].split(".")[0]]<=maxcoverage:
                        cmd='disttbfast.exe '+parstring+' -i '+os.path.join(self.outpath,indir2,name)
                    else:
                        sampleids[name.split("_all.fa")[0].split(".")[0]]=subset_bylength(os.path.join(self.outpath,indir,name),os.path.join(self.outpath,indir2,name),maxcoverage,plen)
                        cmd='disttbfast.exe '+parstring+' -i '+os.path.join(self.outpath,indir2,name)
                with open(os.devnull, 'w') as devnull:
                    stdout=subprocess.check_output(cmd,shell=True, stderr=devnull)
                with open(os.path.join(self.outpath,"Aligned",name + '_aln.fasta'), "wb") as handle:
                    handle.write(stdout)        
                transcheckeach,conseq,flag,cov=callconsensus(os.path.join(self.outpath,"Aligned",name+"_aln.fasta"),fixthresh,5,name)
                otherconseqs=[]
                if flag==True:
                    transcheck[name.split("_all.fa")[0].split(".")[0]]=transcheckeach
                    conseqs[name.split("_all.fa")[0].split(".")[0]]=conseq
                    flags[name.split("_all.fa")[0].split(".")[0]]=flag
                    if cov!="NA":
                        coverages[name.split("_all.fa")[0].split(".")[0]]=cov
                    self.ndone+=1
                    self.notifyMessage.emit(self.ndone)
                else:
                    n=rangefreq[1]
                    while n>=rangefreq[0]:
                        if n!=fixthresh:
                            transcheckeach2,conseq2,flag2,cov2=callconsensus(os.path.join(self.outpath,"Aligned",name+"_aln.fasta"),n,5,name)
                            if flag2==True:
                                if conseq2 not in otherconseqs:
                                    otherconseqs.append(conseq2)
                        n-=stepsize
                if len(otherconseqs)==1:
                    transcheck[name.split("_all.fa")[0].split(".")[0]]="1"
                    conseqs[name.split("_all.fa")[0].split(".")[0]]=otherconseqs[0]
                    flags[name.split("_all.fa")[0].split(".")[0]]=True
                    if cov!="NA":
                        coverages[name.split("_all.fa")[0].split(".")[0]]=cov
                        coverages[name.split("_all.fa")[0].split(".")[0]]=cov
                    self.ndone+=1
                    self.notifyMessage.emit(self.ndone)
                else:
                    transcheck[name.split("_all.fa")[0].split(".")[0]]=transcheckeach
                    conseqs[name.split("_all.fa")[0].split(".")[0]]=conseq
                    flags[name.split("_all.fa")[0].split(".")[0]]=flag
                    if cov!="NA":
                        coverages[name.split("_all.fa")[0].split(".")[0]]=cov           
            except Application.ApplicationError:
                transcheck[name.split("_all.fa")[0].split(".")[0]]="NA"
                conseqs[name.split("_all.fa")[0].split(".")[0]]=""
                flags[name.split("_all.fa")[0].split(".")[0]]=False
                if cov!='NA':
                    coverages[name.split("_all.fa")[0].split(".")[0]]="NA"
        self.result=[transcheck,conseqs,flags,coverages]
        self.donelist=[]
        with open(os.path.join(self.outpath,"consensus_good.fa"),'a') as outfile:

            for each in self.result[2].keys():
                if flags[each]==True:
                    outfile.write(">"+each+";"+str(coverages[each])+'\n'+conseqs[each]+'\n')
                    self.donelist.append(each)
                    with open(os.path.join(self.outpath,"Consensus",each),'w') as o:
                        o.write(">"+each+";"+str(coverages[each])+'\n'+conseqs[each]+'\n')
                else:
                    with open(os.path.join(self.outpath,"Consensus",each),'w') as o:
                        o.write(">"+each+";"+str(coverages[each])+'\n'+conseqs[each]+'\n')
        with open(os.path.join(self.outpath,"consensus_all.fa"),'w') as outfile:
            for filename in os.listdir(os.path.join(self.outpath,"Consensus")):
                with open(os.path.join(self.outpath,"Consensus",filename)) as readfile:
                    shutil.copyfileobj(readfile,outfile)
                
        self.taskFinished.emit(self.donelist)
    #    print "running consensus complete",self.ndone


class rundemultiplexlive(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
    notifyProgress=QtCore.pyqtSignal(int)
    notifyMessage=QtCore.pyqtSignal(int)
    notifyMessage2=QtCore.pyqtSignal(int,int)
    def __init__(self,livedir,newfiles,outpath,tagdict,muttags_fr,sampledict,typedict,primerfset,primerrset,taglen,primerlensum,samplecounts,totalseqs,ndemultiplexed,ndemultiplexedused,minlen,explen,demlen,primersearchlen,tagmm,compression,primermismatch,parent=None):
        super(rundemultiplexlive, self).__init__(parent)
        self.livedir=livedir
        self.newfiles=newfiles
        self.outpath=outpath
        self.tagdict=tagdict
        self.muttags_fr=muttags_fr
        self.sampledict=sampledict
        self.typedict=typedict
        self.primerfset=primerfset
        self.primerrset=primerrset
        self.taglen=taglen
        self.primerlensum=primerlensum
        self.samplecounts=samplecounts
        self.totalseqs=totalseqs
        self.ndemultiplexed=ndemultiplexed
        self.ndemultiplexedused=ndemultiplexedused
        self.minlen=minlen
        self.explen=explen
        self.demlen=demlen
        self.primersearchlen=primersearchlen
        self.tagmm=tagmm
        self.compression=compression
        self.primermismatch=primermismatch                                  
    def run(self):

        minlen=self.minlen
        explen=self.explen
        demlen=self.demlen
        primersearchlen=self.primersearchlen
    #   self.totalseqs=0
        self.nseqspasslen=0
        self.nseqsfordemultiplexing=0
        self.nseqsfordemultiplexingpresplit1=0
        self.nseqsfordemultiplexingpresplit2=0
        self.nwronglengthwindow=0
  #      print "Totalseqs before dem",self.totalseqs
  #      print "Demultiplexed before dem",self.ndemultiplexed
        def findmatch_m2(taglist,muttags_fr,indict,seqdict,sampledict,typedict,samplecounts,filename):
            
        #   print taglist
        #   print "sample",sampledict
            dmpfile=open(os.path.join(self.outpath,"dmpfile"),'a')
            idcombs={}
            cumscores={}
            for each in indict.keys():
                try:
                    idcombs[each]=(taglist[indict[each][0]], taglist[indict[each][1]])
                    cumscore=typedict[indict[each][0]]+typedict[indict[each][1]]
                    cumscores[each]=[str(typedict[indict[each][0]]),str(typedict[indict[each][1]]),str(cumscore)]
                except KeyError:
                    pass
            
            matched_idcombs={}
            for each in idcombs.keys():
                try:                  
                    sampledict[idcombs[each]]
                    self.ndemultiplexed+=1

                    with open(os.path.join(self.outpath,"Demultiplexed",sampledict[idcombs[each]]+"_all.fa"),'a') as outfile:
                        outfile.write(">"+filename+"_"+each+" _"+"_".join(cumscores[each])+"\n"+seqdict[each]+"\n")
                        self.samplecounts[sampledict[idcombs[each]]]+=1
                        self.ndemultiplexedused+=1                        
                except KeyError:
                    pass
                        
            dmpfile.close() 
        def readprimertagfasta(filef,filer):
            indict2,indict1,indict={},{},{}
            with open(filef) as infile1:
                with open(filer) as infile2:
                    l1=infile1.readlines()
                    l2=infile2.readlines()
                    for i,j in enumerate(l1):
                        if ">" in j:
                            indict1[j.strip().replace(">","")]=l1[i+1].strip()
                            indict[j.strip().replace(">","")]=[0,0]
                    for i,j in enumerate(l2):
                        if ">" in j:
                            indict2[j.strip().replace(">","")]=l2[i+1].strip()
                            indict[j.strip().replace(">","")]=[0,0]
            return indict1,indict2,indict
        def builddict_sequences(infile):
            seqdict={}
            with open(infile) as inseqs:
                l=inseqs.readlines()
                for i,j in enumerate(l):
                    if ">" in j:
                        seqdict[j.strip().replace(">","")]=l[i+1].strip()
            return seqdict
        def modseq(seq):
            for bp in ["A","T","G","C"]:
                n=20
                while n>2:
                    seq=seq.replace(bp*n,bp*2)
                    n-=1
            return seq
        def demultiplexfunc(pf,pr,seqdict,tagdict,muttags_fr,sampledict,typedict,num_row,filename):
            indict1,indict2,indict=readprimertagfasta(pf,pr)

            for each in list(indict.keys()):
                try: 
                    indict[each][0]=indict1[each]
                    indict[each][1]=indict2[each]
                except KeyError:
                    del indict[each]
                    
            try:
                with open(os.path.join(self.outpath,"dmpfile_tagfr"),'w') as tagfrdmp:  
                    for tag in muttags_fr.keys():
                        tagfrdmp.write(str(tag)+'\t'+muttags_fr[tag]+'\n')
                findmatch_m2(tagdict, muttags_fr,indict,seqdict,sampledict,typedict,num_row,filename)
            except UnboundLocalError:
                findmatch_m2(tagdict, {},indict,seqdict,sampledict,typedict,num_row,filename)
        def cleantagfile(infile):
            with open(infile) as primertagfile:
                with open(infile+"cleaned",'w') as primertagfile_cleaned:
                    taglines=primertagfile.readlines()
                    ids=[]
                    seqs={}
                    for n,line in enumerate(taglines):
                        if ">" in line:
                            ids.append(line)
                            seqs[line]=modseq(taglines[n+1])[-(taglen+1):]

                    idcounts=Counter(ids)
                    nseqs=0
                    for id in ids:
                        if idcounts[id]==1:
                            nseqs+=1
                            primertagfile_cleaned.write(id+seqs[id])
        for newfile in self.newfiles:
            if self.compression==0:
            #    print "uncompressed"
                with open(os.path.join(self.livedir,newfile)) as infile:
                    with open(os.path.join(self.outpath,"Processing",newfile+"_reformat_out"),'w') as outfile:
                        n=1
                        for line1,line2,line3,line4 in itertools.zip_longest(*[infile]*4):
                            if line1[0]=="@":
                                seqid=">"+line1[1:]
                                sequence=line2.strip()
                                if len(sequence)>minlen:
                                    outfile.write(seqid+line2)
                                    self.nseqspasslen+=1
                                n+=1
                                self.totalseqs+=1
            elif self.compression==1:
            #    print "compressed"
                with gzip.open(os.path.join(self.livedir,newfile),'rb') as infile:
                    with open(os.path.join(self.outpath,"Processing",newfile+"_reformat_out"),'w') as outfile:
                        n=1
                        for line1,line2,line3,line4 in itertools.zip_longest(*[infile]*4):
                            if line1[0]=="@":
                                seqid=">"+line1[1:].strip()
                                sequence=line2.strip()
                                if len(sequence)>minlen:
                                    outfile.write(seqid+'\n'+sequence+'\n')
                                    self.nseqspasslen+=1
                                n+=1
                                self.totalseqs+=1                
            self.notifyMessage.emit(self.totalseqs)
            with open(os.path.join(self.outpath,"Processing",newfile+"_reformat_out_1pdt"),'w') as file1:
                with open(os.path.join(self.outpath,"Processing",newfile+"_reformat_out_2pdt"),'w') as file2:
                    with open(os.path.join(self.outpath,"Processing",newfile+"_reformat_out")) as infile:
                        for line1,line2 in itertools.zip_longest(*[infile]*2):
                            if len(line1.strip())!=0:
                                seqid=line1
                                sequence=line2.strip()
                                if len(sequence)<explen+self.taglen*2+self.primerlensum+demlen:
                                    file1.write(seqid+line2)
                                    self.nseqsfordemultiplexing+=1
                                    self.nseqsfordemultiplexingpresplit1+=1
                                elif len(sequence)>(explen+self.taglen*2+self.primerlensum)*2 - demlen and len(sequence)<(explen+self.taglen*2+self.primerlensum+demlen)*2 :
                                    file2.write(seqid.strip()+" p1\n"+sequence[:explen+self.taglen*2+self.primerlensum+demlen]+'\n')
                                    file2.write(seqid.strip()+" p2\n"+sequence[(explen+self.taglen*2+self.primerlensum-demlen):]+'\n')
                                    self.nseqsfordemultiplexing+=2
                                    self.nseqsfordemultiplexingpresplit2+=1
                                else:
                                                      
                                    self.nwronglengthwindow+=1
            typedict2={newfile+"_reformat_out_1pdt":primersearchlen,newfile+"_reformat_out_2pdt":primersearchlen*2}
            c=0
            nseqs=0
            
            for infile2 in [newfile+"_reformat_out_1pdt",newfile+"_reformat_out_2pdt"]:
                ambiguity_codes=[("R", "A"), ("R", "G"),("M", "A"), ("M", "C"),("S", "C"), ("S", "G"),("Y", "C"), ("Y", "T"),("K", "G"), ("K", "T"),("W", "A"), ("W", "T"),("V", "A"), ("V", "C"),("V", "G"),("H", "A"),("H", "C"),("H", "T"),("D", "A"), ("D", "G"),("D", "T"), ("B", "C"),("B", "G"), ("B", "T"),("N", "A"), ("N", "G"), ("N", "C"), ("N", "T")]
                inputseqs=builddict_sequences(os.path.join(self.outpath,"Processing",infile2))
                seqlens=[len(i.strip()) for i in inputseqs.values()]

                with open(os.path.join(self.outpath,"Processing",infile2+"_all_glsearch1.parsed.lencutoff5parsed_f"),'w') as tagfoutfile:
                    with open(os.path.join(self.outpath,"Processing",infile2+"_all_glsearchR.parsed.lencutoff5parsed_r"),'w') as tagroutfile:
                        with open(os.path.join(self.outpath,"Processing",infile2+"_all_glsearchR.parsed.lencutoff5endr"),'w') as fprimercleanfile:
                            for pf in self.primerfset:
                                for pr in self.primerrset:
                                    pr=seqpy.revcomp(pr)
                                    for n,inseq in enumerate(inputseqs.keys()):
                                        c+=1
                                        if n%1000==0:
                                            progress = float(c) 
                                        compseq=seqpy.revcomp(inputseqs[inseq])
                                        k1 = edlib.align(pf, inputseqs[inseq][:typedict2[infile2]], mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                        k2 = edlib.align(pf, compseq[:typedict2[infile2]], mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                        d1 = k1['editDistance']
                                        d2 = k2['editDistance']
                                        orient=0
                                        revseq=''
                                        if d1<=d2:
                                            if d1<=self.primermismatch:
                                                loc=k1['locations'][0]
                                                tagseq=inputseqs[inseq][loc[0]-self.taglen:loc[0]]
                                                if len(tagseq)!=0:
                                                  
                                                #   print "pf",inseq,d1,loc,len(rev_comp(compseq[loc[1]+1:]))
                                                    tagfoutfile.write(">"+inseq+'\n'+tagseq+'\n')
                                                    revseq=inputseqs[inseq][loc[1]+1:]
                                        else:
                                            if d2<=self.primermismatch:
                                                loc=k2['locations'][0]
                                                
                                                tagseq=compseq[loc[0]-self.taglen:loc[0]]
                                                if len(tagseq)!=0:
                                                #   print "pfR",inseq,d2,loc,len(rev_comp(compseq[loc[1]+1:]))
                                                    tagfoutfile.write(">"+inseq+'\n'+tagseq+'\n')
                                                    revseq=compseq[loc[1]+1:]
                                                    orient=1
                                        
                                        if revseq!='':
                                            k = edlib.align(pr, revseq[-typedict2[infile2]:], mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                            d = k['editDistance']
                                            
                                            if d<=self.primermismatch:
                                                loc=k['locations'][0]
                                                
                                                startpoint=len(revseq)-typedict2[infile2]
                                                
                                                tagseq=seqpy.revcomp(revseq[startpoint+loc[1]+1:startpoint+loc[1]+self.taglen+1])
                                                if len(tagseq)!=0:
                                                    tagroutfile.write(">"+inseq+'\n'+tagseq+'\n')
                                                    if orient==0:
                                                        fprimercleanfile.write(">"+inseq+'\n'+revseq[:startpoint+loc[0]]+'\n')
                                                    else:
                                                        fprimercleanfile.write(">"+inseq+'\n'+revseq[:startpoint+loc[0]].replace("A","E").replace("G","F").replace("C","Q").replace("T","P")+'\n')
                                        
            #cleantagfile(outpath+"/"+infile+"_all_glsearch1.parsed.lencutoff5parsed_f")
            #cleantagfile(outpath+"/"+ infile+"_all_glsearchR.parsed.lencutoff5parsed_r")
                inputseqs=builddict_sequences(os.path.join(self.outpath,"Processing",infile2+"_all_glsearchR.parsed.lencutoff5endr"))
                demultiplexfunc(os.path.join(self.outpath,"Processing",infile2+"_all_glsearch1.parsed.lencutoff5parsed_f"),os.path.join(self.outpath,"Processing",infile2+"_all_glsearchR.parsed.lencutoff5parsed_r"), inputseqs,self.tagdict,self.muttags_fr,self.sampledict,self.typedict,0,infile2)
              #  self.ndemultiplexed+=subdemultiplexedn
            #    print infile2,self.ndemultiplexedused,self.ndemultiplexed
            self.notifyMessage2.emit(self.ndemultiplexedused,self.ndemultiplexed)
        self.taskFinished.emit(0)


class prepdemultiplexlive(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
    notifyProgress=QtCore.pyqtSignal(int)
    notifyMessage=QtCore.pyqtSignal(str)
    def __init__(self,demfile,logfile,tagmm,parent=None):
        super(prepdemultiplexlive, self).__init__(parent)
        self.demfile=demfile
        self.logfile=logfile
        self.tagmm=tagmm
    def run(self):

        def crmutant_m2(tfile,nbp):
            tagdict={}
            sampledict={}
            with open(tfile,'rb') as tagfile:
                t=tagfile.readlines()
                for each in t:
                    each=each.decode("utf-8-sig")
                    tagdict[each.split(',')[1].upper()]=''
                    tagdict[each.split(',')[2].upper()]=''
            counter=1
            typedict={}
            for each in tagdict.keys():
                tagdict[each]="t"+str(counter)
                typedict[each]=0
                counter+=1
            with open(tfile,'rb') as tagfile:
                t=tagfile.readlines()
                for each in t:
                    each=each.decode("utf-8-sig")
                    sampledict[(tagdict[each.split(',')[1].upper()],tagdict[each.split(',')[2].upper()])]=each.split(',')[0].replace(" ","_")
            n=1
            while n<=nbp:
                muttags_fr,newtags_fr=create_all_mutants(tagdict)
        #       newtags_fr_counts=Counter(newtags_fr)   

                with open("conflicts",'w') as conflictfile:
                    for k in newtags_fr.keys():
                        if len(newtags_fr[k])>1:
                            conflictfile.write(k+'\t'+",".join(newtags_fr[k])+'\n')
                            del muttags_fr[k]
                ntagset=list(set(muttags_fr.keys())-set(tagdict.keys()))
        #       print len(tagdict),len(muttags_fr),len(ntagset)
                for k in ntagset:
                    tagdict[k]=muttags_fr[k]
                    typedict[k]=n
                n+=1
        #   with open(self.outdir+"/temp1.fas",'w') as outfile:
        #       for k in tagdict.keys():
        #           outfile.write(k+'\t'+tagdict[k]+'\n')
            if nbp>0:
                return tagdict,muttags_fr,sampledict,typedict
            else:
                return tagdict,tagdict,sampledict,typedict

        def create_all_mutants(tagdict):
            muttags,newtags={},{}
            for tag in tagdict.keys():
                submutant=createsubmutant(tag)
                delmutant=createdelmutant(tag)
                insmutant=createinsmutant(tag)
                mutantset=list(set(submutant)|set(delmutant)|set(insmutant))
                mutantset[:]=[x for x in mutantset if x != tag]
                for mutant in mutantset:
                    try:
                        if tagdict[tag] not in newtags[mutant]:
                            newtags[mutant].append(tagdict[tag])
                    except:
                        newtags[mutant]=[tagdict[tag]]
                    muttags[mutant]= tagdict[tag]
            return muttags,newtags
        def createsubmutant(tag):
            newlist=[]
            bpset=["A","T","G","C"]
            for i,v in enumerate(tag):
                for bp in [x for x in bpset if x != v]:
                    if i==0:
                        newlist.append(bp+tag[i+1:])
                    if i>0 and i<len(tag)-1:
                        newlist.append(tag[0:i]+bp+tag[i+1:])
                    if i==len(tag)-1:
                        newlist.append(tag[0:i]+bp)
            return newlist

        def createdelmutant(tag):
            newlist=[]
            bpset=["A","T","G","C"]
            for i,v in enumerate(tag):
                for bp in bpset:
                    if i>0 and i<len(tag)-1:
                        newlist.append(bp+tag[0:i]+tag[i+1:])
                    if i==len(tag)-1:
                        newlist.append(bp+tag[0:i])
            return newlist
            
        def createinsmutant(tag):
            newlist=[]
            bpset=["A","T","G","C"]
            for i,v in enumerate(tag):
                for bp in bpset:
                    if i>0 and i<len(tag)-1:
                        newlist.append(tag[1:i]+bp+tag[i:])
                    if i==len(tag)-1:
                        newlist.append(tag[1:]+bp)
            return newlist

        self.sampleids={}
        self.primerlensum=0
        with open(self.demfile,'rb') as demulfile:
            demullines=demulfile.readlines()
            demullines=[x.decode("utf-8-sig") for x in demullines]
            for each in demullines:
                self.sampleids[each.split(',')[0].replace(" ","_")]=0
            primerf=demullines[0].split(',')[3].upper()
            primerr=demullines[0].split(',')[4].strip().upper()
            self.taglen=len(demullines[0].split(',')[1])
            self.primerfset=[primerf]
            self.primerlensum+=len(primerf)+len(primerr)
            self.primerrset=[primerr]
        self.logfile.write("<br><br>Read demultiplexing file. There are "+str("{:,}".format(len(self.sampleids))) + " in your experiment.\n")

        self.tagdict,self.muttags_fr,self.sampledict,self.typedict=crmutant_m2(self.demfile,self.tagmm)
        self.taskFinished.emit(0)

class prepdemultiplex(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
    notifyProgress=QtCore.pyqtSignal(int)
    notifyMessage=QtCore.pyqtSignal(str)
    def __init__(self,demfile,infastq,outdir,minlen,explen,demlen,logfile,tagmm,parent=None):
        super(prepdemultiplex, self).__init__(parent)
        self.demfile=demfile
        self.infastq=infastq
        self.outdir=outdir
        self.minlen=minlen
        self.explen=explen
        self.logfile=logfile
        self.demlen=demlen
        self.tagmm=tagmm
    def run(self):

        def crmutant_m2(tfile,nbp):
            tagdict={}
            sampledict={}
            with open(tfile,'rb') as tagfile:
                t=tagfile.readlines()
                for each in t:
                    each=each.decode("utf-8-sig")
                    tagdict[each.split(',')[1].upper()]=''
                    tagdict[each.split(',')[2].upper()]=''
            counter=1
            typedict={}
            for each in tagdict.keys():
                tagdict[each]="t"+str(counter)
                typedict[each]=0
                counter+=1
            with open(tfile,'rb') as tagfile:
                t=tagfile.readlines()
                for each in t:
                    each=each.decode("utf-8-sig")
                    sampledict[(tagdict[each.split(',')[1].upper()],tagdict[each.split(',')[2].upper()])]=each.split(',')[0].replace(" ","_")
            n=1
            while n<=nbp:
                muttags_fr,newtags_fr=create_all_mutants(tagdict)
        #       newtags_fr_counts=Counter(newtags_fr)   

                with open("conflicts",'w') as conflictfile:
                    for k in newtags_fr.keys():
                        if len(newtags_fr[k])>1:
                            conflictfile.write(k+'\t'+",".join(newtags_fr[k])+'\n')
                            del muttags_fr[k]
                ntagset=list(set(muttags_fr.keys())-set(tagdict.keys()))
        #       print len(tagdict),len(muttags_fr),len(ntagset)
                for k in ntagset:
                    tagdict[k]=muttags_fr[k]
                    typedict[k]=n
                n+=1
            with open(self.outdir+"/temp1.fas",'w') as outfile:
                for k in tagdict.keys():
                    outfile.write(k+'\t'+tagdict[k]+'\n')
            if nbp>0:
                return tagdict,muttags_fr,sampledict,typedict
            else:
                return tagdict,tagdict,sampledict,typedict

        def create_all_mutants(tagdict):
            muttags,newtags={},{}
            for tag in tagdict.keys():
                submutant=createsubmutant(tag)
                delmutant=createdelmutant(tag)
                insmutant=createinsmutant(tag)
                mutantset=list(set(submutant)|set(delmutant)|set(insmutant))
                mutantset[:]=[x for x in mutantset if x != tag]
                for mutant in mutantset:
                    try:
                        if tagdict[tag] not in newtags[mutant]:
                            newtags[mutant].append(tagdict[tag])
                    except:
                        newtags[mutant]=[tagdict[tag]]
                    muttags[mutant]= tagdict[tag]
            return muttags,newtags
        def createsubmutant(tag):
            newlist=[]
            bpset=["A","T","G","C"]
            for i,v in enumerate(tag):
                for bp in [x for x in bpset if x != v]:
                    if i==0:
                        newlist.append(bp+tag[i+1:])
                    if i>0 and i<len(tag)-1:
                        newlist.append(tag[0:i]+bp+tag[i+1:])
                    if i==len(tag)-1:
                        newlist.append(tag[0:i]+bp)
            return newlist

        def createdelmutant(tag):
            newlist=[]
            bpset=["A","T","G","C"]
            for i,v in enumerate(tag):
                for bp in bpset:
                    if i>0 and i<len(tag)-1:
                        newlist.append(bp+tag[0:i]+tag[i+1:])
                    if i==len(tag)-1:
                        newlist.append(bp+tag[0:i])
            return newlist
            
        def createinsmutant(tag):
            newlist=[]
            bpset=["A","T","G","C"]
            for i,v in enumerate(tag):
                for bp in bpset:
                    if i>0 and i<len(tag)-1:
                        newlist.append(tag[1:i]+bp+tag[i:])
                    if i==len(tag)-1:
                        newlist.append(tag[1:]+bp)
            return newlist
        basename=os.path.basename(self.infastq)

        self.sampleids={}
        primerlensum=0
        with open(self.demfile,'rb') as demulfile:
            demullines=demulfile.readlines()
            demullines=[x.decode("utf-8-sig") for x in demullines]

            for each in demullines:
                self.sampleids[each.split(',')[0].replace(" ","_")]=''
            primerf=demullines[0].split(',')[3].upper()
            primerr=demullines[0].split(',')[4].strip().upper()
            self.taglen=len(demullines[0].split(',')[1])
            self.primerfset=[primerf]
            primerlensum+=len(primerf)+len(primerr)
            self.primerrset=[primerr]
        self.nseqspasslen=0
        self.totalseqs=0
        self.logfile.write("<br><br>Read demultiplexing file. There are "+str("{:,}".format(len(self.sampleids))) + " in your experiment.\n")
        with open(self.infastq) as infile:
            with open(os.path.join(self.outdir,basename+"_reformat_out"),'w') as outfile:
                n=1
                for line1,line2,line3,line4 in itertools.zip_longest(*[infile]*4):
                    if len(line1.strip())!=0:
                        seqid=">"+line1[1:]
                        sequence=line2.strip()
                        if len(sequence)>self.minlen:
                            outfile.write(seqid+line2)
                            self.nseqspasslen+=1
                        n+=1
                        self.totalseqs+=1
        self.logfile.write(" Generated length filtered file. Your raw file contains "+str(n-1) + " sequences.\n")
        self.nseqsfordemultiplexingpresplit1=0
        self.nseqsfordemultiplexingpresplit2=0
        self.nwronglengthwindow=0
        self.nseqsfordemultiplexing=0       

        with open(os.path.join(self.outdir,basename+"_reformat_out_1pdt"),'w') as file1:
            with open(os.path.join(self.outdir,basename+"_reformat_out_2pdt"),'w') as file2:
                with open(os.path.join(self.outdir,basename+"_reformat_out")) as infile:
                    for line1,line2 in itertools.zip_longest(*[infile]*2):
                        if len(line1.strip())!=0:
                            seqid=line1
                            sequence=line2.strip()
                            if len(sequence)<self.explen+self.taglen*2+primerlensum+self.demlen:
                                file1.write(seqid+line2)
                                self.nseqsfordemultiplexing+=1
                                self.nseqsfordemultiplexingpresplit1+=1
                            elif len(sequence)>(self.explen+self.taglen*2+primerlensum)*2 - self.demlen and len(sequence)<(self.explen+self.taglen*2+primerlensum+self.demlen)*2 :
                                file2.write(seqid.strip()+"p1\n"+sequence[:self.explen+self.taglen*2+primerlensum+self.demlen]+'\n')
                                file2.write(seqid.strip()+"p2\n"+sequence[(self.explen+self.taglen*2+primerlensum-self.demlen):]+'\n')
                                self.nseqsfordemultiplexing+=2
                                self.nseqsfordemultiplexingpresplit2+=1
                            else:
                                self.nwronglengthwindow+=1
        self.notifyMessage.emit("<br>Your raw file contains "+str("{:,}".format(self.totalseqs)) + " reads.<br>You have "+str("{:,}".format(self.nseqspasslen))+" reads passing the length filter. <br> Of these, "+str("{:,}".format(self.nseqsfordemultiplexingpresplit1)) +" reads are of the correct length for containing one barcode, while " + str("{:,}".format(self.nseqsfordemultiplexingpresplit2)) +" may contain two ligated  barcodes<br>Total reads of expected length ="+str("{:,}".format(self.nseqsfordemultiplexingpresplit1+self.nseqsfordemultiplexingpresplit2)) +"<br>After splitting the long products, a total of "+str("{:,}".format(self.nseqsfordemultiplexing)) +" reads are used for demultiplexing.")
        self.logfile.write(str(round(time.time())) +": Split files into correct product length. There are "+str("{:,}".format(self.nseqspasslen)) + " single product sequences.\n")
        self.lastbitn=self.nseqspasslen%40000
#       print self.lastbitn
        self.tagdict,self.muttags_fr,self.sampledict,self.typedict=crmutant_m2(self.demfile,self.tagmm)
        def grouper(n, iterable, fillvalue=None):
            args = [iter(iterable)] * n
            return zip_longest(fillvalue=fillvalue, *args)

        n = 40000
        self.maxid=0

        with open(os.path.join(self.outdir,basename+"_reformat_out_1pdt")) as f:
            for i, g in enumerate(grouper(n, f, fillvalue=''), 1):
                with open(os.path.join(self.outdir,basename+"_reformat_out_1pdt_p"+str(i*n)), 'w') as fout:
                    fout.writelines(g)
                    self.maxid=basename+"_reformat_out_1pdt_p"+str(i*n)
        with open(os.path.join(self.outdir,basename+"_reformat_out_2pdt")) as f:
            for i, g in enumerate(grouper(n, f, fillvalue=''), 1):
                with open(os.path.join(self.outdir,basename+"_reformat_out_2pdt_p"+str(i*n)), 'w') as fout:
                    fout.writelines(g)
        self.taskFinished.emit(0)


class mergedemfiles(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
    notifyProgress=QtCore.pyqtSignal(int)
    def __init__(self,sampleids,indir,outdir,parent=None):
        super(mergedemfiles, self).__init__(parent)
        self.sampleids=sampleids
        self.indir=indir
        self.outdir=outdir
    def run(self):
        def mergefiles(inlist, outfilename):
            o = open(outfilename, 'w')
                   
            for each in inlist:
                for line in fileinput.input([each]):
                    o.write(line)
                                   
                                
                fileinput.close()
            o.close()
                        
        c=1
        for each in self.sampleids.keys():
    #       print each
            n=0
            eachlist=[]
            while n<4:
                dirlist=os.listdir(os.path.join(self.indir,str(n)))
                if each+"_all.fa" in dirlist:
                    eachlist.append(os.path.join(self.indir,str(n),each+"_all.fa"))
                n+=1
            if len(eachlist)>0:
                mergefiles(eachlist,os.path.join(self.outdir,each+"_all.fa"))
                                          
            self.notifyProgress.emit(c)
            c+=1
        shutil.rmtree(self.indir)
        self.taskFinished.emit(0)


class calculatecoverage(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
                                         
    def __init__(self,indir,parent=None):
        super(calculatecoverage, self).__init__(parent)
        self.indir=indir
    def run(self):
        dirlist=os.listdir(self.indir)
        self.dirdict={}
        self.counter={}
        for i,fname in enumerate(dirlist):
            with open(os.path.join(self.indir,fname)) as infile:
                if fname.endswith("_all.fa")==True:
                    self.dirdict[fname]=fname
                else:
                    self.dirdict[fname.split(".")[0]+"_all.fa"]=fname
                l=infile.readlines()
                for i,j in enumerate(l):
                    if ">" in j:
                        try:
                            self.counter[fname.split("_all.fa")[0]]+=1
                        except KeyError:
                            self.counter[fname.split("_all.fa")[0]]=1
                                         
        self.taskFinished.emit(0)

class runpaircomparisons(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(list)
    notifyProgress=QtCore.pyqtSignal(int)
    def __init__(self,inseqdict,inseqnamedict,outdir,filelist,reffilename,parent=None):
        super(runpaircomparisons, self).__init__(parent)
        self.inseqdict=inseqdict
        self.inseqnamedict=inseqnamedict
        self.outdir=outdir
        self.filelist=filelist
        self.reffilename=reffilename
    def run(self):
        ambiguity_codes=[("R", "A"), ("R", "G"),("M", "A"), ("M", "C"),("S", "C"), ("S", "G"),("Y", "C"), ("Y", "T"),("K", "G"), ("K", "T"),("W", "A"), ("W", "T"),("V", "A"), ("V", "C"),("V", "G"),("H", "A"),("H", "C"),("H", "T"),("D", "A"), ("D", "G"),("D", "T"), ("B", "C"),("B", "G"), ("B", "T"),("N", "A"), ("N", "G"), ("N", "C"), ("N", "T")]
    #   print os.path.join(self.outdir,"identicalbarcodes_658chosen.fa")
        filenamepairs=[]
        self.filelist.remove(self.reffilename)
        for each in self.filelist:
            if each!=self.reffilename:
                filenamepairs.append([self.reffilename,each])
        os.mkdir(os.path.join(self.outdir,"pairwise_comparisons"))
        wb=xlsxwriter.Workbook(os.path.join(self.outdir,"statistics.xlsx"))
        allsheet=wb.add_worksheet("Pairwise Multisample comparison")
        goodsheet=wb.add_worksheet("Correct")
        errsheet=wb.add_worksheet("Erroneous")
        multisamplegood,multisamplebad=0,0
        multisampleincompatible={}
        counterdict,sheetdict,pstatdict={},{},{}
        for each in filenamepairs:
            each1=each[0].split(".")[0]
            each2=each[1].split(".")[0]
            os.mkdir(os.path.join(self.outdir,"pairwise_comparisons",each2+"_comps"))
            os.mkdir(os.path.join(self.outdir,"pairwise_comparisons",each2+"_comps","badbarcodes"))
            sheetdict[each2]=wb.add_worksheet(each2[:31])
        uniqlist={} 
        filenamedict={}
        for i,each in enumerate(self.filelist):
            filenamedict[each]=i
            uniqlist[each.split(".")[0]]=0
            counterdict[each.split(".")[0]]=[0,0,0,0,0,0,0,0]
        inseqdict=self.inseqdict
        inseqnamedict=self.inseqnamedict
        goodbarcodefile=open(os.path.join(self.outdir,"identical_compatible_barcodes.fa"),'w')
        onlyfilelist=[]
        os.mkdir(os.path.join(self.outdir,"badbarcodes"))
        counter=0
        refcounter=0
        inlistorder=[]
        for k in inseqdict.keys():
            subfilelist=inseqdict[k].keys()
            if len(subfilelist)>1:
                inlistorder=subfilelist
                break
        identicalcounts,compatiblecounts,incompatiblecounts=0,0,0
        for k in inseqdict.keys():
            identicalpairs=[]
            compatiblepairs=[]
            incompatiblepairs=[]
            incompatiblecounter=[]
            incompatiblecounter1=[]
            incompatiblecounter2=[]
            incompatiblecounter3=[]
            incompatibleseqs=[]
            subfilelist=list(inseqdict[k].keys())
            try:
                subfilelist.remove(self.reffilename)
                for i,fname in enumerate(self.filelist):
                    if fname in subfilelist:
                        incompatiblecounter.append(0)
                        incompatiblecounter1.append(0)
                        incompatiblecounter2.append(0)
                        incompatiblecounter3.append(0)
                    else:
                        incompatiblecounter.append("NA")
                        incompatiblecounter1.append(0)
                        incompatiblecounter2.append(0)
                        incompatiblecounter3.append(0)
                if len(subfilelist)>0:
                    for p,mk in enumerate(inseqdict[k].keys()):
                        if mk==self.reffilename:
                            refcounter+=1
                            for q,id in enumerate(subfilelist):
                                if id!=self.reffilename:
                                    counterdict[id.split(".")[0]][3]+=1
                                    counterdict[id.split(".")[0]][4]+=1
                                    seq1=inseqdict[k][mk]
                                    seq2=inseqdict[k][id]
                                    if len(seq1)==len(seq2):
                                        noambgcodealn= edlib.align(seq1, seq2, mode='NW',task='path')
                                        ambgcodealn= edlib.align(seq1, seq2, mode='NW',task='path',additionalEqualities=ambiguity_codes)
                                    if len(seq1)<len(seq2):
                                        ambgcodealn= edlib.align(seq1, seq2, mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                        noambgcodealn= edlib.align(seq1, seq2, mode='HW',task='path')
                                    if len(seq1)>len(seq2):
                                        ambgcodealn= edlib.align(seq2, seq1, mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                        noambgcodealn= edlib.align(seq2, seq1, mode='HW',task='path')
                                    d1 = noambgcodealn['editDistance']
                                    d2 = ambgcodealn['editDistance']
                                    if d2==0:
                                        if d1>d2:
                                            compatiblepairs.append([mk,id])
                                            compatiblecounts+=1
                                            each1=mk.split(".")[0]
                                            each2=id.split(".")[0]
                                            with open(os.path.join(self.outdir,"pairwise_comparisons",each2+"_comps","compatiblebarcodes.fa"),'a') as pcompfile:
                                                pcompfile.write(self.inseqnamedict[k][id]+self.inseqdict[k][id]+'\n')
                                            counterdict[each2][1]+=1                                    
                                        elif d1==d2:
                                            identicalpairs.append([mk,id])
                                            identicalcounts+=1
                                            each1=mk.split(".")[0]
                                            each2=id.split(".")[0]
                                            with open(os.path.join(self.outdir,"pairwise_comparisons",each2+"_comps","goodbarcodes.fa"),'a') as pgoodfile:
                                                pgoodfile.write(self.inseqnamedict[k][id]+self.inseqdict[k][id]+'\n')
                                            counterdict[each2][0]+=1
                                    else:
                                        compseq1=seqpy.revcomp(seq1)
                                        if len(seq1)==len(seq2):
                                            noambgcodealn= edlib.align(compseq1, seq2, mode='NW',task='path')
                                            ambgcodealn= edlib.align(compseq1, seq2, mode='NW',task='path',additionalEqualities=ambiguity_codes)
                                        if len(seq1)<len(seq2):
                                            ambgcodealn= edlib.align(compseq1, seq2, mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                            noambgcodealn= edlib.align(compseq1, seq2, mode='HW',task='path')
                                        if len(seq1)>len(seq2):
                                            ambgcodealn= edlib.align(seq2, compseq1, mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                            noambgcodealn= edlib.align(seq2, compseq1, mode='HW',task='path')   
                                        cd1 = noambgcodealn['editDistance']
                                        cd2 = ambgcodealn['editDistance']
                                        if cd2==0:
                                            if cd1>cd2:
                                                compatiblepairs.append([mk,id])
                                                compatiblecounts+=1
                                                each1=mk.split(".")[0]
                                                each2=id.split(".")[0]
                                                with open(os.path.join(self.outdir,"pairwise_comparisons",each2+"_comps","compatiblebarcodes.fa"),'a') as pcompfile:
                                                    pcompfile.write(self.inseqnamedict[k][id]+self.inseqdict[k][id]+'\n')
                                                counterdict[each2][1]+=1
                                            elif cd1==cd2:
                                                identicalpairs.append([mk,id])
                                                identicalcounts+=1
                                                each1=mk.split(".")[0]
                                                each2=id.split(".")[0]
                                                with open(os.path.join(self.outdir,"pairwise_comparisons",each2+"_comps","goodbarcodes.fa"),'a') as pgoodfile:
                                                    pgoodfile.write(self.inseqnamedict[k][id]+self.inseqdict[k][id]+'\n')
                                                counterdict[each2][0]+=1
                                        else:
                                            incompatiblepairs.append([mk,id])
                                            incompatiblecounts+=1
                                            if cd2>d2:
                                                incompatibleseqs.append([seq1,seq2])
                                                each1=mk.split(".")[0]
                                                each2=id.split(".")[0]                                      
                                        #       incompatiblecounter[filenamedict[mk]]+=1
                                                incompatiblecounter[filenamedict[id]]=d2                            
                                                with open(os.path.join(self.outdir,"pairwise_comparisons",each2+"_comps","badbarcodes",str(d2)+"_"+k+".fa"),'a') as incompatiblefile:
                                                    incompatiblefile.write(self.inseqnamedict[k][mk]+self.inseqdict[k][mk]+'\n'+self.inseqnamedict[k][id]+self.inseqdict[k][id]+'\n')
                                                try:
                                                    pstatdict[each2].append([k,d2])
                                                except KeyError:
                                                    pstatdict[each2]=[[k,d2]]
                                                counterdict[each2][2]+=1
                                                if d2==1:
                                                    counterdict[each2][5]+=1
                                                elif d2==2:
                                                    counterdict[each2][6]+=1
                                                elif d2>2:
                                                    counterdict[each2][7]+=1
                                            else:
                                                incompatibleseqs.append([compseq1,seq2])
                                                incompatiblecounter[filenamedict[id]]=cd2
                                                each1=mk.split(".")[0]
                                                each2=id.split(".")[0]
                                                with open(os.path.join(self.outdir,"pairwise_comparisons",each2+"_comps","badbarcodes",str(cd2)+"_"+k+".fa"),'a') as incompatiblefile:
                                                    incompatiblefile.write(self.inseqnamedict[k][mk]+self.inseqdict[k][mk]+'\n'+self.inseqnamedict[k][id]+self.inseqdict[k][id]+'\n')
                                                try:
                                                    pstatdict[each2].append([k,cd2])
                                                except KeyError:
                                                    pstatdict[each2]=[[k,cd2]]
                                                counterdict[each2][2]+=1
                                                if cd2==1:
                                                    counterdict[each2][5]+=1
                                                elif cd2==2:
                                                    counterdict[each2][6]+=1
                                                elif cd2>2:
                                                    counterdict[each2][7]+=1
                    fuseidenticalscompatible=self.fusesets(identicalpairs+compatiblepairs)
                    if len(incompatiblepairs)==0:
                        if len(fuseidenticalscompatible)>0:
                            goodbarcodefile.write(self.inseqnamedict[k][fuseidenticalscompatible[0][0]]+self.inseqdict[k][fuseidenticalscompatible[0][0]]+'\n')
                            multisamplegood+=1
                    else:
                        with open(os.path.join(self.outdir,"badbarcodes",k+".fa"),'w') as badfile:
                            for inlistcount,inlist in enumerate(incompatiblepairs):
                                badfile.write(self.inseqnamedict[k][inlist[0]]+incompatibleseqs[inlistcount][0]+'\n'+self.inseqnamedict[k][inlist[1]]+incompatibleseqs[inlistcount][1]+'\n')
                        multisamplebad+=1
                    multisampleincompatible[k]=incompatiblecounter
                else:
                    for i,each in enumerate(subfilelist):
                        with open(os.path.join(self.outdir,subfilelist[i].split(".")[0]+"_onlybarcodes.fa"),'a') as uniquefile:
                            uniquefile.write(self.inseqnamedict[k][subfilelist[i]]+self.inseqdict[k][subfilelist[i]]+'\n')
                            uniqlist[subfilelist[i].split(".")[0]]+=1
                            counterdict[subfilelist[i].split(".")[0]][4]+=1
            except ValueError:
                    for i,each in enumerate(subfilelist):
                        with open(os.path.join(self.outdir,subfilelist[i].split(".")[0]+"_onlybarcodes.fa"),'a') as uniquefile:
                            uniquefile.write(self.inseqnamedict[k][subfilelist[i]]+self.inseqdict[k][subfilelist[i]]+'\n')
                            uniqlist[subfilelist[i].split(".")[0]]+=1
                            counterdict[subfilelist[i].split(".")[0]][4]+=1
            self.notifyProgress.emit(counter+1)
            counter+=1
        for each in sheetdict.keys():
            sheetdict[each].write(0,0,"Overview")
            sheetdict[each].write(1,0,"Number of specimens with identical barcodes")
            sheetdict[each].write(1,1,counterdict[each.replace(self.reffilename.split(".")[0]+"_","")][0])
            sheetdict[each].write(2,0,"Number of specimens with compatible barcodes")
            sheetdict[each].write(2,1,counterdict[each.replace(self.reffilename.split(".")[0]+"_","")][1])
            sheetdict[each].write(3,0,"Number of specimens with incompatible barcodes")
            sheetdict[each].write(3,1,counterdict[each.replace(self.reffilename.split(".")[0]+"_","")][2])
            sheetdict[each].write(5,0,"Sample")
            sheetdict[each].write(5,1,"Distance")
            r=6
            if each in pstatdict.keys():
                for row in pstatdict[each]:
        #           print row
                    sheetdict[each].write_url(r,0,os.path.join(self.outdir,"pairwise_comparisons",each,"badbarcodes",str(row[1])+"_"+row[0]+".fa"),string=row[0])
                    sheetdict[each].write(r,1,row[1])
                    r+=1
        allsheet.write(0,0,"Overview")
        allsheet.write(1,0,"Number of samples in the reference")
        allsheet.write(1,1,refcounter)
        allsheet.write(5,0,"Total Number of Barcodes in the dataset")
        allsheet.write(6,0,"Number of barcodes compared to the reference")
        allsheet.write(7,0,"Number of identical barcodes")
        allsheet.write(8,0,"Number of compatible barcodes")
        allsheet.write(9,0,"Number of incompatible barcodes")
        allsheet.write(10,0,"Number of incompatible barcodes with 1 error")
        allsheet.write(11,0,"Number of incompatible barcodes with 2 errors")
        allsheet.write(12,0,"Number of incompatible barcodes with >2 errors")
        goodsheet.write(1,0,"Total Number of Barcodes in the dataset")
        goodsheet.write(2,0,"Number of barcodes compared to the reference")
        goodsheet.write(3,0,"Number of identical barcodes")
        goodsheet.write(4,0,"Number of compatible barcodes")
        errsheet.write(1,0,"Total Number of Barcodes in the dataset")
        errsheet.write(2,0,"Number of barcodes compared to the reference")
        errsheet.write(3,0,"Number of incompatible barcodes")
        errsheet.write(4,0,"Number of incompatible barcodes with 1 error")
        errsheet.write(5,0,"Number of incompatible barcodes with 2 errors")
        errsheet.write(6,0,"Number of incompatible barcodes with >2 errors")


        c=1
        for fname in self.filelist:
            allsheet.write(4,c,fname)
            each=fname.split(".")[0]
            allsheet.write(5,c,counterdict[each][4])
            allsheet.write(6,c,counterdict[each][3])
            allsheet.write(7,c,counterdict[each][0])
            allsheet.write(8,c,counterdict[each][1])
            allsheet.write(9,c,counterdict[each][2])
            allsheet.write(10,c,counterdict[each][5])
            allsheet.write(11,c,counterdict[each][6])
            allsheet.write(12,c,counterdict[each][7])
            goodsheet.write(0,c,fname)
            goodsheet.write(1,c,counterdict[each][4])
            goodsheet.write(2,c,counterdict[each][3])
            goodsheet.write(3,c,counterdict[each][0])
            goodsheet.write(4,c,counterdict[each][1])
            errsheet.write(0,c,fname)
            errsheet.write(1,c,counterdict[each][4])
            errsheet.write(2,c,counterdict[each][3])
            errsheet.write(3,c,counterdict[each][2])
            errsheet.write(4,c,counterdict[each][5])
            errsheet.write(5,c,counterdict[each][6])
            errsheet.write(6,c,counterdict[each][7])
            
            c+=1
        r=5
        
        for row in multisampleincompatible.keys():
            templist=filter(lambda a: a != "NA", multisampleincompatible[row])
            totaldistance=sum(templist)
            if totaldistance==0:
                goodsheet.write(r,0,row)
                
                c=1
                for v in multisampleincompatible[row]:
                    goodsheet.write(r,c,v)
                    c+=1
                r+=1
        r=7
        for row in multisampleincompatible.keys():
            templist=filter(lambda a: a != "NA", multisampleincompatible[row])
            totaldistance=sum(templist)
            if totaldistance!=0:
                errsheet.write_url(r,0,os.path.join(self.outdir,"badbarcodes",row+'.fa'),string=row)
                
                c=1
                for v in multisampleincompatible[row]:
                    errsheet.write(r,c,v)
                    c+=1
                r+=1
        wb.close()
        self.taskFinished.emit([refcounter,counterdict])
    def fusesets(self,inlist):
        flag=0
        while flag==0:
            new=[]
            flag=1
            while inlist:
                cur,comp=inlist[0],inlist[1:]
                inlist=[]
                for q in comp:
                    if len(list(set(cur)&set(q)))>0:
                        flag=0
                        cur=tuple(set(cur)|set(q))
                    else:
                        inlist.append(q)
                new.append(cur)
            inlist=new
        return new
    
class runtwocomparisons(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(list)
    notifyProgress=QtCore.pyqtSignal(int)
    def __init__(self,inseqdict,inseqnamedict,outdir,parent=None):
        super(runtwocomparisons, self).__init__(parent)
        self.inseqdict=inseqdict
        self.inseqnamedict=inseqnamedict
        self.outdir=outdir
    def run(self):
        ambiguity_codes=[("R", "A"), ("R", "G"),("M", "A"), ("M", "C"),("S", "C"), ("S", "G"),("Y", "C"), ("Y", "T"),("K", "G"), ("K", "T"),("W", "A"), ("W", "T"),("V", "A"), ("V", "C"),("V", "G"),("H", "A"),("H", "C"),("H", "T"),("D", "A"), ("D", "G"),("D", "T"), ("B", "C"),("B", "G"), ("B", "T"),("N", "A"), ("N", "G"), ("N", "C"), ("N", "T")]
    #    print os.path.join(self.outdir,"identicalbarcodes_658chosen.fa")
        inseqdict=self.inseqdict
        inseqnamedict=self.inseqnamedict
        goodbarcodefile=open(os.path.join(self.outdir,"identicalbarcodes_658chosen.fa"),'w')
        compatiblebarcodefile=open(os.path.join(self.outdir,"compatiblebarcodes_658chosen_or_lowern.fa"),'w')
        statsfile=open(os.path.join(self.outdir,"statsfile.csv"),'w')
        onlyfilelist=[]
        os.mkdir(os.path.join(self.outdir,"badbarcodes"))
        counter=0
        inlistorder=[]
        for k in inseqdict.keys():
            subfilelist=inseqdict[k].keys()
            if len(subfilelist)>1:
                inlistorder=subfilelist
                break
        uniqlist={}
        statslines=''
        identicalcounts,compatiblecounts,incompatiblecounts=0,0,0
        for k in inseqdict.keys():
            identicalpairs=[]
            compatiblepairs=[]
            incompatiblepairs=[]
            incompatiblealns=[]
            incompatibleseqs=[]
            incompatibledists=[]
            subfilelist=list(inseqdict[k].keys())
            if len(subfilelist)>1:
                for p,mk in enumerate(inseqdict[k].keys()):
                    if p!=len(subfilelist)-1:
                        for q,id in enumerate(subfilelist[p+1:]):
                            seq1=inseqdict[k][mk]
                            seq2=inseqdict[k][id]

                            if len(seq1)==len(seq2):
                                noambgcodealn= edlib.align(seq1, seq2, mode='NW',task='path')
                                ambgcodealn= edlib.align(seq1, seq2, mode='NW',task='path',additionalEqualities=ambiguity_codes)
                            #   nice = edlib.getNiceAlignment(ambgcodealn, seq1, seq2)
                            if len(seq1)<len(seq2):
                                ambgcodealn= edlib.align(seq1, seq2, mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                noambgcodealn= edlib.align(seq1, seq2, mode='HW',task='path')
                                
                            if len(seq1)>len(seq2):
                                ambgcodealn= edlib.align(seq2, seq1, mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                noambgcodealn= edlib.align(seq2, seq1, mode='HW',task='path')
                            d1 = noambgcodealn['editDistance']
                            d2 = ambgcodealn['editDistance']
                            if d2==0:
                                if d1>d2:
                                    compatiblepairs.append([mk,id])
                                    compatiblecounts+=1
                                    compatiblebarcodefile.write(self.inseqnamedict[k][mk]+self.inseqdict[k][mk]+'\n')
                                elif d1==d2:
                                    identicalpairs.append([mk,id])
                                    identicalcounts+=1
                                    goodbarcodefile.write(self.inseqnamedict[k][mk]+self.inseqdict[k][mk]+'\n')
                            else:
                                compseq1=seqpy.revcomp(seq1)
                                if len(seq1)==len(seq2):
                                    noambgcodealn= edlib.align(compseq1, seq2, mode='NW',task='path')
                                    ambgcodealn= edlib.align(compseq1, seq2, mode='NW',task='path',additionalEqualities=ambiguity_codes)

                                if len(seq1)<len(seq2):
                                    ambgcodealn= edlib.align(compseq1, seq2, mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                    noambgcodealn= edlib.align(compseq1, seq2, mode='HW',task='path')

                                if len(seq1)>len(seq2):
                                    ambgcodealn= edlib.align(seq2, compseq1, mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                    noambgcodealn= edlib.align(seq2, compseq1, mode='HW',task='path')   
                                cd1 = noambgcodealn['editDistance']
                                cd2 = ambgcodealn['editDistance']
                                if cd2==0:
                                    if cd1>cd2:
                                        compatiblepairs.append([mk,id])
                                        compatiblecounts+=1
                                        compatiblebarcodefile.write(self.inseqnamedict[k][mk]+self.inseqdict[k][mk]+'\n')
                                    elif cd1==cd2:
                                        identicalpairs.append([mk,id])
                                        identicalcounts+=1
                                        goodbarcodefile.write(self.inseqnamedict[k][mk]+self.inseqdict[k][mk]+'\n')
                                else:
                                    
                                    incompatiblepairs.append([mk,id])
                                    incompatiblecounts+=1
                                    if cd2>d2:
                                        string=k+','+str(d2)+'\n'
                                        statslines+=string
                                    #   incompatibleseqs.append([seq1,seq2])
                                    #   incompatibledists.append(str(d2))
                                        with open(os.path.join(self.outdir,"badbarcodes",str(d2)+"_"+k+".fa"),'w') as badfile:
                                            badfile.write(self.inseqnamedict[k][mk]+seq1+'\n'+self.inseqnamedict[k][id]+seq2+'\n')
                                    else:
                                        string=k+','+str(cd2)+'\n'
                                        statslines+=string
                                #       incompatibleseqs.append([compseq1,seq2])
                                #       incompatibledists.append(str(cd2))
                                        with open(os.path.join(self.outdir,"badbarcodes",str(cd2)+"_"+k+".fa"),'w') as badfile:
                                            badfile.write(self.inseqnamedict[k][mk]+compseq1+'\n'+self.inseqnamedict[k][id]+seq2+'\n')
            #   if len(incompatiblepairs)==1:
            #       with open(os.path.join(self.outdir,"badbarcodes",incompatibledists[0]+"_"+k+".fa"),'w') as badfile:
            #           badfile.write(self.inseqnamedict[k][incompatibleseqs[0][0]]+incompatibleseqs[0][0]+'\n'+self.inseqnamedict[k][incompatibleseqs[0][1]]+incompatibleseqs[0][1]+'\n')
            else:
                with open(os.path.join(self.outdir,subfilelist[0].split(".")[0]+"_onlybarcodes.fa"),'a') as uniquefile:
                    uniquefile.write(self.inseqnamedict[k][subfilelist[0]]+self.inseqdict[k][subfilelist[0]]+'\n')
                    try:
                        uniqlist[subfilelist[0].split(".")[0]]+=1
                    except KeyError:
                        uniqlist[subfilelist[0].split(".")[0]]=1
                        
            
            self.notifyProgress.emit(counter+1)
            counter+=1
        statsfile.write("Overview\nNumber of specimens with identical barcodes,"+str(identicalcounts)+"\n")
        statsfile.write("Number of specimens with compatible barcodes,"+str(compatiblecounts)+"\n")
        statsfile.write("Number of specimens with incompatible barcodes,"+str(incompatiblecounts)+"\n")
        strdictlens=""
        for each in uniqlist.keys():
            strdictlens+="Number of barcodes in "+ each+" only,"+str(uniqlist[each])+"\n"

        statsfile.write("\n\nOverview of incompatible barcodes\nSampleID,Distance\n")
        statsfile.write(statslines)
        self.taskFinished.emit([identicalcounts,compatiblecounts,incompatiblecounts,uniqlist])
class runmulticomparisons(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(list)
    notifyProgress=QtCore.pyqtSignal(int)
    def __init__(self,inseqdict,inseqnamedict,outdir,filelist,parent=None):
        super(runmulticomparisons, self).__init__(parent)
        self.inseqdict=inseqdict
        self.inseqnamedict=inseqnamedict
        self.outdir=outdir
        self.filelist=filelist
    def run(self):
        ambiguity_codes=[("R", "A"), ("R", "G"),("M", "A"), ("M", "C"),("S", "C"), ("S", "G"),("Y", "C"), ("Y", "T"),("K", "G"), ("K", "T"),("W", "A"), ("W", "T"),("V", "A"), ("V", "C"),("V", "G"),("H", "A"),("H", "C"),("H", "T"),("D", "A"), ("D", "G"),("D", "T"), ("B", "C"),("B", "G"), ("B", "T"),("N", "A"), ("N", "G"), ("N", "C"), ("N", "T")]
    #   print os.path.join(self.outdir,"identicalbarcodes_658chosen.fa")
        filenamepairs=list(itertools.combinations(self.filelist, 2))
    #   print filenamepairs
        os.mkdir(os.path.join(self.outdir,"pairwise_comparisons"))
        wb=xlsxwriter.Workbook(os.path.join(self.outdir,"statistics.xlsx"))
        allsheet=wb.add_worksheet("Multisample comparison")
        multisamplegood,multisamplebad=0,0
        multisampleincompatible={}
        counterdict,sheetdict,pstatdict={},{},{}
        for each in filenamepairs:
            each1=each[0].split(".")[0]
            each2=each[1].split(".")[0]
            os.mkdir(os.path.join(self.outdir,"pairwise_comparisons",each1+"_"+each2+"_comps"))
            os.mkdir(os.path.join(self.outdir,"pairwise_comparisons",each1+"_"+each2+"_comps","badbarcodes"))
            counterdict[each1+"_"+each2]=[0,0,0,0,0,0,0,0]
            sheetdict[each1+"_"+each2]=wb.add_worksheet(each1[:15]+"_"+each2[:15])
        uniqlist={} 
        filenamedict={}
        for i,each in enumerate(self.filelist):
            filenamedict[each]=i
            uniqlist[each.split(".")[0]]=0

        
        inseqdict=self.inseqdict
        inseqnamedict=self.inseqnamedict
        goodbarcodefile=open(os.path.join(self.outdir,"identical_compatible_barcodes.fa"),'w')
        onlyfilelist=[]
        os.mkdir(os.path.join(self.outdir,"badbarcodes"))
        counter=0
              
        inlistorder=[]
        for k in inseqdict.keys():
            subfilelist=inseqdict[k].keys()
            if len(subfilelist)>1:
                inlistorder=subfilelist
                break
        
  
        identicalcounts,compatiblecounts,incompatiblecounts=0,0,0
        for k in inseqdict.keys():
            combodone=[]
            identicalpairs=[]
            compatiblepairs=[]
            incompatiblepairs=[]
            incompatiblecounter=[]
            incompatiblecounter1=[]
            incompatiblecounter2=[]
            incompatiblecounter3=[]
            incompatibleseqs=[]
            subfilelist=list(inseqdict[k].keys())     
            try:
                for i,fname in enumerate(self.filelist):
                    if fname in subfilelist:
                        incompatiblecounter.append(0)
                        incompatiblecounter1.append(0)
                        incompatiblecounter2.append(0)
                        incompatiblecounter3.append(0)
                    else:
                        incompatiblecounter.append("NA")
                        incompatiblecounter1.append(0)
                        incompatiblecounter2.append(0)
                        incompatiblecounter3.append(0)
                if len(subfilelist)>0:
                    for p,mk in enumerate(inseqdict[k].keys()):
                        subfilelist.remove(mk)
                        #refcounter+=1
                        for q,id in enumerate(subfilelist):
                            
                            if mk!=id:
                                if id.split(".")[0]+"_"+mk.split(".")[0] in counterdict.keys():
                                    comboname=id.split(".")[0]+"_"+mk.split(".")[0]
                                elif mk.split(".")[0]+"_"+id.split(".")[0] in counterdict.keys():
                                    comboname=mk.split(".")[0]+"_"+id.split(".")[0]
                                
                                counterdict[comboname][3]+=1
                                counterdict[comboname][4]+=1
                                seq1=inseqdict[k][mk]
                                seq2=inseqdict[k][id]
                                if len(seq1)==len(seq2):
                                    noambgcodealn= edlib.align(seq1, seq2, mode='NW',task='path')
                                    ambgcodealn= edlib.align(seq1, seq2, mode='NW',task='path',additionalEqualities=ambiguity_codes)
                                if len(seq1)<len(seq2):
                                    ambgcodealn= edlib.align(seq1, seq2, mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                    noambgcodealn= edlib.align(seq1, seq2, mode='HW',task='path')
                                if len(seq1)>len(seq2):
                                    ambgcodealn= edlib.align(seq2, seq1, mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                    noambgcodealn= edlib.align(seq2, seq1, mode='HW',task='path')
                                d1 = noambgcodealn['editDistance']
                                d2 = ambgcodealn['editDistance']
                                if d2==0:
                                    if d1>d2:
                                        compatiblepairs.append([mk,id])
                                        compatiblecounts+=1
                                        each1=mk.split(".")[0]
                                        each2=id.split(".")[0]
                                        with open(os.path.join(self.outdir,"pairwise_comparisons",comboname+"_comps","compatiblebarcodes.fa"),'a') as pcompfile:
                                            pcompfile.write(self.inseqnamedict[k][id]+self.inseqdict[k][id]+'\n')
                                        counterdict[comboname][1]+=1                                    
                                    elif d1==d2:
                                        identicalpairs.append([mk,id])
                                        identicalcounts+=1
                                        each1=mk.split(".")[0]
                                        each2=id.split(".")[0]
                                        with open(os.path.join(self.outdir,"pairwise_comparisons",comboname+"_comps","goodbarcodes.fa"),'a') as pgoodfile:
                                            pgoodfile.write(self.inseqnamedict[k][id]+self.inseqdict[k][id]+'\n')
                                        counterdict[comboname][0]+=1
                                else:
                                    compseq1=seqpy.revcomp(seq1)
                                    if len(seq1)==len(seq2):
                                        noambgcodealn= edlib.align(compseq1, seq2, mode='NW',task='path')
                                        ambgcodealn= edlib.align(compseq1, seq2, mode='NW',task='path',additionalEqualities=ambiguity_codes)
                                    if len(seq1)<len(seq2):
                                        ambgcodealn= edlib.align(compseq1, seq2, mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                        noambgcodealn= edlib.align(compseq1, seq2, mode='HW',task='path')
                                    if len(seq1)>len(seq2):
                                        ambgcodealn= edlib.align(seq2, compseq1, mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                        noambgcodealn= edlib.align(seq2, compseq1, mode='HW',task='path')   
                                    cd1 = noambgcodealn['editDistance']
                                    cd2 = ambgcodealn['editDistance']
                                    if cd2==0:
                                        if cd1>cd2:
                                            compatiblepairs.append([mk,id])
                                            compatiblecounts+=1
                                            each1=mk.split(".")[0]
                                            each2=id.split(".")[0]

                                            with open(os.path.join(self.outdir,"pairwise_comparisons",comboname+"_comps","compatiblebarcodes.fa"),'a') as pcompfile:
                                                pcompfile.write(self.inseqnamedict[k][id]+self.inseqdict[k][id]+'\n')
                                            
                                            counterdict[comboname][1]+=1
                                        elif cd1==cd2:
                                            identicalpairs.append([mk,id])
                                            identicalcounts+=1
                                            each1=mk.split(".")[0]
                                            each2=id.split(".")[0]
                                            with open(os.path.join(self.outdir,"pairwise_comparisons",comboname+"_comps","goodbarcodes.fa"),'a') as pgoodfile:
                                                pgoodfile.write(self.inseqnamedict[k][id]+self.inseqdict[k][id]+'\n')
                                            
                                            counterdict[comboname][0]+=1
                                    else:
                                        incompatiblepairs.append([mk,id])
                                        incompatiblecounts+=1
                                        if cd2>d2:
                                            incompatibleseqs.append([seq1,seq2])
                                            each1=mk.split(".")[0]
                                            each2=id.split(".")[0]                                      
                                            incompatiblecounter[filenamedict[mk]]+=1
                                            incompatiblecounter[filenamedict[id]]+=1                            
                                            with open(os.path.join(self.outdir,"pairwise_comparisons",comboname+"_comps","badbarcodes",str(d2)+"_"+k+".fa"),'a') as incompatiblefile:
                                                incompatiblefile.write(self.inseqnamedict[k][mk]+self.inseqdict[k][mk]+'\n'+self.inseqnamedict[k][id]+self.inseqdict[k][id]+'\n')
                                            try:
                                                pstatdict[comboname].append([k,d2])
                                            except KeyError:
                                                pstatdict[comboname]=[[k,d2]]
                                            counterdict[comboname][2]+=1
                                            if d2==1:
                                                counterdict[comboname][5]+=1
                                            elif d2==2:
                                                counterdict[comboname][6]+=1
                                            elif d2>2:
                                                counterdict[comboname][7]+=1
                                        else:
                                            incompatibleseqs.append([compseq1,seq2])
                                            incompatiblecounter[filenamedict[mk]]+=1
                                            incompatiblecounter[filenamedict[id]]+=1
                                            each1=mk.split(".")[0]
                                            each2=id.split(".")[0]
                                            with open(os.path.join(self.outdir,"pairwise_comparisons",comboname+"_comps","badbarcodes",str(cd2)+"_"+k+".fa"),'a') as incompatiblefile:
                                                incompatiblefile.write(self.inseqnamedict[k][mk]+self.inseqdict[k][mk]+'\n'+self.inseqnamedict[k][id]+self.inseqdict[k][id]+'\n')
                                            try:
                                                pstatdict[comboname].append([k,cd2])
                                            except KeyError:
                                                pstatdict[comboname]=[[k,cd2]]
                                            counterdict[comboname][2]+=1
                                            if cd2==1:
                                                counterdict[comboname][5]+=1
                                            elif cd2==2:
                                                counterdict[comboname][6]+=1
                                            elif cd2>2:
                                                counterdict[comboname][7]+=1
                                
                                    
                    fuseidenticalscompatible=self.fusesets(identicalpairs+compatiblepairs)
                    if len(incompatiblepairs)==0:
                        if len(fuseidenticalscompatible)>0:
                            goodbarcodefile.write(self.inseqnamedict[k][fuseidenticalscompatible[0][0]]+self.inseqdict[k][fuseidenticalscompatible[0][0]]+'\n')
                            multisamplegood+=1
                    else:
                        with open(os.path.join(self.outdir,"badbarcodes",k+".fa"),'w') as badfile:
                            for inlistcount,inlist in enumerate(incompatiblepairs):
                                badfile.write(self.inseqnamedict[k][inlist[0]]+incompatibleseqs[inlistcount][0]+'\n'+self.inseqnamedict[k][inlist[1]]+incompatibleseqs[inlistcount][1]+'\n')
                        multisamplebad+=1
                    multisampleincompatible[k]=incompatiblecounter
                else:
                    for i,each in enumerate(subfilelist):
                        with open(os.path.join(self.outdir,subfilelist[i].split(".")[0]+"_onlybarcodes.fa"),'a') as uniquefile:
                                                             
                            uniquefile.write(self.inseqnamedict[k][subfilelist[i]]+self.inseqdict[k][subfilelist[i]]+'\n')
                            uniqlist[subfilelist[i].split(".")[0]]+=1
                        #   counterdict[subfilelist[i].split(".")[0]][4]+=1
            except ValueError:
                    for i,each in enumerate(subfilelist):
                
                        with open(os.path.join(self.outdir,subfilelist[i].split(".")[0]+"_onlybarcodes.fa"),'a') as uniquefile:
                            uniquefile.write(self.inseqnamedict[k][subfilelist[i]]+self.inseqdict[k][subfilelist[i]]+'\n')

                            uniqlist[subfilelist[i].split(".")[0]]+=1
                        #   counterdict[subfilelist[i].split(".")[0]][4]+=1
            self.notifyProgress.emit(counter+1)
            counter+=1
        for each in sheetdict.keys():
            sheetdict[each].write(0,0,"Overview")
            sheetdict[each].write(1,0,"Number of specimens with identical barcodes")
            sheetdict[each].write(1,1,counterdict[each][0])
            sheetdict[each].write(2,0,"Number of specimens with compatible barcodes")
            sheetdict[each].write(2,1,counterdict[each][1])
            sheetdict[each].write(3,0,"Number of specimens with incompatible barcodes")
            sheetdict[each].write(3,1,counterdict[each][2])
            sheetdict[each].write(5,0,"Sample")
            sheetdict[each].write(5,1,"Distance")
            r=6
            try:    
                for row in pstatdict[each]:
                    sheetdict[each].write(r,0,row[0])
                    sheetdict[each].write(r,1,row[1])
                    r+=1
            except KeyError:
                pass        
        allsheet.write(0,0,"Overview")
        allsheet.write(1,0,"Number of samples with completely compatible/identical barcodes across all sets")
        allsheet.write(1,1,multisamplegood)
        allsheet.write(2,0,"Number of samples with incompatible barcodes in at least one set")
        allsheet.write(2,1,multisamplebad)
        c=1
        for fname in counterdict:
            allsheet.write(4,c,fname)
            each=fname.split(".")[0]
            allsheet.write(5,c,counterdict[each][4])
            allsheet.write(6,c,counterdict[each][3])
            allsheet.write(7,c,counterdict[each][0])
            allsheet.write(8,c,counterdict[each][1])
            allsheet.write(9,c,counterdict[each][2])
            allsheet.write(10,c,counterdict[each][5])
            allsheet.write(11,c,counterdict[each][6])
            allsheet.write(12,c,counterdict[each][7])
            c+=1
        r=5
        for row in multisampleincompatible.keys():
            allsheet.write(r,0,row)
            c=1
            for v in multisampleincompatible[row]:
                allsheet.write(r,c,v)
                c+=1
            r+=1
        wb.close()
        self.taskFinished.emit([multisamplegood,multisamplebad,uniqlist])
    def fusesets(self,inlist):
        flag=0
        while flag==0:
            new=[]
            flag=1
            while inlist:
                cur,comp=inlist[0],inlist[1:]
                inlist=[]
                for q in comp:
                    if len(list(set(cur)&set(q)))>0:
                        flag=0
                        cur=tuple(set(cur)|set(q))
                    else:
                        inlist.append(q)
                new.append(cur)
            inlist=new
        return new


class DragPushbtn(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(DragPushbtn, self).__init__(parent)
        self.mk1bpushbtn=QtWidgets.QPushButton("Drag the directory where fastq/fast5 files are generated")
        self.mk1bpushbtn.setEnabled(False)
        self.setAcceptDrops(True)
        layout=QtWidgets.QGridLayout()
        
        layout.addWidget(self.mk1bpushbtn,0,0)
        self.setLayout(layout)
    def dragEnterEvent(self, e):
     #   print e

        if e.mimeData().hasText():
            e.accept()
        else:
            e.ignore()
            
    def dropEvent(self, e):
        urls = e.mimeData().urls()        
        if len(urls)==1:
            fname=str(urls[0].path())
        #   print fname
            if "/" == str(urls[0].path())[0]:
                fname=str(urls[0].path())[1:]
            if os.path.isdir(fname)==True:
                self.livedir=fname
                self.mk1bpushbtn.setEnabled(True)
                self.mk1bpushbtn.setText("Click to Proceed")
class copyfiles(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
    def __init__(self,indir,outdir,inlist,indir2,dirdict,inputmode,parent=None):
        super(copyfiles, self).__init__(parent)
        self.indir=indir
        self.outdir=outdir
        self.inlists=inlist
        self.dirdict=dirdict
        self.indir2=indir2
        self.inputmode=inputmode
    def run(self):
        indir2=self.indir2
        for i,inlist in enumerate(self.inlists):
            if len(inlist)>0:
                for fname in inlist:
                    if self.inputmode==1:
                        shutil.copyfile(os.path.join(self.indir,fname), os.path.join(self.outdir[i],fname))
                    else:
                        shutil.copyfile(os.path.join(self.indir,self.dirdict[fname]), os.path.join(self.outdir[i],self.dirdict[fname]))
            shutil.make_archive(self.outdir[i], 'zip', self.outdir[i])
            shutil.rmtree(self.outdir[i])
        shutil.make_archive(os.path.join(indir2,"2a_ConsensusByLength"), 'zip', os.path.join(indir2,"2a_ConsensusByLength"))
        shutil.make_archive(os.path.join(indir2,"2b_ConsensusBySimilarity"), 'zip', os.path.join(indir2,"2b_ConsensusBySimilarity"))
        shutil.make_archive(os.path.join(indir2,"3_ConsensusByBarcodeComparison"), 'zip', os.path.join(indir2,"3_ConsensusByBarcodeComparison"))
        shutil.rmtree(os.path.join(indir2,"2a_ConsensusByLength"))
        shutil.rmtree(os.path.join(indir2,"2b_ConsensusBySimilarity"))
        shutil.rmtree(os.path.join(indir2,"3_ConsensusByBarcodeComparison"))
        self.taskFinished.emit(1)
class PbarWindow(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super(PbarWindow, self).__init__(parent)
        self.layout=QtWidgets.QGridLayout()
        self.layout.setSpacing(20)
        self.setLayout(self.layout)
        self.logstatus = QtWidgets.QTextEdit()
        self.logstatus.setReadOnly(True)
        self.logstatus.setTextInteractionFlags(QtCore.Qt.NoTextInteraction) 
        self.pbarwidget=QtWidgets.QWidget()
        self.pbarlayout=QtWidgets.QGridLayout()
    #   self.pbarlabel=QtWidgets.QLabel('Initializing')
        self.pbarTitle=QtWidgets.QLabel('Setting up your run, this will take a short while')
        self.pbar=QtWidgets.QProgressBar()
        self.pbarTitle.setFont(QtGui.QFont("Times", 12, QtGui.QFont.Bold))
    #   self.pbarlabel.setStyleSheet("font-weight: bold; color: green")
        self.pbarlayout.addWidget(self.pbarTitle,0,0,1,5,alignment=QtCore.Qt.AlignCenter)
        self.pbarlayout.addWidget(self.pbar,1,0,1,5)
        self.pbarwidget.setLayout(self.pbarlayout)
        self.layout.addWidget(self.logstatus,0,0,1,5)
        self.layout.addWidget(self.pbarwidget,1,0,1,5)
class checkindir(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
    def __init__(self,indir,parent=None):
        super(checkindir, self).__init__(parent)
        self.indir=indir              
    def run(self):
        dirlist=os.listdir(self.indir)
        if len(dirlist)>0: #checks if directory is not empty
            for each in dirlist:
                if each.endswith("_all.fa"):#examines files that end with _all.fa
                    with open(os.path.join(self.indir,each)) as infile:
                        line=infile.readline()
                        if len(line)>0: #checks if length of first line is over 0
                            if line[0]==">": # checks if first character is ">"
                                self.indirstatus=True
                            else: #emits error if first character of file is not ">"
                                self.indirstatus=False
                                break
                        else: #emits error if first line is empty
                            self.indirstatus=False
                            break
                else: #emits error if directory contains a file without "_all.fa"
                    self.indirstatus=False
                    break
        else: #emits error if directory is empty
            self.indirstatus=False
        self.taskFinished.emit(1)

class SelectionTable(QtWidgets.QTableWidget):
    notifydoubleclick=QtCore.pyqtSignal(int)
    def __init__(self, parent=None):
        super(SelectionTable, self).__init__(parent)
        self.setColumnCount(1)
        self.setRowCount(3)
        self.setShowGrid(True)
        
        self.setSizePolicy(QtWidgets.QSizePolicy.Minimum,QtWidgets.QSizePolicy.Minimum)
        self.setMaximumWidth(self.size().width())
        self.setMaximumHeight(self.size().height())
    #   self.cellDoubleClicked.connect(self.emitsignal)
    #   self.setRowHeight(0,(self.size().height()*0.3))
    #   self.setRowHeight(1,(self.size().height()*0.3))
    #   self.setRowHeight(2,(self.size().height()*0.3))
        self.setColumnWidth(0,(int(self.size().width()*0.98)))
                                                             
        self.horizontalHeader().hide()
        self.verticalHeader().hide()
                      
        self.setFrameStyle(QtWidgets.QFrame.NoFrame)
                       
        item=QtWidgets.QTableWidgetItem("   MODE 1: Demultiplexing and/or barcode calling : Please drag an input FASTQ file and demultiplexing file to enable this mode   ")
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.setItem(0, 0, item)
        item=QtWidgets.QTableWidgetItem("   MODE 2: Barcode calling only: Please drag a folder containing demultiplexed fasta to enable this mode. Use naming conventions specified in readme if data was not demultiplexed by ONTbarcoder   ")
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.setItem(1, 0, item)
        item=QtWidgets.QTableWidgetItem("   MODE 3: Fix barcodes based on MSA: Please drag an input fasta of consensus previously called by this software   ")
        
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        
        self.setItem(2, 0, item)
        
        self.setRowHeight(0, 50)
        self.setRowHeight(1, 50)
        self.setRowHeight(2, 50)
        self.resizeRowsToContents()                          
        self.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)

class runtoptwenty(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
    notifyProgress=QtCore.pyqtSignal(int)
    def __init__(self,inlist,parent=None):
        super(runtoptwenty, self).__init__(parent)      
        self.indir=inlist
    def run(self):
        ambiguity_codes=[("R", "A"), ("R", "G"),("M", "A"), ("M", "C"),("S", "C"), ("S", "G"),("Y", "C"), ("Y", "T"),("K", "G"), ("K", "T"),("W", "A"), ("W", "T"),("V", "A"), ("V", "C"),("V", "G"),("H", "A"),("H", "C"),("H", "T"),("D", "A"), ("D", "G"),("D", "T"), ("B", "C"),("B", "G"), ("B", "T"),("N", "A"), ("N", "G"), ("N", "C"), ("N", "T")]
    #   print "running top 20"
        inlist=self.indir
        seqlist=inlist[0]
        seqdict=inlist[1]
        refseqdict=inlist[2]
        outpath=inlist[3]
        num_row=inlist[4]
        for i,each in enumerate(seqlist):
            dists={}
            for other in refseqdict.keys():
                if each!=other:
                    k = edlib.align(each, other, mode='NW',task='path',additionalEqualities=ambiguity_codes)
                    d = k['editDistance']
                    dists[other]=d
            

            sorted_d=sorted(dists.items(), key=lambda x: x[1])
            with open(os.path.join(outpath,seqdict[each][0].split(";")[0]),'w') as outfile:
                outfile.write(">"+seqdict[each][0]+'\n'+each+'\n')
                for other in sorted_d[:20]:
                    outfile.write(">"+refseqdict[other[0]][0]+";dist="+str(float(other[1])/float(len(other[0]))*100)+'\n'+other[0]+'\n')
            cmd='disttbfast.exe -q 0 -E 2 -V -1.53 -s 0.0 -W 6 -O -C 4-4 -b 62 -g 0 -f -1.53 -Q 100.0 -h 0 -F -X 0.1 -x 1000 -i ' +os.path.join(outpath,seqdict[each][0].split(";")[0])
        #   os.system('disttbfast.exe -q 0 -E 2 -V -1.53 -s 0.0 -W 6 -O -C 4-4 -b 62 -g 0 -f -1.53 -Q 100.0 -h 0 -F -X 0.1 -x 1000 -i ' +os.path.join(outpath,seqdict[each][0].split(";")[0])+" > "+os.path.join(outpath,seqdict[each][0].split(";")[0]+"_aln.fa"))
        #   mafft_cline=MuscleCommandline("muscle3.8.31_i86win32.exe",input=os.path.join(outpath,seqdict[each][0].split(";")[0]))
            with open(os.devnull, 'w') as devnull:
                stdout=subprocess.check_output(cmd,shell=True,stderr=devnull)
                                       
            with open(os.path.join(outpath,seqdict[each][0].split(";")[0]+"_aln.fa"), "wb") as handle:
                handle.write(stdout)
            self.notifyProgress.emit(i+1)
        self.taskFinished.emit(0)
class runconsensusparts(QtCore.QThread):
    taskFinished=QtCore.pyqtSignal(int)
    notifyProgress=QtCore.pyqtSignal(int)
    def __init__(self,inlist,parent=None):
        super(runconsensusparts, self).__init__(parent)
        self.inlist=inlist
    def run(self):      
        inlist=self.inlist
    #   print inlist
        inlist1=inlist[0]
        outpath=inlist[1]
        indir=inlist[2]
        indir2=inlist[3]
        subsetval=inlist[4]
        outname=inlist[5]
        num_row=inlist[6]
        plen=inlist[7]
        v=inlist[8]
        outpath2=inlist[10]
        
        postdemlen=int(inlist[9])
        fixthresh=float(inlist[11])
        rangefreq=inlist[12]
        stepsize=float(inlist[13])
        ingencode=int(inlist[14])
        def consensus(indict,perc_thresh,abs_thresh):
            if len(indict.keys())>=abs_thresh:
                poslist=[]
                n=0
                while n<len(list(indict.values())[0]):
                    newlist=[]
                    for i,each in enumerate(indict.keys()):
                        try:
                            newlist.append(indict[each][n])
                        except IndexError:
                #           print each,indict[each],indict.values()[0],len(indict[each]),len(indict.values()[0])
                            break
                    poslist.append(newlist)
                    n+=1
                sequence=[]
                countpos=1
                for character in poslist:
                    charcounter=Counter(character)
                    baseset={}
                    for k,v in charcounter.items():
                        percbp=float(v)/float(len(character))
                        if percbp>perc_thresh:
                            baseset[k]=v
                    if len(baseset)==0:
                        bp='N'
                    if len(baseset)==1:
                        bp=list(baseset.keys())[0]
                    if len(baseset)>1:
                        try:
                            del baseset["-"]
                            if len(baseset)==1:
                                bp=list(baseset.keys())[0]
                            else:
                                bp='N'
                        except KeyError:
                            bp="N"
                    sequence.append(bp)
                    countpos+=1
                return ''.join(sequence)
            else:
                return ''
        def translate_corframe(seq,gencode):
            translatedset=[]
            each=seq
            corframe=get_cor_frame(each.replace("-",""),gencode)
            if corframe==1:
                translation=Seq(each.replace('-','')).translate(table=gencode,to_stop=True).__str__()
                explen=len(each.replace('-',''))
            if corframe==2:
                translation=Seq(each[1:].replace('-','')).translate(table=gencode,to_stop=True).__str__()
                explen=len(each[1:].replace('-',''))
            if corframe==3:
                translation=Seq(each[2:].replace('-','')).translate(table=gencode,to_stop=True).__str__()
                explen=len(each[2:].replace('-',''))
            if corframe==4:
                translation=Seq(each.replace('-','')).reverse_complement().translate(table=gencode,to_stop=True).__str__()
                explen=len(Seq(each.replace('-','')).__str__())
            if corframe==5:
                translation=Seq(each[:-1].replace('-','')).reverse_complement().translate(table=gencode,to_stop=True).__str__()
                explen=len(Seq(each[:-1].replace('-','')).__str__())
            if corframe==6:
                translation=Seq(each[:-2].replace('-','')).reverse_complement().translate(table=gencode,to_stop=True).__str__()
                explen=len(Seq(each[:-2].replace('-','')).__str__())

            if len(translation)<int(explen/3):
                return "0"
            elif len(translation)==int(explen/3):
                return "1"

        def get_cor_frame(seq,gencode):
            seqset=[Seq(seq),Seq(seq[1:]),Seq(seq[2:]),Seq(seq).reverse_complement(),Seq(seq[:-1]).reverse_complement(),Seq(seq[:-2]).reverse_complement()]
            aminoset=[]
            for each in seqset:
                a=each.translate(table=gencode,to_stop=True)
                aminoset.append(a.__str__())
            maxlen,corframe=0,0
            for i,each in enumerate(aminoset):
                if len(each)>maxlen:
                    maxlen=len(each)
                    corframe=i+1
            return corframe 
                                    
        def callconsensus(i,perc_thresh,abs_thresh,name):
            with open(i) as infile:
                l=infile.readlines()
                seqdict,poslist={},[]
                for i,j in enumerate(l):
                    if ">" in j:
                        poslist.append(i)
                for i,j in enumerate(poslist):
                    ambcounts=0
                    k01=l[j].strip().split('>')[1]
                    if i!=len(poslist)-1:
                        k3=l[j+1:poslist[i+1]]
                    if i==len(poslist)-1:
                        k3=l[j+1:]
                    k4=''.join(k3).replace('\n','')
                    seqdict[k01]=k4.replace("E","A").replace("F","G").replace("Q","C").replace("P","T")
                conseq=consensus(seqdict,perc_thresh,abs_thresh)
                conseq=conseq.replace("-",'')
        #       print name,conseq
                flag=False
                if len(conseq)!=0:
                    transcheck=translate_corframe(conseq,ingencode)
                    coverage=len(seqdict.keys())
                    if len(conseq)==plen:
                        if conseq.count("N")==0:
                            if translate_corframe(conseq,ingencode)=="1":
                                flag=True         
                else:
                    transcheck="NA"
                    coverage="NA"
                return transcheck,conseq,flag,coverage
        def subset_randomly (infile,outfile,n):
            with open(infile) as fulldata:
                with open(outfile,'w') as subsetdata:
                    l=fulldata.readlines()
                    if len(l)/2<=n:
                        subsetdata.write("".join(l))
                    else:
                        subsetlist=[x*2 for x in random.sample(range(len(l)/2),n)]
                        for i in subsetlist:
                            subsetdata.write(l[i]+l[i+1])
        def subset_bylength (infile,outfile,n,plen,windowlen):
            samplesize=0
            with open(infile) as fulldata:
                with open(outfile,'w') as subsetdata:
                    l=fulldata.readlines()
                    tseqdict={}
                    tlendict={}
                    for i,j in enumerate(l):
                        if ">" in j:
                            samplesize+=1
                            if abs(plen-len(l[i+1].strip()))<=windowlen:
                                tlendict[j]=abs(plen-len(l[i+1].strip()))
                                tseqdict[j]=l[i+1]
                    d=sorted(tlendict.items(), key=lambda x: x[1])
                    v=0
                    ntosubset=min(len(tseqdict),n)
        #           print "ntosubset",ntosubset
                    while v<ntosubset:
                        subsetdata.write(d[v][0]+tseqdict[d[v][0]])
                        v+=1
            return samplesize

        self.transcheck={}
        self.conseqs={}
        self.flags={}
        self.coverages={}
        self.sampleids={}
        parstring=''
        with open("parfile") as parfile:
            l=parfile.readlines()
            parstring=l[0].strip()
    #   print parstring
                 
        for c,name in enumerate(inlist1):
    #       print name
            try:
                if subsetval!=0:
                    if v==0:
                        self.sampleids[name.split("_all.fa")[0].split(".")[0]]=subset_bylength(os.path.join(outpath,indir,name),os.path.join(outpath,indir2,name),subsetval,plen,postdemlen)
                        cmd='disttbfast.exe '+parstring+' -i '+os.path.join(outpath,indir2,name) 
                    if v==1:
                        self.sampleids[name.split("_all.fa")[0].split(".")[0]]=subset_bylength(os.path.join(indir,name),os.path.join(outpath,indir2,name),subsetval,plen,postdemlen)
                        cmd='disttbfast.exe '+parstring+' -i '+os.path.join(outpath,indir2,name) 
                else:
                    cmd='disttbfast.exe '+parstring+' -i '+os.path.join(outpath,indir2,name)  
                with open(os.devnull, 'w') as devnull:
                    stdout=subprocess.check_output(cmd,shell=True, stderr=devnull)
                with open(os.path.join(outpath,indir2+'_mafft',name + '_aln.fasta'), "wb") as handle:
                    handle.write(stdout)
                
                transcheckeach,conseq,flag,cov=callconsensus(os.path.join(outpath,indir2+"_mafft",name+"_aln.fasta"),fixthresh,5,name)
                otherconseqs=[]
                if flag==True:
                    self.transcheck[name.split("_all.fa")[0].split(".")[0]]=transcheckeach
                    self.conseqs[name.split("_all.fa")[0].split(".")[0]]=conseq
                    self.flags[name.split("_all.fa")[0].split(".")[0]]=flag
                    if cov!="NA":
                        self.coverages[name.split("_all.fa")[0].split(".")[0]]=cov
                else:
                    n=rangefreq[1]
                    while n>=rangefreq[0]:
                        if n!=fixthresh:
                            transcheckeach2,conseq2,flag2,cov2=callconsensus(os.path.join(outpath,indir2+"_mafft",name+"_aln.fasta"),n,5,name)
                            if flag2==True:
                                if conseq2 not in otherconseqs:
                                    otherconseqs.append(conseq2)
                        n-=stepsize
                if len(otherconseqs)==1:
                    self.transcheck[name.split("_all.fa")[0].split(".")[0]]="1"
                    self.conseqs[name.split("_all.fa")[0].split(".")[0]]=otherconseqs[0]
                    self.flags[name.split("_all.fa")[0].split(".")[0]]=True
                    if cov!="NA":
                        self.coverages[name.split("_all.fa")[0].split(".")[0]]=cov
                else:
                    self.transcheck[name.split("_all.fa")[0].split(".")[0]]=transcheckeach
                    self.conseqs[name.split("_all.fa")[0].split(".")[0]]=conseq
                    self.flags[name.split("_all.fa")[0].split(".")[0]]=flag
                    if cov!="NA":
                        self.coverages[name.split("_all.fa")[0].split(".")[0]]=cov
            except Application.ApplicationError:
                self.transcheck[name.split("_all.fa")[0].split(".")[0]]="NA"
                self.conseqs[name.split("_all.fa")[0].split(".")[0]]=""
                self.flags[name.split("_all.fa")[0].split(".")[0]]=False
                if cov!='NA':
                    self.coverages[name.split("_all.fa")[0].split(".")[0]]="NA"
            self.notifyProgress.emit(c+1)
        self.taskFinished.emit(0)#  return 0
def pool_init(queue):
    filterreads.queue = queue
def pool_init3(queue):
    runconsensuspartsmafft.queue = queue
def pool_init2(queue):
    runtoptwenty.queue = queue
def pool_init1(queue):
    rundemultiplex.queue = queue
class MSAcheck(QtCore.QThread):
    taskFinished = QtCore.pyqtSignal(int)
    notifyProgress1 = QtCore.pyqtSignal(list)
    notifyProgress2 = QtCore.pyqtSignal(list)
    notifyProgress3 = QtCore.pyqtSignal(str)
    notifyProgress4 = QtCore.pyqtSignal(str)
    def __init__(self, outpath, plen,filename,task,dirname, mode,indir,goodfile,errfile,n90subset,outdir,ngood,corlist,prefix,dirdict,parent=None):
        super(MSAcheck, self).__init__(parent)
        self.outpath=outpath
        self.plen=plen
        self.filename=filename
        self.task=task
        self.dirname=dirname
        self.mode=mode
        self.indir=indir
        self.goodfile=goodfile
        self.errfile=errfile
        self.n90subset=n90subset
        self.outdir=outdir
        self.ngood=ngood
        self.tocorlist=corlist
    #   print "corlist",corlist,ngood
        self.prefix=prefix
        self.dirdict=dirdict
    def run(self):
        filename=self.filename
        ambiguity_codes=[("R", "A"), ("R", "G"),("M", "A"), ("M", "C"),("S", "C"), ("S", "G"),("Y", "C"), ("Y", "T"),("K", "G"), ("K", "T"),("W", "A"), ("W", "T"),("V", "A"), ("V", "C"),("V", "G"),("H", "A"),("H", "C"),("H", "T"),("D", "A"), ("D", "G"),("D", "T"), ("B", "C"),("B", "G"), ("B", "T"),("N", "A"), ("N", "G"), ("N", "C"), ("N", "T")]
        #print 'usr\lib\mafft\disttbfast.exe -q 0 -E 2 -V -1.53 -s 0.0 -W 6 -O -C 4-4 -b 62 -g 0 -f -1.53 -Q 100.0 -h 0 -F -X 0.1 -x 1000 -i ' +os.path.join(self.outpath,"barcodesets","temps",filename.split("_")[0]+"_temp.fa") +" > " +os.path.join(self.outpath,"barcodesets","temps",filename.split("_")[0]+"_temp_aln.fa")
        if self.ngood>=3:
            cmd='disttbfast.exe -q 0 -E 2 -V -1.53 -s 0.0 -W 6 -O -C 4-4 -b 62 -g 0 -f -1.53 -Q 100.0 -h 0 -F -X 0.1 -x 1000 -i ' +os.path.join(self.outpath,"barcodesets","temps",self.goodfile) 
            stdout=subprocess.check_output(cmd,shell=True)
            with open(os.path.join(self.outpath,"barcodesets","temps",self.goodfile.split(".")[0]+"_aln.fa"), "wb") as handle:
                handle.write(stdout)
            conseq,seqdict=self.callconsensus(os.path.join(self.outpath,"barcodesets","temps",self.goodfile.split(".")[0]+"_aln.fa"),0.5)
            self.notifyProgress3.emit("done")
        seqdictgood={}
        
        badbarcodes={}
        with open(os.path.join(self.outpath,"barcodesets",self.outdir,self.errfile)) as bfile:
            l=bfile.readlines()
            badbarcodes={}
            for i,j in enumerate(l):
                if ">" in j:
                    badbarcodes[j[1:].split(";")[0]]=l[i+1].strip()
        ngoodbarcodes=0
    #   print "good corlist",len(self.tocorlist)
        with open(os.path.join(self.outpath,"barcodesets",self.outdir,self.prefix+"_predgood_barcodes.fa"),'w') as gfile:
            with open(os.path.join(self.outpath,"barcodesets",self.outdir,self.errfile),'a') as bfile:
                if self.ngood>=3:
                    
                    for n,each in enumerate(seqdict.keys()):
                        if each.split(";")[0] in self.tocorlist:
                #           print  each.split(";")[0]
                            flag=True
                            errcount=0
                            for i,j in enumerate(seqdict[each]):
                                if j=="-":
                                    if conseq[i]!="-":
                                        errcount+=1
                                        flag=False
                                else:
                                    if conseq[i]=="-":
                                        errcount+=1
                                        flag=False
                            if flag==True:
                                gfile.write(">"+each+";estgaps="+str(errcount)+'\n'+seqdict[each].replace("-","").upper()+'\n')
                                seqdictgood[each]=seqdict[each]
                                ngoodbarcodes+=1
                                                                             
                            else:
                                bfile.write(">"+each+";estgaps="+str(errcount)+'\n'+seqdict[each].replace("-","").upper()+'\n')
                                badbarcodes[each+";estgaps="+str(errcount)]=seqdict[each]
                                                                             
                        self.notifyProgress2.emit([n+1,len(self.tocorlist)])
                else:
                    seqdict=self.builddict_sequences(os.path.join(self.outpath,"barcodesets","temps",self.goodfile))
                    for n,each in enumerate(seqdict.keys()):
                        if each.split(";")[0] in self.tocorlist: 
                            bfile.write(">"+each+'\n'+seqdict[each].replace("-","").upper()+'\n')
                            badbarcodes[each]=seqdict[each]
                            self.notifyProgress2.emit([n+1,len(self.tocorlist)])
        self.notifyProgress4.emit('done')
        if self.task==1:
            try:
                os.mkdir(os.path.join(self.outpath,"2b_ConsensusBySimilarity",self.dirname))
            except OSError:
                pass
            #os.mkdir(os.path.join(self.outpath,"90perc","demultiplexed"))
            if self.mode==0:
                dirlist=os.listdir(os.path.join(self.outpath,"demultiplexed"))
            else:
                dirlist=os.listdir(os.path.join(self.indir))
         #   print self.dirdict
            refdict={}
            for each in badbarcodes.keys():
                refdict[each.split(';')[0]]=badbarcodes[each].upper().replace("-","")
            with open(os.path.join(self.outpath,"2b_ConsensusBySimilarity","summary"),'w') as outfile2:
                for i,f in enumerate(refdict.keys()):
                    with open(os.path.join(self.outpath,"2b_ConsensusBySimilarity",self.dirname,f),'w') as outfile:
                        ddict={}
                        if self.mode==0:
                            seqdict=self.builddict_sequences(os.path.join(self.outpath,"demultiplexed",f))
                        else:
                        #    print f
                            seqdict=self.builddict_sequences(os.path.join(self.indir,self.dirdict[f]))
                        for seqid in seqdict.keys():
                            try:
                                k = edlib.align(seqdict[seqid].replace("P","T").replace("E","A").replace("F","G").replace("Q","C"), refdict[f], mode='NW',task='path',additionalEqualities=ambiguity_codes)
                                d = k['editDistance']
                                if d<self.plen*0.1:
                                    ddict[seqid]=d
                            except KeyError:
                                pass
                                
                        sorted_d=sorted(ddict.items(), key=lambda x: x[1])
                        outfile2.write(f+'\t'+str(len(sorted_d))+'\n')
                        for k in sorted_d[:self.n90subset]:
                            outfile.write(">"+k[0]+'\n'+seqdict[k[0]]+'\n')
                    self.notifyProgress1.emit([i+1,len(refdict.keys())])
    #   print "msa1 done"
        self.taskFinished.emit(ngoodbarcodes)
    def builddict_sequences(self,infile):
        seqdict={}
        with open(infile) as inseqs:
            l=inseqs.readlines()
            for i,j in enumerate(l):
                if ">" in j:
                    seqdict[j.strip().replace(">","")]=l[i+1].strip()
        return seqdict      

    def consensus(self,indict,perc_thresh):
        poslist=[]
        n=0
        while n<len(list(indict.values())[0]):
            newlist=[]
            for i,each in enumerate(indict.keys()):
                try:
                    newlist.append(indict[each][n])
                except IndexError:
            #       print each,indict[each],indict.values()[0],len(indict[each]),len(indict.values()[0])
                    break
            poslist.append(newlist)
            n+=1
        sequence=[]
        countpos=1
        for character in poslist:
            charcounter=Counter(character)
            baseset={}
            for k,v in charcounter.items():
                percbp=float(v)/float(len(character))
                if percbp>perc_thresh:
                    baseset[k]=v
            if len(baseset)==0:
                bp='N'
            if len(baseset)==1:
                bp=list(baseset.keys())[0]
            if len(baseset)>1:
                bp="N"
            sequence.append(bp)
            countpos+=1
        return ''.join(sequence)

    def callconsensus(self,i,perc_thresh):
        with open(i) as infile:
            l=infile.readlines()
            seqdict,poslist={},[]
            for i,j in enumerate(l):
                if ">" in j:
                    poslist.append(i)
            for i,j in enumerate(poslist):
                ambcounts=0
                k01=l[j].strip().split('>')[1]
                if i!=len(poslist)-1:
                    k3=l[j+1:poslist[i+1]]
                if i==len(poslist)-1:
                    k3=l[j+1:]
                k4=''.join(k3).replace('\n','')
                seqdict[k01]=k4
            conseq=self.consensus(seqdict,perc_thresh)
            return conseq,seqdict

def rundemultiplex(inlist):
    inlist1=inlist[0]
    outpath=inlist[1]
    tagdict=inlist[2]
    muttags_fr=inlist[3]
    sampledict=inlist[4]
    typedict=inlist[5]
    primerfset=inlist[6]
    taglen=inlist[8]
    primerrset=inlist[7]
    num_row=inlist[9]
    maxid=inlist[10]
    maxv=inlist[11]
    typelist=inlist[12]
    start=100
    primermismatch=int(inlist[13])
    def findmatch_m2(taglist,muttags_fr,indict,seqdict,sampledict,typedict,num_row):
    #   print taglist
    #   print "sample",sampledict
        dmpfile=open(os.path.join(outpath,"dmpfile"),'a')
        idcombs={}
        cumscores={}
        for each in indict.keys():
            try:
                idcombs[each]=(taglist[indict[each][0]], taglist[indict[each][1]])
                cumscore=typedict[indict[each][0]]+typedict[indict[each][1]]
                cumscores[each]=[str(typedict[indict[each][0]]),str(typedict[indict[each][1]]),str(cumscore)]
            except KeyError:
                pass
        
        matched_idcombs={}
        for each in idcombs.keys():
            try:
                with open(os.path.join(outpath,str(num_row),sampledict[idcombs[each]]+"_all.fa"),'a') as outfile:
                    outfile.write(">"+each+"_"+"_".join(cumscores[each])+"\n"+seqdict[each]+"\n")
            except KeyError:
                pass
                    
        dmpfile.close() 
    def readprimertagfasta(filef,filer):
        indict2,indict1,indict={},{},{}
        with open(filef) as infile1:
            with open(filer) as infile2:
                l1=infile1.readlines()
                l2=infile2.readlines()
                for i,j in enumerate(l1):
                    if ">" in j:
                        indict1[j.strip().replace(">","")]=l1[i+1].strip()
                        indict[j.strip().replace(">","")]=[0,0]
                for i,j in enumerate(l2):
                    if ">" in j:
                        indict2[j.strip().replace(">","")]=l2[i+1].strip()
                        indict[j.strip().replace(">","")]=[0,0]
        return indict1,indict2,indict
    def builddict_sequences(infile):
        seqdict={}
        with open(infile) as inseqs:
            l=inseqs.readlines()
            for i,j in enumerate(l):
                if ">" in j:
                    seqdict[j.strip().replace(">","")]=l[i+1].strip()
        return seqdict
    def modseq(seq):
        for bp in ["A","T","G","C"]:
            n=20
            while n>2:
                seq=seq.replace(bp*n,bp*2)
                n-=1
        return seq
    def demultiplexfunc(pf,pr,seqdict,tagdict,muttags_fr,sampledict,typedict,num_row):
        indict1,indict2,indict=readprimertagfasta(pf,pr)

        for each in list(indict.keys()):
            try: 
                indict[each][0]=indict1[each]
                indict[each][1]=indict2[each]
            except KeyError:
                del indict[each]
                
        try:
            with open(os.path.join(outpath,"dmpfile_tagfr"),'w') as tagfrdmp:   
                for tag in muttags_fr.keys():
                    tagfrdmp.write(tag+'\t'+muttags_fr[tag]+'\n')
            findmatch_m2(tagdict, muttags_fr,indict,seqdict,sampledict,typedict,num_row)
        except UnboundLocalError:
            findmatch_m2(tagdict, {},indict,seqdict,sampledict,typedict,num_row)
    def cleantagfile(infile):
        with open(infile) as primertagfile:
            with open(infile+"cleaned",'w') as primertagfile_cleaned:
                taglines=primertagfile.readlines()
                ids=[]
                seqs={}
                for n,line in enumerate(taglines):
                    if ">" in line:
                        ids.append(line)
                        seqs[line]=modseq(taglines[n+1])[-(taglen+1):]

                idcounts=Counter(ids)
                nseqs=0
                for id in ids:
                    if idcounts[id]==1:
                        nseqs+=1
                        primertagfile_cleaned.write(id+seqs[id])

    labeltext=""

    c=0
    nseqs=0
    for infile in inlist1:
        if infile==maxid:
            nseqs+=maxv
        else:
            nseqs+=20000
    typedict2={}
    for each in typelist:
        typedict2[each[0]]=each[1]
    for infile in inlist1:
        ambiguity_codes=[("R", "A"), ("R", "G"),("M", "A"), ("M", "C"),("S", "C"), ("S", "G"),("Y", "C"), ("Y", "T"),("K", "G"), ("K", "T"),("W", "A"), ("W", "T"),("V", "A"), ("V", "C"),("V", "G"),("H", "A"),("H", "C"),("H", "T"),("D", "A"), ("D", "G"),("D", "T"), ("B", "C"),("B", "G"), ("B", "T"),("N", "A"), ("N", "G"), ("N", "C"), ("N", "T")]
        inputseqs=builddict_sequences(os.path.join(outpath,infile))
        seqlens=[len(i.strip()) for i in inputseqs.values()]
        with open(os.path.join(outpath,infile+"_all_glsearch1.parsed.lencutoff5parsed_f"),'w') as tagfoutfile:
            with open(outpath+"/"+infile+"_all_glsearchR.parsed.lencutoff5parsed_r",'w') as tagroutfile:
                with open(outpath+"/"+infile+"_all_glsearchR.parsed.lencutoff5endr",'w') as fprimercleanfile:
                    for pf in primerfset:
                        for pr in primerrset:
                            pr=seqpy.revcomp(pr)
                            for n,inseq in enumerate(inputseqs.keys()):
                                c+=1
                                if n%1000==0:
                                    progress = float(c) 
                                    rundemultiplex.queue.put((num_row, progress,labeltext))
                                compseq=seqpy.revcomp(inputseqs[inseq])
                                k1 = edlib.align(pf, inputseqs[inseq][:typedict2[infile]], mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                k2 = edlib.align(pf, compseq[:typedict2[infile]], mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                d1 = k1['editDistance']
                                d2 = k2['editDistance']
                                orient=0
                                revseq=''
                                if d1<=d2:
                                    if d1<=primermismatch:
                                        loc=k1['locations'][0]
                                        tagseq=inputseqs[inseq][loc[0]-taglen:loc[0]]
                                        if len(tagseq)!=0:
                                        #   print "pf",inseq,d1,loc,len(rev_comp(compseq[loc[1]+1:]))
                                            tagfoutfile.write(">"+inseq+'\n'+tagseq+'\n')
                                            revseq=inputseqs[inseq][loc[1]+1:]
                                else:
                                    if d2<=primermismatch:
                                        loc=k2['locations'][0]
                                        
                                        tagseq=compseq[loc[0]-taglen:loc[0]]
                                        if len(tagseq)!=0:
                                        #   print "pfR",inseq,d2,loc,len(rev_comp(compseq[loc[1]+1:]))
                                            tagfoutfile.write(">"+inseq+'\n'+tagseq+'\n')
                                            revseq=compseq[loc[1]+1:]
                                            orient=1
                                
                                if revseq!='':
                                    k = edlib.align(pr, revseq[-typedict2[infile]:], mode='HW',task='path',additionalEqualities=ambiguity_codes)
                                    d = k['editDistance']
                                    
                                    if d<=primermismatch:
                                        loc=k['locations'][0]
                                        
                                        startpoint=len(revseq)-typedict2[infile]
                                        
                                        tagseq=seqpy.revcomp(revseq[startpoint+loc[1]+1:startpoint+loc[1]+taglen+1])
                                        if len(tagseq)!=0:
                                            tagroutfile.write(">"+inseq+'\n'+tagseq+'\n')
                                            if orient==0:
                                                fprimercleanfile.write(">"+inseq+'\n'+revseq[:startpoint+loc[0]]+'\n')
                                            else:
                                                fprimercleanfile.write(">"+inseq+'\n'+revseq[:startpoint+loc[0]].replace("A","E").replace("G","F").replace("C","Q").replace("T","P")+'\n')
                                        
        #cleantagfile(outpath+"/"+infile+"_all_glsearch1.parsed.lencutoff5parsed_f")
        #cleantagfile(outpath+"/"+ infile+"_all_glsearchR.parsed.lencutoff5parsed_r")
        inputseqs=builddict_sequences(outpath+"/"+ infile+"_all_glsearchR.parsed.lencutoff5endr")
        demultiplexfunc(outpath+"/"+infile+"_all_glsearch1.parsed.lencutoff5parsed_f",outpath+"/"+infile+"_all_glsearchR.parsed.lencutoff5parsed_r", inputseqs,tagdict,muttags_fr,sampledict,typedict,num_row)
    rundemultiplex.queue.put((num_row, nseqs,labeltext))
class PicButton(QtWidgets.QAbstractButton):
    def __init__(self, pixmap, pixmap_hover, pixmap_pressed, parent=None):
        super(PicButton, self).__init__(parent)
        self.pixmap = pixmap
        self.pixmap_hover = pixmap_hover
        self.pixmap_pressed = pixmap_pressed
        self.toggled.connect(self.setcheckedbtn)
      #   self.released.connect(self.update)

    def paintEvent(self, event):
        pix = self.pixmap_hover if self.underMouse() else self.pixmap
        if self.isChecked():
            pix = self.pixmap_pressed
        else:
            pix = self.pixmap           

        painter = QtGui.QPainter(self)
        painter.drawPixmap(event.rect(), pix)

    def enterEvent(self, event):
    #    print "enter"
        self.update()

    def leaveEvent(self, event):
     #   print "leave"
        self.update()
    def setcheckedbtn(self):
    #    print "toggling"
        
        self.update()

    def sizeHint(self):
        return QtCore.QSize(300, 300)




class OptWindow(QtWidgets.QWidget):
    def __init__(self,parent=None):
        super(OptWindow,self).__init__(parent)
        self.setAcceptDrops(True)
        
        self.conventionalbutton= PicButton(QtGui.QPixmap("logos/conventional_unclicked.png"),QtGui.QPixmap("logos/conventional_hover.png"),QtGui.QPixmap("logos/conventional_clicked.png"))
        self.conventionalbutton.setCheckable(True)
        self.conventionalbutton.toggled.connect(self.setconventionalmode)
        self.realtimebutton= PicButton(QtGui.QPixmap("logos/realtime_unclicked.png"),QtGui.QPixmap("logos/realtime_hover.png"),QtGui.QPixmap("logos/realtime_clicked.png"))
        self.realtimebutton.setCheckable(True)
        self.realtimebutton.toggled.connect(self.showdeviceopt)
        
        self.mk1bbutton= PicButton(QtGui.QPixmap("logos/mk1b_unclicked.png"),QtGui.QPixmap("logos/mk1b_hover.png"),QtGui.QPixmap("logos/mk1b_clicked.png"))
        self.mk1bbutton.setCheckable(True)
        self.mk1bbutton.toggled.connect(lambda:self.showgpuopt("1b"))
        self.mk1cbutton= PicButton(QtGui.QPixmap("logos/mk1c_unclicked.png"),QtGui.QPixmap("logos/mk1c_hover.png"),QtGui.QPixmap("logos/mk1c_clicked.png"))
        self.mk1cbutton.setCheckable(True)
        self.mk1cbutton.toggled.connect(lambda:self.showgpuopt("1c"))
        self.gpu1cbutton1= PicButton(QtGui.QPixmap("logos/gpumk1b-1_unclicked.png"),QtGui.QPixmap("logos/gpumk1b_hover.png"),QtGui.QPixmap("logos/gpumk1b_clicked.png"))
        self.gpu1cbutton1.setCheckable(True)
        self.gpu1cbutton1.toggled.connect(self.checkgpuopt1)
        self.gpu1cbutton2= PicButton(QtGui.QPixmap("logos/gpumk1b-2_unclicked.png"),QtGui.QPixmap("logos/gpumk1b-2_hover.png"),QtGui.QPixmap("logos/gpumk1b-2_clicked.png"))
        self.gpu1cbutton2.setCheckable(True)
        self.gpu1cbutton2.toggled.connect(self.checkgpuopt2)
        self.emptywidget1=QtWidgets.QWidget()
        self.emptywidget2=QtWidgets.QWidget()
        self.emptywidget3=QtWidgets.QWidget()
        self.emptywidget4=QtWidgets.QWidget()
        self.emptywidget5=QtWidgets.QWidget()
        self.emptywidget6=QtWidgets.QWidget()
        self.bioinfotitlelayout=QtWidgets.QHBoxLayout()
        self.bioinfobuttonlayout=QtWidgets.QHBoxLayout()
        self.devicetitlelayout=QtWidgets.QHBoxLayout()
        self.devicebuttonlayout=QtWidgets.QHBoxLayout()
        self.gputitlelayout=QtWidgets.QHBoxLayout()
        self.gpubuttonlayout=QtWidgets.QHBoxLayout()
        
        self.bioinfotitle=QtWidgets.QLabel("Bioinformatics")
        self.bioinfotitlelayout.addWidget(self.bioinfotitle,0,QtCore.Qt.AlignCenter)
        self.bioinfotitlelayout.setContentsMargins(0, 0, 0, 0)
        self.bioinfotitle.setAttribute(QtCore.Qt.WA_TranslucentBackground)

        self.devicetitle=QtWidgets.QLabel("Sequencing Device")
        self.devicetitlelayout.addWidget(self.emptywidget5,0,QtCore.Qt.AlignCenter)
        self.devicetitlelayout.setContentsMargins(0, 0, 0, 0)

        self.gputitle=QtWidgets.QLabel("Basecaller")
        self.gputitlelayout.addWidget(self.emptywidget6,0,QtCore.Qt.AlignCenter)
        self.gputitlelayout.setContentsMargins(0, 0, 0, 0)

        self.bioinfobuttonlayout.addWidget(self.conventionalbutton)
        self.bioinfobuttonlayout.addWidget(self.realtimebutton)
        self.devicebuttonlayout.addWidget(self.emptywidget1)
        self.devicebuttonlayout.addWidget(self.emptywidget2)
        self.gpubuttonlayout.addWidget(self.emptywidget3)
        self.gpubuttonlayout.addWidget(self.emptywidget4)

        
        self.bioinfotitlewidget=QtWidgets.QWidget()
        self.bioinfobuttonwidget=QtWidgets.QWidget()
        self.devicetitlewidget=QtWidgets.QWidget()
        self.devicebuttonwidget=QtWidgets.QWidget()
        self.gputitlewidget=QtWidgets.QWidget()
        self.gpubuttonwidget=QtWidgets.QWidget()
        self.bioinfotitlewidget.setFixedHeight(40)
        self.bioinfotitlewidget.setFixedWidth(500)
        self.bioinfobuttonwidget.setFixedHeight(200)
        self.bioinfobuttonwidget.setFixedWidth(500)
        self.devicetitlewidget.setFixedHeight(40)
        self.devicetitlewidget.setFixedWidth(500)
        self.devicebuttonwidget.setFixedHeight(200)
        self.devicebuttonwidget.setFixedWidth(500)
        self.gputitlewidget.setFixedHeight(40)
        self.gputitlewidget.setFixedWidth(500)
        self.gpubuttonwidget.setFixedHeight(200) 
        self.gpubuttonwidget.setFixedWidth(500)
        self.bioinfotitlewidget.setLayout(self.bioinfotitlelayout)
        self.bioinfobuttonwidget.setLayout(self.bioinfobuttonlayout)
        self.devicetitlewidget.setLayout(self.devicetitlelayout)
        self.devicebuttonwidget.setLayout(self.devicebuttonlayout)
        self.gputitlewidget.setLayout(self.gputitlelayout)
        self.gpubuttonwidget.setLayout(self.gpubuttonlayout)
        self.bioinfotitlewidget.setStyleSheet("""QWidget {background-color: qlineargradient(spread:pad, x1:0 y1:0, x2:1 y2:0, stop:0 rgb(122, 173, 235), stop:1 white);}""")
        
   #     self.bioinfobuttonwidget.setStyleSheet("""QWidget {background-color: qlineargradient( x1:0 y1:0, x2:1 y2:0, stop:0 8D98E7, stop:1 e3e5f9;}""")
 #       self.devicetitlewidget.setStyleSheet("""QWidget {background-color: qlineargradient(spread:pad, x1:0 y1:0, x2:1 y2:0, stop:0 grey, stop:1 white);}""")
    #    self.devicebuttonwidget.setStyleSheet("""QWidget {background-color: rgba(124, 164, 229,100);}""")
  #      self.gputitlewidget.setStyleSheet("""QWidget {background-color: qlineargradient(spread:pad, x1:0 y1:0, x2:1 y2:0, stop:0 rgb(252, 185, 62), stop:1 white);}""")        
     #   self.gpubuttonwidget.setStyleSheet("""QWidget {background-color: rgba(252, 185, 62,100);}""")
        self.bioinfotitle.setStyleSheet('font-size: 30px; font-family: Bahnschrift;')
        self.devicetitle.setStyleSheet('font-size: 30px; font-family: Bahnschrift;')
        self.gputitle.setStyleSheet('font-size: 30px; font-family: Bahnschrift;')        



        self.initstacklayout=QtWidgets.QVBoxLayout()

        self.initstacklayout.addWidget(self.bioinfotitlewidget,0,alignment=QtCore.Qt.AlignCenter)
        self.initstacklayout.addWidget(self.bioinfobuttonwidget,0,alignment=QtCore.Qt.AlignCenter)
        self.initstacklayout.addWidget(self.devicetitlewidget,0,alignment=QtCore.Qt.AlignCenter)
        self.initstacklayout.addWidget(self.devicebuttonwidget,0,alignment=QtCore.Qt.AlignCenter)
        self.initstacklayout.addWidget(self.gputitlewidget,0,alignment=QtCore.Qt.AlignCenter)
        self.initstacklayout.addWidget(self.gpubuttonwidget,0,alignment=QtCore.Qt.AlignCenter)

        
        self.processlabel=QtWidgets.QLabel()
        QtWidgets.QToolTip.setFont(QtGui.QFont('SansSerif',10))
        InputMessage=QtWidgets.QLabel('Following information is required:')
        avGeom = QtWidgets.QDesktopWidget().screenGeometry()

    #   #print avGeom.height()
        self.threadpool = QtCore.QThreadPool()
        self.pbar=QtWidgets.QProgressBar()
        self.pbar.setRange(0,1)
        self.h_splitter=QtWidgets.QSplitter(QtCore.Qt.Vertical)


        self.rightframe=QtWidgets.QFrame(self)
        self.rtopframe=QtWidgets.QFrame(self)
        self.h_splitter.addWidget(self.rtopframe)
        self.h_splitter.addWidget(self.rightframe)
        self.Tab = QtWidgets.QTabWidget()

        self.rtopframelayout=QtWidgets.QHBoxLayout()
        self.rtopframe.setLayout(self.rtopframelayout)

        self.comps={'A':'T','T':'A','G':'C','C':'G','N':'N','M':'K','R':'Y','W':'W','S':'S','Y':'R','K':'M','V':'B','H':'D','D':'H','B':'V'}
        self.ambiguity_codes=[("R", "A"), ("R", "G"),("M", "A"), ("M", "C"),("S", "C"), ("S", "G"),("Y", "C"), ("Y", "T"),("K", "G"), ("K", "T"),("W", "A"), ("W", "T"),("V", "A"), ("V", "C"),("V", "G"),("H", "A"),("H", "C"),("H", "T"),("D", "A"), ("D", "G"),("D", "T"), ("B", "C"),("B", "G"), ("B", "T"),("N", "A"), ("N", "G"), ("N", "C"), ("N", "T")]

        # List of views.
        initialtext="Please drag and drop your input FASTA file which will be split into haplotypes"
        self.filepath,self.csvdemfile,self.exceldemfile='','',''
        self.scrolllive=QtWidgets.QScrollArea()                            
        scrollarea=QtWidgets.QScrollArea()
        self.scrollconventional=QtWidgets.QScrollArea()  
        views = []
        layout = QtWidgets.QHBoxLayout(self)
        layout.addWidget(self.h_splitter)

        self.h_splitter.resize(int(self.size().width()*0.95),self.size().height())
        self.rtopframe.resize(int(self.size().width()*0.95),int(self.size().height()*0.99))
        self.rightframe.resize(int(self.size().width()*0.95),int(self.size().height()*0.01))
        #for ViewType in (QtWidgets.QColumnView, QtWidgets.QTreeView):
        # Create the view in the splitter.

        self.grid_right=QtWidgets.QHBoxLayout()
        self.grid_right.setSpacing(0)
        self.grid_right.setContentsMargins(0, 0, 0, 0)                                                      
        self.rframelabel=QtWidgets.QLabel(initialtext)
    #   self.setupbtn=BlinkButton("Initializing")
        self.compbar=QtWidgets.QProgressBar()
        self.quitbtn=QtWidgets.QPushButton("Quit")
        self.quitbtn.adjustSize()
        self.quitbtn.clicked.connect(self.terminate)
        axislabelStyle = {'color': '#000000', 'font-size': '11pt'}
        chartlabelStyle = {'color': '#000000', 'font-size': '16pt','font-weight': 'bold'}
        self.sumtext='<center><b><u> ONTBarcoder </u></b> <br><br>Minion based DNA barcoding<br>'
        self.sumdesc='<center><p>This pipeline aims to demultiplex and call consensus barcodes for specimen based DNA barcoding.<br><br> Please start by dragging files depending on which mode you want to run<br><br>For further instructions please see: <br><a href=https://github.com/asrivathsan/ONTbarcoder/blob/main/ONTBarcoder_manual.pdf>README</a></center>'
        self.opttable=SelectionTable()
        
        self.dembtn=QtWidgets.QPushButton("Let's go")
        self.dembtn.clicked.connect(self.setpostmode)
        self.dembtn.setStyleSheet("background-color: #0d6efd;")
        # self.dembtn.setStyleSheet("background-color: lightgrey;")
        #labelImage.setPixmap(pixmap)
        self.sumtablabel=QtWidgets.QLabel(self.sumtext)
        self.sumtablabel.setStyleSheet("margin-left:30;" "margin-right:30;")
        self.sumtablabel.setFont(QtGui.QFont("Arial", 12, QtGui.QFont.Bold))
        self.sumdescbel=QtWidgets.QLabel(self.sumdesc)
        self.sumdescbel.setStyleSheet("margin-left:30;" "margin-right:30;")
        self.sumdescbel.setOpenExternalLinks(True)
        self.sumdescbel.setTextFormat(QtCore.Qt.RichText)
        self.sumdescbel.setFont(QtGui.QFont("Arial", 10))
        self.sumstack=QtWidgets.QStackedWidget()
        self.sumstack.setStyleSheet("background-color: white;")
        self.conventionalstack=QtWidgets.QStackedWidget()
        self.conventionalstack.setStyleSheet("background-color: white;")

        self.sumstacklayout=QtWidgets.QGridLayout()
        self.sumwidget=QtWidgets.QWidget()
        self.livestack=QtWidgets.QStackedWidget()
        self.livestack.setStyleSheet("background-color: white;")
        self.livestacklayout=QtWidgets.QGridLayout()
        self.livewidget=QtWidgets.QWidget()
        self.liveprogresswidget=QtWidgets.QWidget()
        self.liveprogresslayout=QtWidgets.QGridLayout()
        self.livelogstatus = QtWidgets.QTextEdit()
        self.livelogstatus.setReadOnly(True)
        self.livelogstatus.setTextInteractionFlags(QtCore.Qt.NoTextInteraction) 
        self.timestamp=datetime.datetime.fromtimestamp(int(time.time()))
        font=QtGui.QFont()
        font.setPixelSize(12)
        font2=QtGui.QFont()
        font2.setPixelSize(14)
        self.livereadgraphwidget=pg.PlotWidget()                     # width of the window displaying the curve
   #     self.livereadgraphwidget=pg.PlotWidget(title="Total number of reads")
        self.livereadgraphwidget.setTitle("Total number of reads", **chartlabelStyle)
        self.livereadgraphwidget.setLabel('bottom',"Time (minutes)", **axislabelStyle)
        self.livereadgraphwidget.setLabel('left',"Number of reads" , **axislabelStyle)
        self.livereadgraphwidget.setBackground('w')
        self.livereadgraphwidget.getAxis("bottom").setTickFont(font)
        self.livereadgraphwidget.getAxis("left").setTickFont(font)
        self.livereadgraphwidget.getAxis("bottom").setPen('k')
        self.livereadgraphwidget.getAxis("left").setPen('k')
        self.livereadgraphwidget.getAxis("bottom").setTextPen('k')
        self.livereadgraphwidget.getAxis("left").setTextPen('k') 
        self.livereadgraph=self.livereadgraphwidget.plot(symbol='o', pen =(195, 46, 212), symbolBrush =(195, 46, 212))
        self.livereadgraphXm=np.full(1,float(0))
        self.livereadgraphYm=np.full(1,float(0))
        self.livereadgraph.setData(self.livereadgraphXm,self.livereadgraphYm)
    #   print  self.livereadgraphXm                  # update x position for displaying the curve
    #   self.livereadgraph.setData(self.livereadgraphXm)                     # set the curve with this data
        self.livereadgraph.setPos(0,0) 
        self.livedemultiplexedusedgraphwidget=pg.PlotWidget()
        self.livedemultiplexedusedgraphwidget.setTitle("Number of demultiplexed reads used", **chartlabelStyle)
        self.livedemultiplexedusedgraph=self.livedemultiplexedusedgraphwidget.plot(symbol='o',pen =(0, 0, 200), symbolBrush =(0, 0, 200))
        self.livedemultiplexedusedgraphwidget.setLabel('bottom',"Time (minutes)", **axislabelStyle)
        self.livedemultiplexedusedgraphwidget.setLabel('left',"Number of reads", **axislabelStyle)
        self.livedemultiplexedusedgraphwidget.setBackground('w')
        self.livedemultiplexedusedgraphwidget.getAxis("bottom").setTickFont(font)
        self.livedemultiplexedusedgraphwidget.getAxis("left").setTickFont(font)
        self.livedemultiplexedusedgraphwidget.getAxis("bottom").setPen('k')
        self.livedemultiplexedusedgraphwidget.getAxis("left").setPen('k')
        self.livedemultiplexedusedgraphwidget.getAxis("bottom").setTextPen('k')
        self.livedemultiplexedusedgraphwidget.getAxis("left").setTextPen('k')  

        self.livedemusedgraphXm=np.full(1,float(0))
        self.livedemusedgraphYm=np.full(1,float(0))
        self.livedemultiplexedusedgraph.setData(self.livedemusedgraphXm,self.livedemusedgraphYm)
        self.livebarcodegraphwidget=pg.PlotWidget()
        self.livebarcodegraphwidget.setTitle("Number of barcodes meeting QC", **chartlabelStyle)
        self.livebarcodegraphwidget.setLabel('bottom',"Time (minutes)", **axislabelStyle)
        self.livebarcodegraphwidget.setLabel('left',"Number of barcodes", **axislabelStyle)
        self.livebarcodegraph=self.livebarcodegraphwidget.plot(symbol='o', pen =(0, 128, 0), symbolBrush =(0, 128, 0))


        self.livebarcodegraphwidget.setBackground('w')
        self.livebarcodegraphwidget.getAxis("bottom").setPen('k')
        self.livebarcodegraphwidget.getAxis("left").setPen('k')
        self.livebarcodegraphwidget.getAxis("bottom").setTextPen('k')
        self.livebarcodegraphwidget.getAxis("left").setTextPen('k')          
        self.livebarcodegraphwidget.getAxis("bottom").setTickFont(font)
        self.livebarcodegraphwidget.getAxis("left").setTickFont(font)
        self.livebarcodegraphXm=np.full(1,float(0))
        self.livebarcodegraphYm=np.full(1,float(0))
        self.livebarcodegraph.setData(self.livebarcodegraphXm,self.livebarcodegraphYm)
        self.livedemultiplexedgraphwidget=pg.PlotWidget()
        self.livedemultiplexedgraphwidget.setTitle("Number of demultiplexed reads",**chartlabelStyle)
        self.livedemultiplexedgraph=self.livedemultiplexedgraphwidget.plot(symbol='o',pen =(0, 0, 200), symbolBrush =(0, 0, 200))
        self.livedemultiplexedgraphwidget.setLabel('bottom',"Time (minutes)", **axislabelStyle)
        self.livedemultiplexedgraphwidget.setLabel('left',"Number of reads", **axislabelStyle)
        self.livedemultiplexedgraphwidget.setBackground('w')
        self.livedemultiplexedgraphwidget.getAxis("bottom").setTickFont(font)
        self.livedemultiplexedgraphwidget.getAxis("left").setTickFont(font)       
        self.livedemultiplexedgraphwidget.getAxis("bottom").setPen('k')
        self.livedemultiplexedgraphwidget.getAxis("left").setPen('k')
        self.livedemultiplexedgraphwidget.getAxis("bottom").setTextPen('k')
        self.livedemultiplexedgraphwidget.getAxis("left").setTextPen('k')   

        self.livedemgraphXm=np.full(1,float(0))
        self.livedemgraphYm=np.full(1,float(0))
        self.livedemultiplexedgraph.setData(self.livedemgraphXm,self.livedemgraphYm)

                                                                  
        self.liveprogresslayout.addWidget(self.livereadgraphwidget,0,0,1,1)
        self.liveprogresslayout.addWidget(self.livedemultiplexedgraphwidget,0,1,1,1)
        self.liveprogresslayout.addWidget(self.livedemultiplexedusedgraphwidget,1,0,1,1)
        self.liveprogresslayout.addWidget(self.livebarcodegraphwidget,1,1,1,1)
        self.liveprogresswidget.setLayout(self.liveprogresslayout)
                                            
        scrollarea.setWidget(self.sumstack)
        scrollarea.setWidgetResizable(True)
        scrollarea.verticalScrollBar().setSliderPosition(scrollarea.verticalScrollBar().maximum())
        self.scrollconventional.setWidget(self.conventionalstack)
        self.scrollconventional.setWidgetResizable(True)
        self.scrollconventional.verticalScrollBar().setSliderPosition(self.scrollconventional.verticalScrollBar().maximum())             
        self.scrolllive.setWidget(self.livestack)
        self.scrolllive.setWidgetResizable(True)
        self.scrolllive.verticalScrollBar().setSliderPosition(self.scrolllive.verticalScrollBar().maximum())
        self.livetext='<center><b><u> ONTBarcoder </u></b> <br><br>Real-time barcoding<br>'
        self.livedesc='<center><p>This pipeline aims to demultiplex and call consensus barcodes for specimen based DNA barcoding real time.<br><br> Please start by dragging a demultiplexing file</center>'
        self.livetablabel=QtWidgets.QLabel(self.livetext)
        self.livetablabel.setStyleSheet("margin-left:30;" "margin-right:30;")
        self.livetablabel.setFont(QtGui.QFont("SansSerif", 16, QtGui.QFont.Bold))
        self.livedescbel=QtWidgets.QLabel(self.livedesc)
        self.livedescbel.setStyleSheet("margin-left:30;" "margin-right:30;")
        self.livedescbel.setFont(QtGui.QFont("SansSerif", 10))
        self.livedescbel.setOpenExternalLinks(True)
        self.livebtn=QtWidgets.QPushButton("Let's go")
        self.livebtn.setEnabled(False)
        self.livebtn.clicked.connect(self.setlivemode)
    #    self.livebtn.setStyleSheet("background-color: lightgrey;")
        self.livebtn.setAutoFillBackground(True)
     #   self.livestack.setStyleSheet("background-color: white;")
        self.livestacklayout=QtWidgets.QGridLayout()
        self.livestacklayout.addWidget(self.livetablabel,0,0,1,5)
        self.livestacklayout.addWidget(self.livedescbel,1,0,3,5)
        self.livestacklayout.addWidget(self.livebtn,3,2,1,1)
        self.livewidget.setLayout(self.livestacklayout)
        self.livestack.addWidget(self.livewidget)
        self.livestack.addWidget(self.liveprogresswidget)
        self.livestack.setCurrentIndex(0)
        self.sumstacklayout=QtWidgets.QGridLayout()
        self.sumstacklayout.addWidget(self.sumtablabel,0,0,1,5)
        self.sumstacklayout.addWidget(self.sumdescbel,1,0,1,5)
        self.sumstacklayout.addWidget(self.opttable,2,1,4,3)
        self.sumstacklayout.addWidget(self.dembtn,6,2,1,1)
        self.sumwidget.setLayout(self.sumstacklayout)
        self.initwidget=QtWidgets.QWidget()
        self.initwidget.setLayout(self.initstacklayout)
        self.initwidget.resize(int(self.size().width()*0.9),self.size().height())        

        self.sumstack.addWidget(self.initwidget)
        self.conventionalstack.addWidget(self.sumwidget)
        self.sumstack.setCurrentIndex(0)

        self.compstack=QtWidgets.QStackedWidget()
        self.compstack.setStyleSheet("background-color: white;")
        self.compstacklayout=QtWidgets.QGridLayout()        
        self.comptext='<center><b><u> Compare barcodes: </u></b> Please drag in 2 or more fasta files simultaneously. <br><br> This module helps compare the barcode sets, either to <br>(1) compare the barcodes obtained by the software against a reference set of barcodes or <br>(2) compare the different barcode sets obtained by this software. <br><br> Given that pairwise comparisons are done it is essential that the barcode names are identical before the delimiters "_/-/;" for this module</center><br><br>'
        self.comptablabel=QtWidgets.QLabel(self.comptext)
        self.comptablabel.setStyleSheet("margin-left:30;" "margin-right:30;")   
        self.compwidget=QtWidgets.QWidget()
        self.compstacklayout.addWidget(self.comptablabel,0,0,1,3)
        self.compwidget.setLayout(self.compstacklayout)
        self.compstack.addWidget(self.compwidget)
        
        self.rtopframelayout.addWidget(self.Tab)    
        self.selectlenscounter=0
        self.ngoodbarcodescounter=0
        self.con200errn=0
        self.n90errn=0
        self.nerr=0
        self.nfinal=0
        self.nfixed=0
        self.con200trans={}
        self.con200length={}
        self.con200barcodes={}
        self.con200cov={}
        self.con200flags={}
        self.n90goodn=0
    #   self.n90errn=0
        self.Tab.addTab(scrollarea,"Setup")
                                                   
        self.Tab.addTab(self.compstack,"Compare barcode sets")
     #   self.Tab.addTab(scrollconventional,"Conventional barcoding")
    #    self.Tab.addTab(scrolllive,"Real-time barcoding")
        self.Tab.resize(int(self.size().width()*0.95) ,self.size().height())
        self.rightframe.setLayout(self.grid_right)
        self.startbutton=QtWidgets.QPushButton("Start")

        self.startbutton.clicked.connect(self.readsetup)
        self.startbutton.setFixedWidth(100)
        self.startbutton.setFixedHeight(50)
        self.grid_right.setSpacing(0)
        self.grid_right.addWidget(self.startbutton,0,alignment=QtCore.Qt.AlignRight)
        self.grid_right.addWidget(self.quitbtn,0,alignment=QtCore.Qt.AlignRight)
        self.tfileflag=False
        self.statuswidget=PbarWindow()
        self.conventionalstack.addWidget(self.statuswidget)
        self.setWindowTitle('ONTBarcoder')
        self.setMinimumSize(500,950)
        self.show()
    def checkgpuopt1(self):
        if self.gpu1cbutton1.isChecked()==True:
            self.gpu1cbutton2.setChecked(False)

    def checkgpuopt2(self):
        if self.gpu1cbutton2.isChecked()==True:
            self.gpu1cbutton1.setChecked(False)

    def readsetup(self):
        if self.realtimebutton.isChecked():
        #    print "2"
            if self.mk1cbutton.isChecked():
            #    print "22"
                if self.gpu1cbutton1.isChecked()==True:
                    self.runmode="221"
                elif self.gpu1cbutton2.isChecked()==True:
                    self.runmode="222"
            elif self.mk1bbutton.isChecked(): 
                if self.gpu1cbutton1.isChecked()==True:
                    self.runmode="211"
                elif self.gpu1cbutton2.isChecked()==True:
                    self.runmode="212"
            try:
            #    print  self.runmode
                self.askfordemfile("live")
            except AttributeError:
                self.showerrordialog("Check your settings again")
        elif self.conventionalbutton.isChecked(): 
            self.runmode="1"
            self.askfordemfile("conventional")
            
    def askfordemfile(self,e):
        if e=="live":
            self.Tab.addTab(self.scrolllive,"Real-time barcoding")
            self.startbutton.setParent(None)
            self.Tab.setCurrentIndex(2)
        elif e=="conventional":
            self.Tab.addTab(self.scrollconventional,"Conventional barcoding")
            self.startbutton.setParent(None)
            self.Tab.setCurrentIndex(2)           
    def livebarcodedialog221(self):
        
        self.livestack.setCurrentIndex(1)
        self.flag=False
        try:
            self.livedemfile                       
            self.flag='live'
        except AttributeError:
            self.showerrordialog("You forgot at least one input")
        if self.flag!=False:
            self.d=QtWidgets.QDialog()
            layout=QtWidgets.QGridLayout()
            layout.setSpacing(10)
            demlabel=QtWidgets.QLabel('<b>Demultiplexing Settings</b>',self.d)
            ConsensusLabel=QtWidgets.QLabel('<b>Consensus settings</b>',self.d)       
            MinimumCoverageLabel=QtWidgets.QLabel('Minimum Coverage',self.d)
            LargeCoverageLabel=QtWidgets.QLabel('Large Coverage, or coverage beyond which you dont examine every increment',self.d)
            LargeCoverageStepLabel=QtWidgets.QLabel('Step size for larger coverages',self.d)
            MaximumCoverageLabel=QtWidgets.QLabel('Maximum Coverage, set 0 for no limit',self.d)
            Geneticcodelabel=QtWidgets.QLabel('Genetic Code: ',self.d)
            self.gencodebox=QtWidgets.QComboBox()
            self.gencodebox.addItem("1. The Standard Code")
            self.gencodebox.addItem("2. The Vertebrate Mitochondrial Code")
            self.gencodebox.addItem("3. The Yeast Mitochondrial Code")
            self.gencodebox.addItem("4. The Mold, Protozoan, and Coelenterate Mitochondrial Code")
            self.gencodebox.addItem("5. The Invertebrate Mitochondrial Code")
            self.gencodebox.addItem("6. The Ciliate, Dasycladacean and Hexamita Nuclear Code")
            self.gencodebox.addItem("9. The Echinoderm and Flatworm Mitochondrial Code")
            self.gencodebox.addItem("10. The Euplotid Nuclear Code")
            self.gencodebox.addItem("11. The Bacterial, Archaeal and Plant Plastid Code")
            self.gencodebox.addItem("12. The Alternative Yeast Nuclear Code")
            self.gencodebox.addItem("13. The Ascidian Mitochondrial Code")
            self.gencodebox.addItem("14. The Alternative Flatworm Mitochondrial Code")
            self.gencodebox.addItem("16. Chlorophycean Mitochondrial Code")
            self.gencodebox.addItem("21. Trematode Mitochondrial Code")
            self.gencodebox.addItem("22. Scenedesmus obliquus Mitochondrial Code")
            self.gencodebox.addItem("23. Thraustochytrium Mitochondrial Code")
            self.gencodebox.addItem("24. Rhabdopleuridae Mitochondrial Code")
            self.gencodebox.addItem("25. Candidate Division SR1 and Gracilibacteria Code")
            self.gencodebox.addItem("26. Pachysolen tannophilus Nuclear Code")
            self.gencodebox.addItem("27. Karyorelict Nuclear Code")
            self.gencodebox.addItem("28. Condylostoma Nuclear Code")
            self.gencodebox.addItem("29. Mesodinium Nuclear Code")
            self.gencodebox.addItem("30. Peritrich Nuclear Code")
            self.gencodebox.addItem("31. Blastocrithidia Nuclear Code")
            self.gencodebox.addItem("33. Cephalodiscidae Mitochondrial UAA-Tyr Code")
            self.gencodebox.activated[str].connect(self.gencodechoice)
            self.gencode=5
            self.gencodebox.setCurrentIndex(4)
            MinLengthLabel = QtWidgets.QLabel('Minimum Length: ',self.d)
            ExpLengthLabel = QtWidgets.QLabel('Length of barcode: ',self.d)
            demlenlab=QtWidgets.QLabel('Window to define product length (length of barcode +/-): ',self.d)
            primermismatchlab=QtWidgets.QLabel('Number of mismatches for primer search: ',self.d)
            primersearchlab=QtWidgets.QLabel('Window for primer and tag search: ',self.d)
      #      postdemlenlab=QtWidgets.QLabel('Maximum deviation of read length from barcode length: ',self.d2)
            consensesfreqfixedlabel=QtWidgets.QLabel('Main consensus calling frequency: ',self.d)
            consensesfreqrangelabel=QtWidgets.QLabel('Range of frequencies to assess: ',self.d)
            consensesfreqsteplabel=QtWidgets.QLabel('to be examined at step size of ',self.d)
            self.MinLengthTextBox = QtWidgets.QLineEdit()
            self.MinLengthTextBox.setText("658")
            self.ExpLengthTextBox = QtWidgets.QLineEdit()
            self.ExpLengthTextBox.setText("658")
            self.PrimerSearchTextBox = QtWidgets.QLineEdit()
            self.PrimerSearchTextBox.setText("100")
            self.PostdemLengthTextBox = QtWidgets.QLineEdit()
            self.PostdemLengthTextBox.setText("50")
            self.DemLengthTextBox = QtWidgets.QLineEdit()
            self.DemLengthTextBox.setText("100")
            tagmismatchlabel = QtWidgets.QLabel('Tag Mismatch: ',self.d)
            self.combobox1 =  QtWidgets.QComboBox()
            self.combobox1.addItem('0')
            self.combobox1.addItem('1')
            self.combobox1.addItem('2')
            self.combobox1.setCurrentText('2')
            self.combobox1.currentTextChanged.connect(self.tagmismatchchanged)
            self.PrimersearchMismatchTextBox = QtWidgets.QLineEdit()
            self.PrimersearchMismatchTextBox.setText("10")
            self.MinimumCoverageTextBox = QtWidgets.QLineEdit()
            self.MinimumCoverageTextBox.setText("20")
            self.LargeCoverageTextBox = QtWidgets.QLineEdit()
            self.LargeCoverageTextBox.setText("100")
            self.LargeCoverageStepTextBox = QtWidgets.QLineEdit()
            self.LargeCoverageStepTextBox.setText("10")
            self.MaximumCoverageTextBox = QtWidgets.QLineEdit()
            self.MaximumCoverageTextBox.setText("0")
            self.subsetmaxcheckbox = QtWidgets.QCheckBox("Subset higher coverage to Large coverage based on read lengths?")
            self.subsetmaxcheckbox.setChecked(True)
            self.subsetmaxcheckbox.stateChanged.connect(self.subsetornot)
            self.consensesfreqfixedTextbox= QtWidgets.QLineEdit()
            self.consensesfreqrangeTextbox= QtWidgets.QLineEdit()
            self.consensesfreqstepTextbox= QtWidgets.QLineEdit()
            self.consensesfreqfixedTextbox.setText("0.3")
            self.consensesfreqrangeTextbox.setText("0.2,0.5")
            self.consensesfreqstepTextbox.setText("0.05")
      #      self.basecallingcheckbox=QtWidgets.QCheckBox("Do you want to basecall as well as do barcode calling?")
        #    self.basecallingcheckbox.stateChanged.connect(lambda:self.stepcheck(self.basecallingcheckbox))
        #    self.basecallingcheckbox.setChecked(False)
       #     self.basecallingstat=False

           
            self.tagmm=2
            self.subsetmax=True
            self.minlen=658
            self.explen=658
            self.postdemlen=50
            self.mincoverage=20
            self.maxcoverage=0
            self.largecoverage=100
            self.largecoveragestep=10
            self.demlen=100
            self.primermismatch=10
            self.primersearchlen=100
            self.consensesfreqfixed=0.3
            self.consensesfreqrange=[0.2,0.5]
            self.consensesfreqstep=0.05
            regex = QtCore.QRegExp("^[0-9]*$")
            validator = QtGui.QRegExpValidator(regex,self.MinLengthTextBox)
            if self.runmode=="212":
                self.liveornot="basecall"
                self.procbutton=QtWidgets.QPushButton("Proceed",self.d)
                self.procbutton.clicked.connect(self.livebasecallingsoftwareparams)
                self.remotetransferstat=False
         
            elif self.runmode=="211":
                self.procbutton=QtWidgets.QPushButton("Proceed",self.d)
                self.procbutton.clicked.connect(self.setdemoutpath)
                self.remotetransferstat=False
            elif self.runmode=="222":
                self.liveornot="basecall"
                self.procbutton=QtWidgets.QPushButton("Proceed",self.d)
                self.procbutton.clicked.connect(self.livebasecallingsoftwareparams)
                self.remotetransferstat=True
                self.remotetransferlabel=QtWidgets.QLabel("Please start your run before entering remote path, don't worry if there is a lag, ONTbarcoder will catch up")
           #     self.remotetransfercheckbox.stateChanged.connect(lambda:self.stepcheck(self.remotetransfercheckbox))
           #     self.remotetransfercheckbox.setChecked(False)
          #      self.remotetransferstat=False
                hostlabel=QtWidgets.QLabel('Hostname',self.d)
                usernamelabel=QtWidgets.QLabel('Username',self.d)       
                passwordlabel=QtWidgets.QLabel('Password',self.d)
                remotepathlabel=QtWidgets.QLabel('Path to remote directory',self.d)            
                
                self.HostTextBox = QtWidgets.QLineEdit()
                self.HostTextBox.setText("mc-XXXXXX.local")
                self.UsernameTextBox = QtWidgets.QLineEdit()
                self.UsernameTextBox.setText("")
                self.PasswordTextBox = QtWidgets.QLineEdit()
                self.PasswordTextBox.setText("")
                self.RemotepathTextbox = QtWidgets.QLineEdit()
                self.RemotepathTextbox.setText("")                       
            elif self.runmode=="221":
                self.procbutton=QtWidgets.QPushButton("Proceed",self.d)
                self.procbutton.clicked.connect(self.setdemoutpath)                   
                self.remotetransferstat=True
                self.remotetransferlabel=QtWidgets.QLabel("Please start your run before entering remote path, don't worry if there is a lag, ONTbarcoder will catch up")
           #     self.remotetransfercheckbox.stateChanged.connect(lambda:self.stepcheck(self.remotetransfercheckbox))
           #     self.remotetransfercheckbox.setChecked(False)
          #      self.remotetransferstat=False
                hostlabel=QtWidgets.QLabel('Hostname',self.d)
                usernamelabel=QtWidgets.QLabel('Username',self.d)       
                passwordlabel=QtWidgets.QLabel('Password',self.d)
                remotepathlabel=QtWidgets.QLabel('Path to remote directory',self.d)            
                
                self.HostTextBox = QtWidgets.QLineEdit()
                self.HostTextBox.setText("mc-XXXXXX.local")
                self.UsernameTextBox = QtWidgets.QLineEdit()
                self.UsernameTextBox.setText("")
                self.PasswordTextBox = QtWidgets.QLineEdit()
                self.PasswordTextBox.setText("")
                self.RemotepathTextbox = QtWidgets.QLineEdit()
                self.RemotepathTextbox.setText("")                       
            layout.addWidget(demlabel,0,0,1,1)
            layout.addWidget(MinLengthLabel,1,0,1,2)
            layout.addWidget(self.MinLengthTextBox,1,2,1,1)
            layout.addWidget(ExpLengthLabel,2,0,1,2)
            layout.addWidget(self.ExpLengthTextBox,2,2,1,1)
            layout.addWidget(demlenlab,3,0,1,2)
            layout.addWidget(self.DemLengthTextBox,3,2,1,1)

            layout.addWidget(primersearchlab,4,0,1,1)
            layout.addWidget(self.PrimerSearchTextBox,4,2,1,1)
            layout.addWidget(primermismatchlab,4,3,1,1)
            layout.addWidget(self.PrimersearchMismatchTextBox,4,4,1,1)              
            layout.addWidget(tagmismatchlabel,5,0,1,2) 
            layout.addWidget(self.combobox1,5,2,1,1)  
            layout.addWidget(ConsensusLabel,6,0,1,1)
            layout.addWidget(MinimumCoverageLabel,7,0,1,2)
            layout.addWidget(self.MinimumCoverageTextBox,7,2,1,1)
            layout.addWidget(LargeCoverageLabel,8,0,1,2)
            layout.addWidget(self.LargeCoverageTextBox,8,2,1,1)     
            layout.addWidget(LargeCoverageStepLabel,9,0,1,2)
            layout.addWidget(self.LargeCoverageStepTextBox,9,2,1,1)
            layout.addWidget(MaximumCoverageLabel,10,0,1,2)
            layout.addWidget(self.MaximumCoverageTextBox,10,2,1,1)
            layout.addWidget(self.subsetmaxcheckbox,11,0,1,1)          
            layout.addWidget(consensesfreqfixedlabel,12,0,1,2)
            layout.addWidget(self.consensesfreqfixedTextbox,12,2)
            layout.addWidget(consensesfreqrangelabel,13,0,1,2)
            layout.addWidget(self.consensesfreqrangeTextbox,13,2)
            layout.addWidget(consensesfreqsteplabel,14,0,1,2)
            layout.addWidget(self.consensesfreqstepTextbox,14,2)
            layout.addWidget(Geneticcodelabel,15,0,1,2)
            layout.addWidget(self.gencodebox,15,2,1,2)
            if self.runmode=="211":
                layout.addWidget(self.procbutton,16,4)
            elif self.runmode=="212":
                layout.addWidget(self.procbutton,16,4)
            elif self.runmode=="221":
                layout.addWidget(self.remotetransferlabel,16,0,1,4)
                layout.addWidget(hostlabel,17,0,1,2)
                layout.addWidget(self.HostTextBox,17,2,1,1)     
                layout.addWidget(usernamelabel,18,0,1,2)
                layout.addWidget(self.UsernameTextBox,18,2,1,1)
                layout.addWidget(passwordlabel,19,0,1,2)
                layout.addWidget(self.PasswordTextBox,19,2,1,1)
                layout.addWidget(remotepathlabel,20,0,1,1)          
                layout.addWidget(self.RemotepathTextbox,20,2,1,1)
                layout.addWidget(self.procbutton,21,4)
            elif self.runmode=="222":
                layout.addWidget(self.remotetransferlabel,16,0,1,4)
                layout.addWidget(hostlabel,17,0,1,2)
                layout.addWidget(self.HostTextBox,17,2,1,1)     
                layout.addWidget(usernamelabel,18,0,1,2)
                layout.addWidget(self.UsernameTextBox,18,2,1,1)
                layout.addWidget(passwordlabel,19,0,1,2)
                layout.addWidget(self.PasswordTextBox,19,2,1,1)
                layout.addWidget(remotepathlabel,20,0,1,1)          
                layout.addWidget(self.RemotepathTextbox,20,2,1,1)
                layout.addWidget(self.procbutton,21,4)
            self.d.setLayout(layout)
            self.d.exec_()
    def getlivedir(self):
        self.mk1bdirbtn=DragPushbtn()
        self.mk1bdirbtn.mk1bpushbtn.clicked.connect(self.setuplive)
        self.mk1bdirbtn.exec_()
    #    self.mk1bdirbtn.setAutoFillBackground(True)

    def setconventionalmode(self):
     #   print "conventional"#
        if self.conventionalbutton.isChecked()==True:
            if self.realtimebutton.isChecked()==True:
                self.realtimebutton.setChecked(False)
                self.gpu1cbutton2.setChecked(False)
                self.gpu1cbutton1.setChecked(False)
                self.mk1cbutton.setChecked(False)
                self.mk1bbutton.setChecked(False)
    def showgpuopt(self,device):
     #   print "gpu",device#
        self.gputitlewidget.setStyleSheet("""QWidget {background-color: qlineargradient(spread:pad, x1:0 y1:0, x2:1 y2:0, stop:0 rgb(252, 185, 62), stop:1 white);}""")        
        try:
            self.emptywidget6.setParent(None)
            self.gputitlelayout.addWidget(self.gputitle,0,QtCore.Qt.AlignCenter)
            self.gputitle.setAttribute(QtCore.Qt.WA_TranslucentBackground)   
        except AttributeError:
            pass
        try:
            self.emptywidget4.setParent(None)
            
        except AttributeError:
            pass
        try:
            self.emptywidget3.setParent(None)
            
        except AttributeError:
            pass
        self.gpubuttonlayout.addWidget(self.gpu1cbutton1)
        self.gpubuttonlayout.addWidget(self.gpu1cbutton2)
        if device=="1b":
            if self.mk1bbutton.isChecked()==True:
                self.conventionalbutton.setChecked(False)
                self.mk1cbutton.setChecked(False)

                self.realtimebutton.setChecked(True)


        if device=="1c":
            if self.mk1cbutton.isChecked()==True:
 
                self.conventionalbutton.setChecked(False)
                self.mk1bbutton.setChecked(False)

                self.realtimebutton.setChecked(True)

    def showdeviceopt(self):
     #   print "device"
        if self.realtimebutton.isChecked()==True:

            self.conventionalbutton.setChecked(False)
            try:
                self.devicetitlewidget.setStyleSheet("""QWidget {background-color: qlineargradient(spread:pad, x1:0 y1:0, x2:1 y2:0, stop:0 grey, stop:1 white);}""")
                self.devicebuttonlayout.removeWidget(self.emptywidget1)
                self.devicebuttonlayout.removeWidget(self.emptywidget2)
                self.devicetitlelayout.removeWidget(self.emptywidget5)
                self.devicetitlelayout.addWidget(self.devicetitle,0,QtCore.Qt.AlignCenter)
                self.devicetitle.setAttribute(QtCore.Qt.WA_TranslucentBackground)
                self.devicebuttonlayout.addWidget(self.mk1bbutton)
                self.devicebuttonlayout.addWidget(self.mk1cbutton)
            except AttributeError:
                pass

        if self.realtimebutton.isChecked()==False:
            self.devicetitlewidget.setStyleSheet("""QWidget {background-color: qlineargradient(spread:pad, x1:0 y1:0, x2:1 y2:0, stop:0 white, stop:1 white);}""")
            self.gputitlewidget.setStyleSheet("""QWidget {background-color: qlineargradient(spread:pad, x1:0 y1:0, x2:1 y2:0, stop:0 white, stop:1 white);}""")
            self.gpu1cbutton2.setChecked(False)
            self.gpu1cbutton1.setChecked(False)
            self.mk1cbutton.setChecked(False)
            self.mk1bbutton.setChecked(False)

#
    def setlivemode(self):
        self.liveornot='live'
        self.livebarcodedialog221()
    def setpostmode(self):
        self.liveornot='post'
        self.getparams()
    def quit(self):
        self.close()
    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            e.accept()
        else:
            e.ignore()
    def checkindirstart(self):
        self.mycheckindir.start()
    def indircheckcomplete(self):
    #   print "checkdone",self.mycheckindir.indirstatus
        if  self.mycheckindir.indirstatus==False:
            self.opttable.item(1,0).setText("Your input directory seems to contain invalid files. Please reload correct files, else software will crash")
            self.opttable.item(1,0).setBackground(QtGui.QColor(255,0,0))
    def dropEvent(self, e):
        urls = e.mimeData().urls()
        if self.Tab.currentIndex()==2:
            if self.runmode=="1":
                if len(urls)==1:
                    fname=str(urls[0].path())
                 #   print fname

                    if "/" == str(urls[0].path())[0]:
                        fname=str(urls[0].path())[1:]
                        if os.path.isdir(fname)==True:
                            self.indir=fname
                            try:
                                del self.infastq
                            except AttributeError:
                                pass
                            self.opttable.item(1,0).setText("Input is demultiplexed reads "+fname+". Please proceed for barcode calling")
                            self.opttable.item(1,0).setBackground(QtGui.QColor(50,205,50))
                            self.opttable.item(0,0).setBackground(QtGui.QColor(255,255,255))
                            self.opttable.item(2,0).setBackground(QtGui.QColor(255,255,255))

                            self.mycheckindir=checkindir(self.indir)
                            self.mycheckindir.taskFinished.connect(self.indircheckcomplete)

                            self.checkindirstart()
                        else:
                            with open(fname,'rb') as infile:
                                line=infile.readline().decode("utf-8-sig")
                                if line[0]=="@":
                                    self.infastq=fname
                                    try:
                                        del self.forfixfname
                                    except AttributeError:
                                        pass
                                    try:
                                        self.demfile
                                        self.opttable.item(0,0).setText("You have provided both FASTQ and demultiplexing file, you can proceed")
                                        self.opttable.item(0,0).setBackground(QtGui.QColor(50,205,50))
                                        self.opttable.item(1,0).setBackground(QtGui.QColor(255,255,255))
                                        self.opttable.item(2,0).setBackground(QtGui.QColor(255,255,255))
                                        self.inputmode=1
                                    except AttributeError:
                                        self.opttable.item(0,0).setText("You have dragged a Fastq, but a demultiplexing file remains")
                                        self.opttable.item(0,0).setBackground(QtGui.QColor(50,205,50))
                                        self.opttable.item(1,0).setBackground(QtGui.QColor(255,255,255))
                                        self.opttable.item(2,0).setBackground(QtGui.QColor(255,255,255))
                                elif len(line.split(","))==5:
                                    self.demfile=fname
                                #   self.opttable.item(0,0).setText(fname)
                                    try:
                                        self.infastq
                                        self.opttable.item(0,0).setText("You have provided both FASTQ and demultiplexing file, you can proceed")
                                        self.opttable.item(0,0).setBackground(QtGui.QColor(50,205,50))
                                        self.opttable.item(1,0).setBackground(QtGui.QColor(255,255,255))
                                        self.opttable.item(2,0).setBackground(QtGui.QColor(255,255,255))
                                    except AttributeError:
                                        self.opttable.item(0,0).setText("You have dragged a demultiplexing file, but a FASTQ file remains")
                                        self.opttable.item(0,0).setBackground(QtGui.QColor(50,205,50))
                                        self.opttable.item(1,0).setBackground(QtGui.QColor(255,255,255))
                                        self.opttable.item(2,0).setBackground(QtGui.QColor(255,255,255))
                                elif line[0]==">":
                                    self.forfixfname=fname
                                    self.opttable.item(2,0).setText("You have hopefully provided a FASTA of consensus for fixing based on MSA, please proceed if so")
                                    self.opttable.item(2,0).setBackground(QtGui.QColor(50,205,50))
                                    self.opttable.item(1,0).setBackground(QtGui.QColor(255,255,255))
                                    self.opttable.item(0,0).setBackground(QtGui.QColor(255,255,255))
                                    self.inputmode=3
                                    self.flag=3
                                else:
                                    self.showerrordialog("This file cannot be recognized")
                if len(urls)==2:
                    for name in urls:
                        fname=str(name.path())
                        if "/" == fname[0]:
                            fname=fname[1:]
                        if os.path.isdir(fname)==True:
                            self.indir=fname
                            self.mycheckindir=checkindir(self.indir)
                            self.mycheckindir.taskFinished.connect(self.indircheckcomplete)
                            self.checkindirstart()
                        else:
                            with open(fname,'rb') as infile:
                                line=infile.readline().decode("utf-8-sig")
                                if line[0]=="@":
                                    self.infastq=fname
                                    self.opttable.item(0,0).setText("You have provided both FASTQ and demultiplexing file, you can proceed")
                                    self.opttable.item(0,0).setBackground(QtGui.QColor(50,205,50))
                                    self.opttable.item(1,0).setBackground(QtGui.QColor(255,255,255))
                                    self.opttable.item(2,0).setBackground(QtGui.QColor(255,255,255))
                                elif len(line.split(","))==5:
                                    self.demfile=fname
                                    self.opttable.item(0,0).setText("You have provided both FASTQ and demultiplexing file, you can proceed")
                                    self.opttable.item(0,0).setBackground(QtGui.QColor(50,205,50))
                                    self.opttable.item(1,0).setBackground(QtGui.QColor(255,255,255))
                                    self.opttable.item(2,0).setBackground(QtGui.QColor(255,255,255))
                                else:   
                                    self.showerrordialog("At least one of your files is wrong")
                if len(urls)>2:
                    self.showerrordialog("You can load only two files")
            else:
                if len(urls)==1:
                    for name in urls:
                        fname=str(name.path())
                        if "/" == fname[0]:
                            fname=fname[1:]
                        if os.path.isdir(fname)==True:
                            pass
                        else:
                            self.livedemfile=fname
                    self.livebtn.setEnabled(True)
                    self.livebtn.setStyleSheet("background-color: #0d6efd;")
        if self.Tab.currentIndex()==1:
            if len(urls)>=2:
                self.compfilelist=[]
                for name in urls:
                    fname=str(name.path())
                    if "/" == fname[0]:
                        fname=fname[1:]
                    self.compfilelist.append(fname)
            #   print urls
                self.setcompoutpath()
    def setcompoutpath(self):
        self.messagebox2("By default, files will be saved in current directory under name ComparisonOut+timestamp. Do you want to choose another folder?")
    def setcompoutpath2(self):
        self.compoutpath = QtWidgets.QFileDialog.getExistingDirectory(self, "Select Directory for Output files")
        if self.compoutpath!='':
            dirlist=os.listdir(self.compoutpath)
            if " " not in self.compoutpath:
                if len(dirlist)==0:
                    self.showrefdalog()
                else:
                    self.messagebox("Please select an empty output folder. This pipeline generates many files, and needs a clean folder to start with")
                    self.setcompoutpath2()
            else:
                self.messagebox("Please select an empty output folder without space in name")
                self.setcompoutpath2()              
        else:
            self.setcompoutpath2()
    def setcompoutpath3(self):
        ts = int(time.time())
        os.mkdir("ONTBarcoderComparisonOut"+str(ts))
        self.compoutpath=os.path.join(os.getcwd(),"ONTBarcoderComparisonOut"+str(ts))
        self.showrefdalog()
    
    def setdemoutpath(self):
        try:         
            self.d.close()
        except:
            pass
        try:         
            self.livebasecalldialog.close()
        except:
            pass
        try:
            self.livedoradobasecalldialog.close()
        except:
            pass    #   self.statuswidget=PbarWindow()
    #   self.sumstack.addWidget(self.statuswidget)
    #   self.sumstack.setCurrentIndex(1)
        self.messagebox3("By default, files will be saved in current directory under name Output+timestamp. Do you want to choose another folder?")

    def setdemoutpath2(self):   
        self.mb.close()               
                               
        self.demoutpath = QtWidgets.QFileDialog.getExistingDirectory(self, "Select Directory for Output files")
        if self.demoutpath !='':
            dirlist=os.listdir(self.demoutpath)
            if " " not in self.demoutpath:                                                              
                if len(dirlist)==0:
                    if self.liveornot=='post':
                        self.setuprun()
                    else:
                        if self.remotetransferstat==True:
                            self.setuplive()
                        else:
                            self.getlivedir()
                else:
                    self.messagebox("Please select an empty output folder. This pipeline generates many files, and needs a clean folder to start with")
                    self.setdemoutpath2()
            else:
                self.messagebox("Please select an empty output folder without space in name")
                self.setdemoutpath2()               
                                   
        else:
            self.sumstack.setCurrentIndex(0)

    def setdemoutpath3(self):
        self.mb.close()                                                      
        ts = int(time.time())
        os.mkdir("ONTBarcoderOutput"+str(ts))
        self.demoutpath=os.path.join(os.getcwd(),"ONTBarcoderOutput"+str(ts))
        if self.liveornot=='post':
            self.setuprun()
        else:
            if self.remotetransferstat==True:
                self.setuplive()
            else:
                self.getlivedir()
    def setuplive(self):
        valuecheck=[]
        try:
            self.demlen=int(self.DemLengthTextBox.text())
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)
        try:
            self.postdemlen=int(self.PostdemLengthTextBox.text())
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)
        try:
            self.primersearchlen=int(self.PrimerSearchTextBox.text())
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)            
        try:
            self.primermismatch=int(self.PrimersearchMismatchTextBox.text())
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)

        try:
            self.minlen=int(self.MinLengthTextBox.text())
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)
        try:
            self.explen=int(self.ExpLengthTextBox.text())
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)

        try:
            self.mincoverage=int(str(self.MinimumCoverageTextBox.text()))
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)
        try:
            self.largecoverage=int(str(self.LargeCoverageTextBox.text()))
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)
        try:
            self.largecoveragestep=int(str(self.LargeCoverageStepTextBox.text()))
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)            
        try:
            self.maxcoverage=int(str(self.MaximumCoverageTextBox.text()))
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)
        try:
            self.consensesfreqfixed=float(self.consensesfreqfixedTextbox.text())
            if self.consensesfreqfixed<0:
                self.showwarningdialog("The values for consensus frequency should be in range 0-1")
                valuecheck.append(False)
            elif self.consensesfreqfixed>1:
                self.showwarningdialog("The values for consensus frequency  should be in range 0-1")
                self.runstat=False
            else:
                valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values for consensus frequency must be a floating point value")
        if str(self.consensesfreqrangeTextbox.text())!="":
            try:
                self.consensesfreqrange=[float(x) for x in str(self.consensesfreqrangeTextbox.text()).split(",")]
                
                if len(self.consensesfreqrange)!=2:
                    self.showwarningdialog("Two values need to be specified for the range of consensus frequencies")
                    valuecheck.append(False)
                stat1,stat2=False,False
                if self.consensesfreqrange[0]<0:
                    self.showwarningdialog("The values for consensus frequency ranges should be in range 0-1")
                    valuecheck.append(False)
                elif self.consensesfreqrange[0]>1:
                    self.showwarningdialog("The values for consensus frequency  ranges should be in range 0-1")
                    valuecheck.append(False)
                else:
                    stat1=True
                    
                if self.consensesfreqrange[1]<0:
                    self.showwarningdialog("The values for consensus frequency  ranges should be in range 0-1")
                    valuecheck.append(False)
                elif self.consensesfreqrange[1]>1:
                    self.showwarningdialog("The values for consensus frequency  ranges should be in range 0-1")
                    valuecheck.append(False)
                else:
                    stat2=True

                if stat1==True:
                    if stat2==True:
                        valuecheck.append(True)     

            except ValueError:
                self.showwarningdialog("The values for consensus frequency range should be floating point values")
                valuecheck.append(False)
        else:
            self.consensesfreqrange=[self.consensesfreqfixed,self.consensesfreqfixed]
            valuecheck.append(True)
            
                        
        try:    
            self.consensesfreqstep=float(self.consensesfreqstepTextbox.text())
            if self.consensesfreqstep<0:
                self.showwarningdialog("The values for consensus frequency  step size should be in range 0-1")
                valuecheck.append(False)
                                     
                               
            elif self.consensesfreqstep>1:
                self.showwarningdialog("The values for consensus frequency  step size should be in range 0-1")
                valuecheck.append(False)
            else: 
                valuecheck.append(True)
                    
        except ValueError:
            self.showwarningdialog("The values for consensus frequency step size must be a floating point value")
            valuecheck.append(False)
       # print valuecheck
        if self.remotetransferstat==True:
            try:    
                self.hostname=str(self.HostTextBox.text())
                if len(self.hostname)==0:
                    self.showwarningdialog("device name should not be empty")
                else: 
                    valuecheck.append(True)
            except ValueError:
                self.showwarningdialog("Something wrong with remote transfer setting")
                valuecheck.append(False)
            try:    
                self.username=str(self.UsernameTextBox.text())
                if len(self.username)==0:
                    self.showwarningdialog("device name should not be empty")
                else: 
                    valuecheck.append(True)
            except ValueError:
                self.showwarningdialog("Something wrong with remote transfer setting")
                valuecheck.append(False)    
            try:    
                self.password=str(self.PasswordTextBox.text())
                if len(self.password)==0:
                    self.showwarningdialog("device name should not be empty")
                else: 
                    valuecheck.append(True)
            except ValueError:
                self.showwarningdialog("Something wrong with remote transfer setting")
                valuecheck.append(False)    
            try:    
                self.remotepath=str(self.RemotepathTextbox.text())
                if len(self.remotepath)==0:
                    self.showwarningdialog("device name should not be empty")
                else: 
                    valuecheck.append(True)
            except ValueError:
                self.showwarningdialog("Something wrong with remote transfer setting")
                valuecheck.append(False)
             
        if self.liveornot=='basecall':
            if self.basecallertype=="Guppy":
                try:    
                    self.devicename=str(self.DeviceTextBox.text())
                    if len(self.devicename)==0:
                        self.showwarningdialog("device name should not be empty")
                    else: 
                        valuecheck.append(True)
                except ValueError:
                    self.showwarningdialog("The values for consensus frequency step size must be a floating point value")
                    valuecheck.append(False)
                try:    
                    self.numcallers=str(self.NumCallersTextBox.text())
                    if int(self.numcallers)<1:
                        self.showwarningdialog("The values for consensus frequency  step size should be over 1")
                    else: 
                        valuecheck.append(True)
                except ValueError:
                    self.showwarningdialog("The values for consensus frequency step size must be a floating point value")
                    valuecheck.append(False)
                  
                try:    
                    self.gpucallers=str(self.GPURunnerTextBox.text())
                    if int(self.gpucallers)<1:
                        self.showwarningdialog("The values for consensus frequency  step size should be over 1")
                    else: 
                        valuecheck.append(True)
                except ValueError:
                    self.showwarningdialog("The values for consensus frequency step size must be a floating point value")
                    valuecheck.append(False)
                try:    
                    self.chunksperrunner=str(self.ChunksPerRunnerTextBox.text())
                    if int(self.chunksperrunner)<1:
                        self.showwarningdialog("The values for consensus frequency  step size should be over 1")
                    else: 
                        valuecheck.append(True)
                except ValueError:
                    self.showwarningdialog("The values for consensus frequency step size must be a floating point value")
                    valuecheck.append(False)
            elif self.basecallertype=="Dorado":
                try:    
                    self.devicename=str(self.DeviceTextBox2.text())
                    if len(self.devicename)==0:
                        self.showwarningdialog("device name should not be empty")
                    else: 
                        valuecheck.append(True)
                except ValueError:
                    self.showwarningdialog("The values for consensus frequency step size must be a floating point value")
                    valuecheck.append(False)
                try:    
                    self.chunksnum=str(self.ChunkTextBox.text())
                    if len(self.chunksnum)==0:
                        self.showwarningdialog("device name should not be empty")
                    else: 
                        valuecheck.append(True)
                except ValueError:
                    self.showwarningdialog("The values for consensus frequency step size must be a floating point value")
                    valuecheck.append(False)
                try:    
                    self.doradooverlap=str(self.OverlapTextBox.text())
                    if len(self.doradooverlap)==0:
                        self.showwarningdialog("device name should not be empty")
                    else: 
                        valuecheck.append(True)
                except ValueError:
                    self.showwarningdialog("The values for consensus frequency step size must be a floating point value")
                    valuecheck.append(False)                    


        os.mkdir(os.path.join(self.demoutpath,"livefastq"))
        if self.remotetransferstat==True:
            self.livedir=os.path.join(self.demoutpath,"livefastq")
        elif self.remotetransferstat==False:
            if self.liveornot=='basecall':
                if self.basecallertype=='Guppy':
                    self.inputfast5folder=self.mk1bdirbtn.livedir
                elif self.basecallertype=='Dorado':
                    self.inputpod5folder=self.mk1bdirbtn.livedir
                self.livedir=os.path.join(self.demoutpath,"livefastq")
             #   print "no remote, basecalling"
                
            else:
                self.livedir=self.mk1bdirbtn.livedir
            self.mk1bdirbtn.close()
        
        self.logfile=open(os.path.join(self.demoutpath,"log.txt"),'w')
        self.livestack.setCurrentIndex(1)
        self.myprepdemultiplexlive=prepdemultiplexlive(self.livedemfile,self.logfile,self.tagmm)
    #    print self.basecallertype,"basecallertype"
        if self.liveornot=='basecall':
            if self.basecallertype=='Guppy':
                os.mkdir(os.path.join(self.livedir,"temp"))
                os.mkdir(os.path.join(self.demoutpath,"livetransfers"))
                if self.remotetransferstat==True:
                    self.inputfast5folder=os.path.join(self.demoutpath,"livetransfers")
                with open(os.path.join(os.path.dirname(sys.argv[0]),"config","guppyconfig.txt"),'w') as guppyconfigfile:
                    guppyconfigfile.write("guppypath="+self.guppypath+'\n'+"configpath="+self.configpath)
                self.guppycommand=str(self.CmdLineTextBox.toPlainText()).replace("ignorethis",os.path.join(self.livedir,"temp")).replace("ignoreinputfolder",self.inputfast5folder)
            #    print(self.guppycommand)
                self.myprepdemultiplexlive.taskFinished.connect(self.setlivedetectorandguppy)
            if self.basecallertype=='Dorado':
                os.mkdir(os.path.join(self.livedir,"temp"))
                os.mkdir(os.path.join(self.demoutpath,"livetransfers"))
                if self.remotetransferstat==True:
                    self.inputpod5folder=os.path.join(self.demoutpath,"livetransfers")
                with open(os.path.join(os.path.dirname(sys.argv[0]),"config","doradoconfig.txt"),'w') as doradoconfigfile:
                    doradoconfigfile.write("doradopath="+self.doradopath+'\n'+"configpath="+self.doradoconfigpath)
                self.doradooutfastqpath=os.path.join(self.livedir,"temp")
                self.doradocommand=self.doradocommand.replace("ignoreinputfolder",os.path.join(self.inputpod5folder,"ignorethis"))
            #    print(self.doradocommand)
                self.myprepdemultiplexlive.taskFinished.connect(self.setlivedetectoranddorado)
        else:
            self.myprepdemultiplexlive.taskFinished.connect(self.setlivedetector)

        self.runprepdemultiplexlive()
    def readremotedir(self,remotepath,client):
        stdin, stdout, stderr = client.exec_command('ls '+remotepath)
        newlist=[]
        for line in stdout:
            newlist.append(line.encode(encoding="utf-8").strip())
     #   print newlist
        return newlist        
    def runprepdemultiplexlive(self):
        self.myprepdemultiplexlive.start()                                     
    def setlivedetectorandguppy(self):
        self.iternum=0
        os.mkdir(os.path.join(self.demoutpath,"Demultiplexed"))
        os.mkdir(os.path.join(self.demoutpath,"Subsets"))
        os.mkdir(os.path.join(self.demoutpath,"Aligned"))
        os.mkdir(os.path.join(self.demoutpath,"Consensus"))
        os.mkdir(os.path.join(self.demoutpath,"Processing"))
        self.sampleids=self.myprepdemultiplexlive.sampleids
        self.totalseqs=0
        self.donelist=[]
        self.donefast5=[]
        self.ndemultiplexed=0
        self.ndemultiplexedused=0    
        self.ndone=0
        self.mylivedetector=livedetector(self.livedir)
        self.mylivedetector.notifyProgress.connect(self.adddonefiles)
        self.mylivesamplecount=livesamplecount(self.livedir,self.sampleids,self.mincoverage,self.largecoverage,self.largecoveragestep,self.maxcoverage)
        self.mylivesamplecount.notifyProgress.connect(self.updatelivesamplecounter)
        self.mylivesamplecount.taskFinished.connect(self.killlivesamplecount)
        self.mylivefast5detector=livefast5detector(self.inputfast5folder)
        self.mylivefast5detector.notifyProgress.connect(self.addfast5torun)
    #    print "remote transfer",self.remotetransferstat
        if self.remotetransferstat==True:
            self.donetransfer=[]
            self.myliveremotedetector=liveremotedetector(self.remotepath,self.hostname,self.username,self.password)
            self.myliveremotedetector.notifyProgress.connect(self.addfiletotransfer)
        self.runlivedetector2()                       
    def setlivedetectoranddorado(self):
        self.iternum=0
        os.mkdir(os.path.join(self.demoutpath,"Demultiplexed"))
        os.mkdir(os.path.join(self.demoutpath,"Subsets"))
        os.mkdir(os.path.join(self.demoutpath,"Aligned"))
        os.mkdir(os.path.join(self.demoutpath,"Consensus"))
        os.mkdir(os.path.join(self.demoutpath,"Processing"))
        self.sampleids=self.myprepdemultiplexlive.sampleids
        self.totalseqs=0
        self.donelist=[]
        self.donepod5=[]
        self.ndemultiplexed=0
        self.ndemultiplexedused=0    
        self.ndone=0
        self.mylivedetector=livedetector(self.livedir)
        self.mylivedetector.notifyProgress.connect(self.adddonefiles)
        self.mylivesamplecount=livesamplecount(self.livedir,self.sampleids,self.mincoverage,self.largecoverage,self.largecoveragestep,self.maxcoverage)
        self.mylivesamplecount.notifyProgress.connect(self.updatelivesamplecounter)
        self.mylivesamplecount.taskFinished.connect(self.killlivesamplecount)
        self.mylivepod5detector=livepod5detector(self.inputpod5folder)
        self.mylivepod5detector.notifyProgress.connect(self.addpod5torun)
     #   print("remote transfer",self.remotetransferstat)
        if self.remotetransferstat==True:
            self.donetransfer=[]
            self.myliveremotedetector=liveremotedetector(self.remotepath,self.hostname,self.username,self.password)
            self.myliveremotedetector.notifyProgress.connect(self.addfiletotransfer)
        self.runlivedetector3()                                                               
    def setlivedetector(self):
        os.mkdir(os.path.join(self.demoutpath,"Demultiplexed"))
        os.mkdir(os.path.join(self.demoutpath,"Subsets"))
        os.mkdir(os.path.join(self.demoutpath,"Aligned"))
        os.mkdir(os.path.join(self.demoutpath,"Consensus"))
        os.mkdir(os.path.join(self.demoutpath,"Processing"))
        self.sampleids=self.myprepdemultiplexlive.sampleids
        self.totalseqs=0
        self.donelist=[]
        self.ndemultiplexed=0
        self.ndemultiplexedused=0
        self.ndone=0
        self.mylivedetector=livedetector(self.livedir)
        self.mylivedetector.notifyProgress.connect(self.adddonefiles)
        self.mylivesamplecount=livesamplecount(self.livedir,self.sampleids,self.mincoverage,self.largecoverage,self.largecoveragestep,self.maxcoverage)
        self.mylivesamplecount.notifyProgress.connect(self.updatelivesamplecounter)
        self.mylivesamplecount.taskFinished.connect(self.killlivesamplecount)
        if self.remotetransferstat==True:
            self.donetransfer=[]
            self.checkdifftransfer()
            
            
          #  self.myliveremotedetector=liveremotedetector(self.remotepath,self.client)
       #     self.myliveremotedetector.notifyProgress.connect(self.addfiletotransfer)
        self.runlivedetector()
    def killlivesamplecount(self):
    #   self.mylivesamplecount.stop()
    #   self.mylivesamplecount.quit()  
        self.previoussampleids=deepcopy(self.sampleids)
        self.myrunconsensuslive=runconsensuslive(self.demoutpath,self.mylivesamplecount.todolist,self.consensesfreqfixed,self.consensesfreqrange,self.consensesfreqstep,self.gencode,self.explen,self.ndone,self.largecoverage,self.subsetmax,self.sampleids)
        self.myrunconsensuslive.notifyMessage.connect(self.addbarcodelog)
        self.myrunconsensuslive.taskFinished.connect(self.restartlivesamplecount)
        self.runrunconsensuslive()
    def addbarcodelog(self,i):
    #    print "ndone",i
        timediff=datetime.datetime.fromtimestamp(int(time.time()))-self.timestamp
        minutes=timediff.total_seconds()/60 
        self.livebarcodegraphXm=np.append(self.livebarcodegraphXm,float(minutes)) 
        self.livebarcodegraphYm=np.append(self.livebarcodegraphYm,float(i))
     #   print self.livebarcodegraphYm,self.livebarcodegraphXm            # update x position for displaying the curve
        self.livebarcodegraph.setData(self.livebarcodegraphXm,self.livebarcodegraphYm)
        self.livebarcodegraph.setPos(0,0)                      # set the curve with this data
    #       QtGui.QApplication.processEvents()      
    def runrunconsensuslive(self):
        self.myrunconsensuslive.start()
    def restartlivesamplecount(self,l):
        donelist=l
        for each in donelist:
            del(self.sampleids[each])
            del(self.previoussampleids[each])
        self.ndone=self.myrunconsensuslive.ndone
     #   print "Round of consensus completed!",self.ndone
        self.mylivesamplecount.flag=True
    #    print "changing samplecount flag"
      #  self.mylivesamplecount=livesamplecount(self.livedir,self.previoussampleids,self.mincoverage,self.largecoverage,self.largecoveragestep,self.maxcoverage)
      #  self.mylivesamplecount.notifyProgress.connect(self.updatelivesamplecounter)
      #  self.mylivesamplecount.taskFinished.connect(self.killlivesamplecount)
     #   self.runlivesamplecount()
    def adddonefiles(self,i):
        inputfiles=self.mylivedetector.new_files
        self.donelist+=inputfiles
     #   print "input is", inputfiles
        self.setdemultiplexrunlive(inputfiles,i)
    def addfast5torun(self):
        inputfiles=self.mylivefast5detector.new_files
        self.donefast5+=inputfiles
        self.setguppyrunlive()
    def addpod5torun(self):
        self.pod5list=self.mylivepod5detector.new_files
    #    print "new pod5list", self.pod5list
        self.donepod5+=self.pod5list
        self.setdoradorunlive()
    def addfiletotransfer(self):
    #    print "added to list from remote"
        inputfiles=self.myliveremotedetector.new_files
        self.donetransfer+=inputfiles
        self.setremotetransfer(inputfiles)
    def runlivedetector(self):
        self.mylivedetector.start()
        self.mylivesamplecount.start()
    def runlivedetector2(self):
        self.mylivedetector.start()
        self.mylivesamplecount.start()
        self.mylivefast5detector.start()
        if self.remotetransferstat==True:
            self.myliveremotedetector.start()
    def runlivedetector3(self):
        self.mylivedetector.start()
        self.mylivesamplecount.start()
        self.mylivepod5detector.start()
        if self.remotetransferstat==True:
            self.myliveremotedetector.start()
    def runlivedetectoronly(self):
        self.mylivedetector.start()
    def runlivepod5detectoronly(self):
        self.mylivepod5detector.start()
    def runlivefast5detectoronly(self):
        self.mylivefast5detector.start()
    def runliveremotedetectoronly(self):
        self.myliveremotedetector.start()
    def runlivesamplecount(self):
        self.mylivesamplecount.start()
    def updatelivesamplecounter(self):
        oldsampleids=deepcopy(self.mylivesamplecount.localsampleids)    
  #      print self.sampleids
        self.mylivesamplecount.update_sampleids(self.sampleids)
    def setdemultiplexrunlive(self,inputfiles,i):
    #   self.nseqspasslen=self.myprepdemultiplexlive.nseqspasslen
                                                                 
     #   print self.flag
        sampledict=self.myprepdemultiplexlive.sampledict
    #    print "input2 is ",inputfiles
        self.livelogstatus.textCursor().insertHtml('<br>Read '+" ".join(inputfiles))
        tagdict=self.myprepdemultiplexlive.tagdict
        muttags_fr=self.myprepdemultiplexlive.muttags_fr
        typedict=self.myprepdemultiplexlive.typedict
        primerfset=self.myprepdemultiplexlive.primerfset
        primerrset=self.myprepdemultiplexlive.primerrset
        taglen=self.myprepdemultiplexlive.taglen
        primerlensum=self.myprepdemultiplexlive.primerlensum
        self.compressedmode=i
      #Builds dictionary of sequences from reformated file
        self.myrundemultiplexlive=rundemultiplexlive(self.livedir,inputfiles,self.demoutpath,tagdict,muttags_fr,sampledict,typedict,primerfset,primerrset,taglen,primerlensum,self.sampleids,self.totalseqs,self.ndemultiplexed,self.ndemultiplexedused,self.minlen,self.explen,self.demlen,self.primersearchlen,self.tagmm,i,self.primermismatch)
        self.myrundemultiplexlive.notifyMessage.connect(self.addloglive)
        self.myrundemultiplexlive.notifyMessage2.connect(self.addloglivedem)
        self.myrundemultiplexlive.taskFinished.connect(self.checkdiff)
        self.runrundemultiplexlive()
    def setdoradorunlive(self):
      # self.doradocommand=self.doradocommand.replace("dorado","./dorado")
      #Builds dictionary of sequences from reformated file
        self.mylivedoradorun=livedoradorun(self.doradocommand,self.demoutpath,self.livedir,self.iternum,self.pod5list,self.inputpod5folder,self.doradooutfastqpath)
        self.mylivedoradorun.taskFinished.connect(self.checkdiffpod5)
        self.runlivedoradorun()
    def setguppyrunlive(self):

      #Builds dictionary of sequences from reformated file
        self.myliveguppyrun=liveguppyrun(self.guppycommand,self.demoutpath,self.livedir,self.iternum)
        self.myliveguppyrun.taskFinished.connect(self.checkdifffast5)
        self.runliveguppyrun()
    def setremotetransfer(self,filelist):
    #    print "setting transfer", self.liveornot
        if self.liveornot=='basecall':
            if self.basecallertype=="Guppy":
                self.myliveremotetransfer=liveremotetransfer(self.remotepath,self.inputfast5folder,filelist,self.hostname,self.username,self.password)
            elif self.basecallertype=="Dorado":
                self.myliveremotetransfer=liveremotetransfer(self.remotepath,self.inputpod5folder,filelist,self.hostname,self.username,self.password)
        else:
            self.myliveremotetransfer=liveremotetransfer(self.remotepath,self.livedir,filelist,self.hostname,self.username,self.password)
        self.myliveremotetransfer.taskFinished.connect(self.checkdifftransfer)
        self.runliveremotetransfer()
    def runrundemultiplexlive(self):
        self.myrundemultiplexlive.start()
    def runlivedoradorun(self):
        self.mylivedoradorun.start()   
    def runliveguppyrun(self):
        self.myliveguppyrun.start()
    def runliveremotetransfer(self):
        self.myliveremotetransfer.start()
    def checkdiff(self):
        i=self.compressedmode
        oldlist=self.donelist
        newlist=os.listdir(self.livedir)
        newlist=[x for x in newlist if ".fastq" in x]
        self.donelist=newlist
        newlist2=[]
        for x in newlist:
            if x[0]!=".":
                newlist2.append(x)
        
        newfiles=list(set(newlist2)-set(oldlist))
      #  print "newfiles",newfiles
        if len(newfiles)>0:
            self.setdemultiplexrunlive(newfiles,i)
         #   print "I am behind",self.ndemultiplexed
        else:
         #   print "I am ahead",self.ndemultiplexed
            self.mylivedetector=livedetector(self.livedir)
            self.mylivedetector.notifyProgress.connect(self.adddonefiles)
            self.runlivedetectoronly()
    def checkdiffpod5(self):
        self.iternum+=1
        oldlist=self.donepod5
     #   print self.donepod5
        newlist=os.listdir(self.inputpod5folder)
        newlist=[x for x in newlist if x.endswith("pod5")]
        self.donepod5=newlist
        newlist2=[]
        for x in newlist:
            if x[0]!=".":
                newlist2.append(x)
        
        newfiles=list(set(newlist2)-set(oldlist))
        self.pod5list=newfiles
        if len(newfiles)>0:
        #    with open("inlistfast5.txt",'w') as outfile:
         #       outfile.write("\n".join(newfiles))
            self.setdoradorunlive()
        else:
            self.mylivepod5detector=livepod5detector(self.inputpod5folder)
            self.mylivepod5detector.notifyProgress.connect(self.addpod5torun)
            self.runlivepod5detectoronly()
    def checkdifffast5(self):
        self.iternum+=1
        oldlist=self.donefast5
        newlist=os.listdir(self.inputfast5folder)
        self.donefast5=newlist
        newlist2=[]
        for x in newlist:
            if x[0]!=".":
                newlist2.append(x)
        
        newfiles=list(set(newlist2)-set(oldlist))
     #   print "newfiles",newfiles
        if len(newfiles)>0:
            with open("inlistfast5.txt",'w') as outfile:
                outfile.write("\n".join(newfiles))
            self.setguppyrunlive()
         #   print "I am behind",self.ndemultiplexed
        else:
          #  print "I am ahead",self.ndemultiplexed
            self.mylivefast5detector=livefast5detector(self.inputfast5folder)
            self.mylivefast5detector.notifyProgress.connect(self.addfast5torun)
            self.runlivefast5detectoronly()
    def checkdifftransfer(self):
        oldlist=self.donetransfer
        try:
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(self.hostname, username=self.username, password=self.password) 
            if self.liveornot=="basecall":
                newlist=self.readremotedir(self.remotepath,client)
            else:
                newlist=self.readremotedir(self.remotepath,client)
            client.close()
            self.donetransfer=newlist
            newlist2=[]
            for x in newlist:
                if x[0]!=".":
                    newlist2.append(x)
            newfiles=list(set(newlist2)-set(oldlist))
        #    print "newremotefiles",newfiles
            if len(newfiles)>0:
                self.setremotetransfer(newfiles)
             #   print "I am behind",self.ndemultiplexed
            else:
             #   print "I am ahead",self.ndemultiplexed
                self.myliveremotedetector=liveremotedetector(self.remotepath,self.hostname,self.username,self.password)
                self.myliveremotedetector.notifyProgress.connect(self.addfiletotransfer)
                self.runliveremotedetectoronly()
        except Exception as e:
            print (e,"checkdifftransfer connection error")
            self.myliveremotedetector=liveremotedetector(self.remotepath,self.hostname,self.username,self.password)
            self.myliveremotedetector.notifyProgress.connect(self.addfiletotransfer)
            self.runliveremotedetectoronly()            
    def setuprun(self):
        self.logfile=open(os.path.join(self.demoutpath,"log.txt"),'w')                    

        os.mkdir(os.path.join(self.demoutpath,"barcodesets"))
        self.wb=xlsxwriter.Workbook(os.path.join(self.demoutpath,"runsummary.xlsx"))
        os.mkdir(os.path.join(self.demoutpath,"barcodesets","consensus_by_length"))
        os.mkdir(os.path.join(self.demoutpath,"barcodesets","consensus_by_similarity"))
        os.mkdir(os.path.join(self.demoutpath,"barcodesets","fixing"))
        os.mkdir(os.path.join(self.demoutpath,"barcodesets","temps"))
        os.mkdir(os.path.join(self.demoutpath,"demultiplexingfiles"))
        self.outpathdemfiles=os.path.join(self.demoutpath,"demultiplexingfiles")
        self.start_time=time.time()
        os.mkdir(os.path.join(self.demoutpath,"demultiplexed"))
        if self.flag==1:
            os.mkdir(os.path.join(self.demoutpath,"1_demultiplexing"))
        if self.flag==4:
            os.mkdir(os.path.join(self.demoutpath,"1_demultiplexing"))
        os.mkdir(os.path.join(self.demoutpath,"2a_ConsensusByLength"))
        os.mkdir(os.path.join(self.demoutpath,"2b_ConsensusBySimilarity"))
        os.mkdir(os.path.join(self.demoutpath,"3_ConsensusByBarcodeComparison"))
        self.runstat=False
        valuecheck=[]
        try:
            self.selectlens=[int(x) for x in self.randomsubsetnTextBox.text().split(",")]
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of coverages should be integers")
            valuecheck.append(False)
        try:
            self.primermismatch=int(self.PrimersearchMismatchTextBox.text())
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)    
        try:
            self.demlen=int(self.DemLengthTextBox.text())
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)
        try:
            self.postdemlen=int(self.PostdemLengthTextBox.text())
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)
        try:
            self.primersearchlen=int(self.PrimerSearchTextBox.text())
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)            
        
        try:
            self.minlen=int(self.MinLengthTextBox.text())
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)
        try:
            self.explen=int(self.ExpLengthTextBox.text())
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)
        try:
            
            self.n90percsubsetn=int(str(self.n90percsubsetnTextBox.text()))
            valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values of length cutoffs must be integers")
            valuecheck.append(False)
            

        try:
            self.consensesfreqfixed=float(self.consensesfreqfixedTextbox.text())
            if self.consensesfreqfixed<0:
                self.showwarningdialog("The values for consensus frequency should be in range 0-1")
                valuecheck.append(False)
            elif self.consensesfreqfixed>1:
                self.showwarningdialog("The values for consensus frequency  should be in range 0-1")
                self.runstat=False
            else:
                valuecheck.append(True)
        except ValueError:
            self.showwarningdialog("The values for consensus frequency must be a floating point value")
        if str(self.consensesfreqrangeTextbox.text())!="":
            try:
                self.consensesfreqrange=[float(x) for x in str(self.consensesfreqrangeTextbox.text()).split(",")]
                
                if len(self.consensesfreqrange)!=2:
                    self.showwarningdialog("Two values need to be specified for the range of consensus frequencies")
                    valuecheck.append(False)
                stat1,stat2=False,False
                if self.consensesfreqrange[0]<0:
                    self.showwarningdialog("The values for consensus frequency ranges should be in range 0-1")
                    valuecheck.append(False)
                elif self.consensesfreqrange[0]>1:
                    self.showwarningdialog("The values for consensus frequency  ranges should be in range 0-1")
                    valuecheck.append(False)
                else:
                    stat1=True
                    
                if self.consensesfreqrange[1]<0:
                    self.showwarningdialog("The values for consensus frequency  ranges should be in range 0-1")
                    valuecheck.append(False)
                elif self.consensesfreqrange[1]>1:
                    self.showwarningdialog("The values for consensus frequency  ranges should be in range 0-1")
                    valuecheck.append(False)
                else:
                    stat2=True

                if stat1==True:
                    if stat2==True:
                        valuecheck.append(True)     

            except ValueError:
                self.showwarningdialog("The values for consensus frequency range should be floating point values")
                valuecheck.append(False)
        else:
            self.consensesfreqrange=[self.consensesfreqfixed,self.consensesfreqfixed]
            valuecheck.append(True)
            
                        
        try:    
            self.consensesfreqstep=float(self.consensesfreqstepTextbox.text())
            if self.consensesfreqstep<0:
                self.showwarningdialog("The values for consensus frequency  step size should be in range 0-1")
                valuecheck.append(False)
                                     
                               
            elif self.consensesfreqstep>1:
                self.showwarningdialog("The values for consensus frequency  step size should be in range 0-1")
                valuecheck.append(False)
            else: 
                valuecheck.append(True)
                    
        except ValueError:
            self.showwarningdialog("The values for consensus frequency step size must be a floating point value")
            valuecheck.append(False)
      #  print valuecheck
        if False not in valuecheck:
            if self.flag==4:
                self.inputmode=4
                self.excelpath=self.indir
                self.indemfiles=os.listdir(self.indir)
                self.statuswidget.logstatus.textCursor().insertHtml("Input FASTQ: "+self.infastq+'<br>Input Demultiplexing file: '+self.excelpath+'<br><b><p style="color:blue;">Analysis setup</p><br> Settings for Phase 1 Demultiplexing:</b><br><br>')
                
                self.statuswidget.logstatus.textCursor().insertHtml('Minimum Length: '+str(self.minlen)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Length of barcode: '+str(self.explen)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Window to define product length (length of barcode +/-): '+str(self.demlen)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Window for primer and tag search: '+str(self.primersearchlen)+'<br>')
                if self.consensuslenstat==True:
                    self.logfile.write('Window to select reads of long length after demultiplexing: '+str(self.postdemlen)+'\n')
                    self.logfile.write('Coverage used for barcode calling based on subsetting reads by lengths: '+str(self.selectlens)+'\n')
                    self.statuswidget.logstatus.textCursor().insertHtml('<br><b>Settings for Phase 2a: Barcode calling with “Consensus by length”</b><br><br>') 
                    self.statuswidget.logstatus.textCursor().insertHtml('Selected Coverage for alignment building based on the reads best matching the expected amplicon length: '+str(self.selectlens)+'<br>')                 
                    self.statuswidget.logstatus.textCursor().insertHtml('Maximum deviation of read length (after primer removal) from barcode length: '+str(self.postdemlen)+'<br>')
                    
                if self.consensus90stat==True:
                    self.logfile.write('Coverage used for barcode calling in 2nd step, based on subsetting reads by similarity: '+str(self.n90percsubsetnTextBox.text())+'\n')
                    self.statuswidget.logstatus.textCursor().insertHtml('<br><b>Settings for Phase 2b: Barcode calling with “Consensus by Similarity”</b><br>')
                    self.statuswidget.logstatus.textCursor().insertHtml('Selected coverage used for alignment building based on the reads most similar to preliminary barcode: '+str(self.n90percsubsetnTextBox.text())+'<br>')
                self.logfile.write('Genetic code: '+str(self.gencode)+'\n')
                self.statuswidget.logstatus.textCursor().insertHtml('<br><b>Other general settings</b><br><br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Genetic code: '+str(self.gencode)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Consensus settings: Fixed threshold:'+str(self.consensesfreqfixed)+' Range to examine: '+str(self.consensesfreqrange)+' at step size of '+str(self.consensesfreqstep)+'<br><br>')
                self.statuswidget.logstatus.textCursor().insertHtml('<b>Analysis phases to use in this run</b><br><br>')                
                self.statuswidget.logstatus.textCursor().insertHtml('Build consensus based on reads subset by length: '+str(self.consensuslenstat)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Build consensus based on reads subset by similarity: '+str(self.consensus90stat)+'<br>')

            if self.flag==1:
                self.inputmode=1
                self.logfile.write("Input FASTQ: "+self.infastq+'\nInput Demultiplexing file: '+self.demfile+'\n\nPARAMETERS:\n')
                self.logfile.write('Minimum Length: '+str(self.minlen)+'\n')
                self.logfile.write('Expected barcode length: '+str(self.explen)+'\n')
                self.logfile.write('Window to select reads of during demultiplexing: '+str(self.demlen)+'\n')
                self.logfile.write('Build consensus based on reads subset by length: '+str(self.consensuslenstat)+'\n')
                self.logfile.write('Build consensus based on reads subset by similarity: '+str(self.consensus90stat)+'\n')
                self.logfile.write('Fix barcodes based on MSA: '+str(self.fixstat)+'\n')
                self.statuswidget.logstatus.textCursor().insertHtml("Input FASTQ: "+self.infastq+'<br>Input Demultiplexing file: '+self.demfile+'<br><b><p style="color:blue;">Analysis setup</p><br> Settings for Phase 1 Demultiplexing:</b><br><br>')
                
                self.statuswidget.logstatus.textCursor().insertHtml('Minimum Length: '+str(self.minlen)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Length of barcode: '+str(self.explen)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Window to define product length (length of barcode +/-): '+str(self.demlen)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Window for primer and tag search: '+str(self.primersearchlen)+'<br>')
    
                if self.consensuslenstat==True:
                    self.logfile.write('Window to select reads of long length after demultiplexing: '+str(self.postdemlen)+'\n')
                    self.logfile.write('Coverage used for barcode calling based on subsetting reads by lengths: '+str(self.selectlens)+'\n')
                    self.statuswidget.logstatus.textCursor().insertHtml('<br><b>Settings for Phase 2a: Barcode calling with “Consensus by length”</b><br><br>') 
                    self.statuswidget.logstatus.textCursor().insertHtml('Selected Coverage for alignment building based on the reads best matching the expected amplicon length: '+str(self.selectlens)+'<br>')                 
                    self.statuswidget.logstatus.textCursor().insertHtml('Maximum deviation of read length (after primer removal) from barcode length: '+str(self.postdemlen)+'<br>')
                    
                if self.consensus90stat==True:
                    self.logfile.write('Coverage used for barcode calling in 2nd step, based on subsetting reads by similarity: '+str(self.n90percsubsetnTextBox.text())+'\n')
                    self.statuswidget.logstatus.textCursor().insertHtml('<br><b>Settings for Phase 2b: Barcode calling with “Consensus by Similarity”</b><br>')
                    self.statuswidget.logstatus.textCursor().insertHtml('Selected coverage used for alignment building based on the reads most similar to preliminary barcode: '+str(self.n90percsubsetnTextBox.text())+'<br>')
                self.logfile.write('Genetic code: '+str(self.gencode)+'\n')
                self.statuswidget.logstatus.textCursor().insertHtml('<br><b>Other general settings</b><br><br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Genetic code: '+str(self.gencode)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Consensus settings: Fixed threshold:'+str(self.consensesfreqfixed)+' Range to examine: '+str(self.consensesfreqrange)+' at step size of '+str(self.consensesfreqstep)+'<br><br>')
                self.statuswidget.logstatus.textCursor().insertHtml('<b>Analysis phases to use in this run</b><br><br>')                
                self.statuswidget.logstatus.textCursor().insertHtml('Build consensus based on reads subset by length: '+str(self.consensuslenstat)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Build consensus based on reads subset by similarity: '+str(self.consensus90stat)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Fix barcodes based on MSA: '+str(self.fixstat)+'<br>')                 
            if self.flag==2:
                self.inputmode=2
                self.logfile.write("Input folder of demultiplexed reads: "+self.indir+'\n')
                self.statuswidget.logstatus.textCursor().insertHtml("Input folder of demultiplexed reads: "+self.indir+'<br>')
                self.logfile.write('Length of barcode: '+str(self.explen)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Length of barcode: '+str(self.explen)+'<br><br>')
                self.statuswidget.logstatus.textCursor().insertHtml('<p style="color:blue;"><b>Analysis setup</p><br><b>Settings for Phase 2a: Barcode calling with “Consensus by length”</b><br><br>')
                
                if self.consensuslenstat==True:
                    self.logfile.write('Window to select reads of long length after demultiplexing: '+str(self.postdemlen)+'\n')
                    self.logfile.write('Coverage used for barcode calling based on subsetting reads by lengths: '+str(self.selectlens)+'\n')
            #       self.statuswidget.logstatus.textCursor().insertHtml('<br><b>Settings for Phase 2a: Barcode calling with “Consensus by length”</b><br><br>') 
                    self.statuswidget.logstatus.textCursor().insertHtml('Selected Coverage for alignment building based on the reads best matching the expected amplicon length: '+str(self.selectlens)+'<br>')                 
                    self.statuswidget.logstatus.textCursor().insertHtml('Maximum deviation of read length (after primer removal) from barcode length: '+str(self.postdemlen)+'<br>')
                if self.consensus90stat==True:
                    self.logfile.write('Coverage used for barcode calling in 2nd step, based on subsetting reads by similarity: '+str(self.n90percsubsetnTextBox.text())+'\n')
                    self.statuswidget.logstatus.textCursor().insertHtml('<br><b>Settings for Phase 2b: Barcode calling with “Consensus by Similarity”</b><br><br>')
                    self.statuswidget.logstatus.textCursor().insertHtml('Selected coverage used for alignment building based on the reads most similar to preliminary barcode: '+str(self.n90percsubsetnTextBox.text())+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('<br><b>Other general settings</b><br><br>')
                self.logfile.write('Genetic code: '+str(self.gencode)+'\n')
                self.statuswidget.logstatus.textCursor().insertHtml('Genetic code: '+str(self.gencode)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Consensus settings: Fixed threshold: '+str(self.consensesfreqfixed)+' Range to examine: '+str(self.consensesfreqrange)+' at step size of '+str(self.consensesfreqstep)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('<br><b>Analysis phases to use in this run</b><br><br>')                            
                self.statuswidget.logstatus.textCursor().insertHtml('Build consensus based on reads subset by length: '+str(self.consensuslenstat)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Build consensus based on reads subset by similarity: '+str(self.consensus90stat)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Fix barcodes based on MSA: '+str(self.fixstat)+'<br>')
                
            if self.flag==3:
                self.inputmode=3
                self.logfile.write("Input file to correct: "+self.forfixfname+'\n')
                self.statuswidget.logstatus.textCursor().insertHtml("Input file to correct: "+self.forfixfname+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Genetic code: '+str(self.gencode)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('<br><b>Analysis phases to use in this run</b><br><br>')                            
                self.statuswidget.logstatus.textCursor().insertHtml('Build consensus based on reads subset by length: '+str(self.consensuslenstat)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Build consensus based on reads subset by similarity: '+str(self.consensus90stat)+'<br>')
                self.statuswidget.logstatus.textCursor().insertHtml('Fix barcodes based on MSA: '+str(self.fixstat)+'<br>')
                
      #      print self.flag, "flag"
            if self.consensuslenstat==False:
                if self.consensus90stat==False:
                    if self.flag==1:
                        self.setdemultiplex()
                                               
                    if self.flag==2:
                        self.showwarningdialog("Since you want to only fix the barcodes, please provide a file containing the barcodes to fix")
                        self.forfixfname=QtWidgets.QFileDialog.getOpenFileName(self, "Select your consensus fasta")
                        self.forfixfname=str(self.forfixfname[0])
                        if self.forfixfname=="":
                            self.runstat=False
                        else:

                            if self.fixstat==True:
                                self.setforupdateconsensus()    
                            else:
                                self.showwarningdialog("You have not selected any action to perform. Returning to settings menu")
                                     
                             
                    if self.flag==3:
                        self.setforupdateconsensus()
                    if self.flag==4:
                        self.parsexls()
                else:
                    self.showwarningdialog("Since you want to improve existing consensus, please provide a file containing the barcodes")
                    self.forconsensus90fname=QtWidgets.QFileDialog.getOpenFileName(self, "Select your consensus fasta")
                #   print self.forconsensus90fname
                    if str(self.forconsensus90fname[0])=="":
                        self.runstat=False
                    else:
                        self.setforupdateconsensus()    
            else:
                if self.flag==1:
                    self.setdemultiplex()
                if self.flag==2:
                    self.setforconsensusonly()
                if self.flag==4:
                    self.parsexls()
        else:
            self.getparams()
    def parsexls(self):
        self.statuswidget.logstatus.textCursor().insertHtml("loading demultiplexing files. ")
        self.myparsexls=parsexlsclass(self.indemfiles,self.excelpath,self.outpathdemfiles,self.primerFseq,self.primerRseq)
        self.myparsexls.notifyFinished.connect(self.startmerge)
        self.myparsexls.start()
    def startmerge(self):
        def mergefiles(inlist, outfilename):
            o = open(outfilename, 'w')
            for each in inlist:
                for line in fileinput.input([each]):
                    o.write(line)
                fileinput.close()
            o.close()
        
        mergefiles([os.path.join(self.outpathdemfiles,x+'.csv') for x in self.indemfiles],os.path.join(self.outpathdemfiles,"mergeddemfile"))
        self.statuswidget.logstatus.textCursor().insertHtml("loading demultiplexing files. "+str(self.myparsexls.n)+" excel files and " +str(self.myparsexls.rcount) + " specimens.")
        self.demfile=os.path.join(self.outpathdemfiles,"mergeddemfile")
        self.setdemultiplex()

    def setforupdateconsensus(self):
    #   print "skipping consensus"
        if self.flag==3:
            self.updateconsensus()
        else:
            self.mycounterdemreads=calculatecoverage(self.indir)
            #if self.consensuslenstat==False:
            #   if self.consensus90stat==True:
                                                                                    
            self.mycounterdemreads.taskFinished.connect(self.updateconsensus)
            self.startcalcdemreads()
                 
    def terminate(self):
    #   print "changetabtab"

        sys.exit()
        
    def setforconsensusonly(self):
        self.sampleids={}
        dirlist=os.listdir(self.indir)
        for fname in dirlist:
            self.sampleids[fname.split("_all.fa")[0]]=""
        self.statuswidget.logstatus.textCursor().insertHtml("You have "+str(len(dirlist))+' input files.')
                                                                        
        self.minlen=int(self.MinLengthTextBox.text())
        self.explen=int(self.ExpLengthTextBox.text())
        self.inlistforconsensus=fnmatch.filter(os.listdir(self.indir),"*")
                                                            
        self.mycounterdemreads=calculatecoverage(self.indir)
                                                                       
        self.mycounterdemreads.taskFinished.connect(self.setprintreadfasta)
        self.startcalcdemreads()
    def startcalcdemreads(self):
    #   print "startcalc"
        self.mycounterdemreads.start()
 
    def setprintreadfasta(self):
                       
        self.statuswidget.logstatus.textCursor().insertHtml("<br>These contain "+str(sum(list(self.mycounterdemreads.counter.values())))+' reads. <br><b><p style="color:blue;">Phase 2a: Consensus by Length</b></p>Reads are sorted by deviation from the length of the amplicon. The reads with the best length match are selected for alignment whereby the maximum number of reads to consider was specified by the user during setup. The reads are aligned and a consensus barcode is called. It is only accepted as a candidate barcode if it has the correct length, is translatable, and free of ambiguous bases. If the iterative mode was chosen during setup, this process is repeated at increasingly higher coverage for those bins that fail to yield a barcode at lower coverages.')
        self.dirdict=self.mycounterdemreads.dirdict
    #   print self.dirdict
        self.printreadfasta(1)
    def getparams2(self):
        self.livestack.setCurrentIndex(1)
        self.flag=False
        try:
            self.livedir
            self.livedemfile                       
            self.flag='live'
        except AttributeError:
            self.showerrordialog("You forgot at least one input")
        if self.flag!=False:
            self.d=QtWidgets.QDialog()
            layout=QtWidgets.QGridLayout()
            layout.setSpacing(20)
            demlabel=QtWidgets.QLabel('<b>Demultiplexing Settings</b>',self.d)
            ConsensusLabel=QtWidgets.QLabel('<b>Consensus settings</b>',self.d)       
            MinimumCoverageLabel=QtWidgets.QLabel('Minimum Coverage',self.d)
            LargeCoverageLabel=QtWidgets.QLabel('Large Coverage, or coverage beyond which you dont examine every increment',self.d)
            LargeCoverageStepLabel=QtWidgets.QLabel('Step size for larger coverages',self.d)
            MaximumCoverageLabel=QtWidgets.QLabel('Maximum Coverage, set 0 for no limit',self.d)
            Geneticcodelabel=QtWidgets.QLabel('Genetic Code: ',self.d)
            self.gencodebox=QtWidgets.QComboBox()
            self.gencodebox.addItem("1. The Standard Code")
            self.gencodebox.addItem("2. The Vertebrate Mitochondrial Code")
            self.gencodebox.addItem("3. The Yeast Mitochondrial Code")
            self.gencodebox.addItem("4. The Mold, Protozoan, and Coelenterate Mitochondrial Code")
            self.gencodebox.addItem("5. The Invertebrate Mitochondrial Code")
            self.gencodebox.addItem("6. The Ciliate, Dasycladacean and Hexamita Nuclear Code")
            self.gencodebox.addItem("9. The Echinoderm and Flatworm Mitochondrial Code")
            self.gencodebox.addItem("10. The Euplotid Nuclear Code")
            self.gencodebox.addItem("11. The Bacterial, Archaeal and Plant Plastid Code")
            self.gencodebox.addItem("12. The Alternative Yeast Nuclear Code")
            self.gencodebox.addItem("13. The Ascidian Mitochondrial Code")
            self.gencodebox.addItem("14. The Alternative Flatworm Mitochondrial Code")
            self.gencodebox.addItem("16. Chlorophycean Mitochondrial Code")
            self.gencodebox.addItem("21. Trematode Mitochondrial Code")
            self.gencodebox.addItem("22. Scenedesmus obliquus Mitochondrial Code")
            self.gencodebox.addItem("23. Thraustochytrium Mitochondrial Code")
            self.gencodebox.addItem("24. Rhabdopleuridae Mitochondrial Code")
            self.gencodebox.addItem("25. Candidate Division SR1 and Gracilibacteria Code")
            self.gencodebox.addItem("26. Pachysolen tannophilus Nuclear Code")
            self.gencodebox.addItem("27. Karyorelict Nuclear Code")
            self.gencodebox.addItem("28. Condylostoma Nuclear Code")
            self.gencodebox.addItem("29. Mesodinium Nuclear Code")
            self.gencodebox.addItem("30. Peritrich Nuclear Code")
            self.gencodebox.addItem("31. Blastocrithidia Nuclear Code")
            self.gencodebox.addItem("33. Cephalodiscidae Mitochondrial UAA-Tyr Code")
            self.gencodebox.activated[str].connect(self.gencodechoice)
            self.gencode=5
            self.gencodebox.setCurrentIndex(4)
            MinLengthLabel = QtWidgets.QLabel('Minimum Length: ',self.d)
            ExpLengthLabel = QtWidgets.QLabel('Length of barcode: ',self.d)
            demlenlab=QtWidgets.QLabel('Window to define product length (length of barcode +/-): ',self.d)
            primermismatchlab=QtWidgets.QLabel('Number of mismatches for primer search: ',self.d)
            primersearchlab=QtWidgets.QLabel('Window for primer and tag search: ',self.d)
      #      postdemlenlab=QtWidgets.QLabel('Maximum deviation of read length from barcode length: ',self.d2)
            consensesfreqfixedlabel=QtWidgets.QLabel('Main consensus calling frequency: ',self.d)
            consensesfreqrangelabel=QtWidgets.QLabel('Range of frequencies to assess: ',self.d)
            consensesfreqsteplabel=QtWidgets.QLabel('to be examined at step size of ',self.d)
            self.MinLengthTextBox = QtWidgets.QLineEdit()
            self.MinLengthTextBox.setText("658")
            self.ExpLengthTextBox = QtWidgets.QLineEdit()
            self.ExpLengthTextBox.setText("658")
            self.PrimerSearchTextBox = QtWidgets.QLineEdit()
            self.PrimerSearchTextBox.setText("100")
            self.PostdemLengthTextBox = QtWidgets.QLineEdit()
            self.PostdemLengthTextBox.setText("50")
            self.DemLengthTextBox = QtWidgets.QLineEdit()
            self.DemLengthTextBox.setText("100")
            tagmismatchlabel = QtWidgets.QLabel('Tag Mismatch: ',self.d)
            self.combobox1 =  QtWidgets.QComboBox()
            self.combobox1.addItem('0')
            self.combobox1.addItem('1')
            self.combobox1.addItem('2')
            self.combobox1.setCurrentText('2')
            self.combobox1.currentTextChanged.connect(self.tagmismatchchanged)
            self.PrimersearchMismatchTextBox = QtWidgets.QLineEdit()
            self.PrimersearchMismatchTextBox.setText("10")
            self.MinimumCoverageTextBox = QtWidgets.QLineEdit()
            self.MinimumCoverageTextBox.setText("20")
            self.LargeCoverageTextBox = QtWidgets.QLineEdit()
            self.LargeCoverageTextBox.setText("100")
            self.LargeCoverageStepTextBox = QtWidgets.QLineEdit()
            self.LargeCoverageStepTextBox.setText("10")
            self.MaximumCoverageTextBox = QtWidgets.QLineEdit()
            self.MaximumCoverageTextBox.setText("0")
            self.subsetmaxcheckbox = QtWidgets.QCheckBox("Subset higher coverage to Large coverage based on read lengths?")
            self.subsetmaxcheckbox.setChecked(True)
            self.subsetmaxcheckbox.stateChanged.connect(self.subsetornot)
            self.consensesfreqfixedTextbox= QtWidgets.QLineEdit()
            self.consensesfreqrangeTextbox= QtWidgets.QLineEdit()
            self.consensesfreqstepTextbox= QtWidgets.QLineEdit()
            self.consensesfreqfixedTextbox.setText("0.3")
            self.consensesfreqrangeTextbox.setText("0.2,0.5")
            self.consensesfreqstepTextbox.setText("0.05")
            self.basecallingcheckbox=QtWidgets.QCheckBox("Do you want to basecall as well as do barcode calling?")
            self.basecallingcheckbox.stateChanged.connect(lambda:self.stepcheck(self.basecallingcheckbox))
            self.basecallingcheckbox.setChecked(False)
            self.basecallingstat=False
            self.remotetransfercheckbox=QtWidgets.QCheckBox("Do you want to transfer files from Mk1C live?")
            self.remotetransfercheckbox.stateChanged.connect(lambda:self.stepcheck(self.remotetransfercheckbox))
            self.remotetransfercheckbox.setChecked(False)
            self.remotetransferstat=False
            hostlabel=QtWidgets.QLabel('Hostname',self.d)
            usernamelabel=QtWidgets.QLabel('Username',self.d)       
            passwordlabel=QtWidgets.QLabel('Password',self.d)
            remotepathlabel=QtWidgets.QLabel('Path to remote directory',self.d)            
            
            self.HostTextBox = QtWidgets.QLineEdit()
            self.HostTextBox.setText("mc-XXXXXX.local")
            self.HostTextBox.setDisabled(True)
            self.UsernameTextBox = QtWidgets.QLineEdit()
            self.UsernameTextBox.setText("")
            self.UsernameTextBox.setDisabled(True)
            self.PasswordTextBox = QtWidgets.QLineEdit()
            self.PasswordTextBox.setText("")
            self.PasswordTextBox.setDisabled(True)
            self.RemotepathTextbox = QtWidgets.QLineEdit()
            self.RemotepathTextbox.setText("")
            self.RemotepathTextbox.setDisabled(True)
           
            self.tagmm=2
            self.subsetmax=True
            self.minlen=658
            self.explen=658
            self.postdemlen=50
            self.mincoverage=20
            self.maxcoverage=0
            self.largecoverage=100
            self.largecoveragestep=10
            self.demlen=100
            self.primermismatch=10
            self.primersearchlen=100
            self.consensesfreqfixed=0.3
            self.consensesfreqrange=[0.2,0.5]
            self.consensesfreqstep=0.05
            regex = QtCore.QRegExp("^[0-9]*$")
            validator = QtGui.QRegExpValidator(regex,self.MinLengthTextBox)
            self.procbutton=QtWidgets.QPushButton("Proceed",self.d)
            self.procbutton.clicked.connect(self.setdemoutpath)

            layout.addWidget(demlabel,0,0,1,1)
            layout.addWidget(MinLengthLabel,1,0,1,2)
            layout.addWidget(self.MinLengthTextBox,1,2,1,1)
            layout.addWidget(ExpLengthLabel,2,0,1,2)
            layout.addWidget(self.ExpLengthTextBox,2,2,1,1)
            layout.addWidget(demlenlab,3,0,1,2)
            layout.addWidget(self.DemLengthTextBox,3,2,1,1)

            layout.addWidget(primersearchlab,4,0,1,1)
            layout.addWidget(self.PrimerSearchTextBox,4,2,1,1)
            layout.addWidget(primermismatchlab,4,3,1,1)
            layout.addWidget(self.PrimersearchMismatchTextBox,4,4,1,1)        
            layout.addWidget(tagmismatchlabel,5,0,1,2) 
            layout.addWidget(self.combobox1,5,2,1,1)  
            layout.addWidget(ConsensusLabel,6,0,1,1)
            layout.addWidget(MinimumCoverageLabel,7,0,1,2)
            layout.addWidget(self.MinimumCoverageTextBox,7,2,1,1)
            layout.addWidget(LargeCoverageLabel,8,0,1,2)
            layout.addWidget(self.LargeCoverageTextBox,8,2,1,1)     
            layout.addWidget(LargeCoverageStepLabel,9,0,1,2)
            layout.addWidget(self.LargeCoverageStepTextBox,9,2,1,1)
            layout.addWidget(MaximumCoverageLabel,10,0,1,2)
            layout.addWidget(self.MaximumCoverageTextBox,10,2,1,1)
            layout.addWidget(self.subsetmaxcheckbox,11,0,1,1)          
            layout.addWidget(consensesfreqfixedlabel,12,0,1,2)
            layout.addWidget(self.consensesfreqfixedTextbox,12,2)
            layout.addWidget(consensesfreqrangelabel,13,0,1,2)
            layout.addWidget(self.consensesfreqrangeTextbox,13,2)
            layout.addWidget(consensesfreqsteplabel,14,0,1,2)
            layout.addWidget(self.consensesfreqstepTextbox,14,2)
            layout.addWidget(Geneticcodelabel,15,0,1,2)
            layout.addWidget(self.gencodebox,15,2,1,2)
            layout.addWidget(self.basecallingcheckbox,16,0,1,4)

            layout.addWidget(self.remotetransfercheckbox,17,0,1,4)
            layout.addWidget(hostlabel,18,0,1,2)
            layout.addWidget(self.HostTextBox,18,2,1,1)     
            layout.addWidget(usernamelabel,19,0,1,2)
            layout.addWidget(self.UsernameTextBox,19,2,1,1)
            layout.addWidget(passwordlabel,20,0,1,2)
            layout.addWidget(self.PasswordTextBox,20,2,1,1)
            layout.addWidget(remotepathlabel,21,0,1,1)          
            layout.addWidget(self.RemotepathTextbox,21,2,1,1)
            layout.addWidget(self.procbutton,22,4)
            self.d.setLayout(layout)
            self.d.exec_()
    def checkbasecallstat(self):
    #    print self.basecallingcheckbox.isChecked()
        if self.basecallingcheckbox.isChecked()==False:
            self.setdemoutpath()
        else:
            self.liveornot='basecall'
            self.livebasecallingparams()
    def stoplivebarcodingwarning(self):
        qm = QtWidgets.QMessageBox()
        qm.question(self,'', "Are you sure you want to stop live barcoding?", qm.Yes | qm.No)

        if qm.Yes:
            self.stoplivebarcoding()
        else:
          qm.information(self,'',"Nothing Changed")
    def stoplivebarcoding(self):
        try:
            if self.mylivedetector.isRunning()==True:
                self.mylivedetector.quit()
        except AttributeError:
            print("no live detector")
        try:
            if self.mylivesamplecount.isRunning()==True:
                self.mylivesamplecount.quit()
        except AttributeError:
            print("no live detector")
        try:
            if self.myrundemultiplexlive.isRunning()==True:
                self.myrundemultiplexlive.quit()
        except AttributeError:
            print("no live detector")
        try:
            if self.mylivefast5detector.isRunning()==True:
                self.mylivefast5detector.quit()
        except AttributeError:
            print("no live detector")
        try:
            if self.myliveguppyrun.isRunning()==True:
                self.myliveguppyrun.quit()
        except AttributeError:
            print("no live detector")
        try:
            if self.myrunconsensuslive.isRunning()==True:
                self.myrunconsensuslive.quit()
        except AttributeError:
            print("no live detector")
        try:
            if self.myliveremotedetector.isRunning()==True:
                self.myliveremotedetector.quit()
        except AttributeError:
            print("no live detector")
        try:
            if self.myliveremotetransfer.isRunning()==True:
                self.myliveremotetransfer.quit()            
        except AttributeError:
            print("no live detector")
    def livebasecallingsoftwareparams(self):
  #      self.stoplivebarcodingbtn=QtWidgets.QPushButton("Stop real time barcoding")
     #   self.grid_right.addWidget(self.stoplivebarcodingbtn,1,0,alignment=QtCore.Qt.AlignRight)
  #      self.stoplivebarcodingbtn.adjustSize()
  #      self.stoplivebarcodingbtn.clicked.connect(self.stoplivebarcodingwarning)


        self.d.close()
        self.livebasecallsoftwaredialog=QtWidgets.QDialog()
        layout=QtWidgets.QGridLayout()
        layout.setSpacing(20)
        basecallerlabel=QtWidgets.QLabel('<b>Basecalling software</b>', self.livebasecallsoftwaredialog)
        self.basecallertypebox=QtWidgets.QComboBox()
        self.basecallertypebox.addItem("Guppy")
        self.basecallertypebox.addItem("Dorado")
        self.basecallertypebox.setCurrentIndex(0)
        self.basecallertype="Guppy"   
        self.basecallertypebox.currentTextChanged.connect(self.basecallertypechoice)
        self.basecallertypebutton=QtWidgets.QPushButton("Proceed", self.livebasecallsoftwaredialog)
        self.basecallertypebutton.clicked.connect(self.setbasecallertype)
        layout.addWidget(basecallerlabel,0,0,1,1)
        layout.addWidget(self.basecallertypebox,2,0,1,1)

    #    layout.addWidget(self.CmdLineTextBox2,8,0,2,4)
        layout.addWidget(self.basecallertypebutton,9,4)

        self.livebasecallsoftwaredialog.setLayout(layout)
        self.livebasecallsoftwaredialog.exec_()
    def livebasecallingparams(self):
        if "guppyconfig.txt" in os.listdir("config"):
            with open(os.path.join("config","guppyconfig.txt")) as infile:
                l=infile.readlines()
                if len(l)==2:
                    self.guppypath=l[0].strip().split("=")[1]
                    self.configpath=l[1].strip().split("=")[1]
                else:
                    self.guppypath=''
                    self.configpath=''
        else:
            self.guppypath=''
            self.configpath=''
        self.livebasecalldialog=QtWidgets.QDialog()
        layout=QtWidgets.QGridLayout()
        layout.setSpacing(20)
        basecalllabel=QtWidgets.QLabel('<b>Basecalling</b>', self.livebasecalldialog)
        TypeLabel=QtWidgets.QLabel('Flow cell type', self.livebasecalldialog)
        GuppyPathLabel=QtWidgets.QLabel('Path to Guppy', self.livebasecalldialog)       
        NumCallersLabel=QtWidgets.QLabel('Number of callers', self.livebasecalldialog)
        GPURunnersPerLabel=QtWidgets.QLabel('GPU Runners per device', self.livebasecalldialog)
        ChunksPerRunnerLabel=QtWidgets.QLabel('Chunks per runner', self.livebasecalldialog)
        ConfigFileLabel=QtWidgets.QLabel('Path to guppy configfile', self.livebasecalldialog)
        DeviceLabel=QtWidgets.QLabel('GPU device', self.livebasecalldialog)

        self.guppypathbutton=QtWidgets.QPushButton("...", self.livebasecalldialog)
        self.guppypathbutton.clicked.connect(self.setguppypath)
        self.configpathbutton=QtWidgets.QPushButton("...", self.livebasecalldialog)
        self.configpathbutton.clicked.connect(self.setconfigpath)

        self.flowcelltypebox=QtWidgets.QComboBox()
        self.flowcelltypebox.addItem("MinION")
        self.flowcelltypebox.addItem("Flongle")
        self.flowcelltype="MinION"
        self.flowcelltypebox.setCurrentIndex(0)       
        self.flowcelltypebox.currentTextChanged.connect(self.flowcelltypechoice)
        self.guppypathtext=QtWidgets.QLabel(self.guppypath, self.livebasecalldialog)
        self.configpathtext=QtWidgets.QLabel(self.configpath, self.livebasecalldialog)

        self.NumCallersTextBox = QtWidgets.QLineEdit()
        self.NumCallersTextBox.setText("1")
        self.GPURunnerTextBox = QtWidgets.QLineEdit()
        self.GPURunnerTextBox.setText("2")
        self.ChunksPerRunnerTextBox = QtWidgets.QLineEdit()
        self.ChunksPerRunnerTextBox.setText("48")
        self.DeviceTextBox = QtWidgets.QLineEdit()
        self.DeviceTextBox.setText("auto")
        self.NumCallersTextBox.textChanged.connect(self.updatenumcallers)
        self.GPURunnerTextBox.textChanged.connect(self.updategpucallers)
        self.ChunksPerRunnerTextBox.textChanged.connect(self.updatechunksperrunner)
        self.DeviceTextBox.textChanged.connect(self.updatedevicename)
        self.commandlinecheckbox=QtWidgets.QCheckBox("Do you instead want to specify your own command? <br> Note that the fastq files must be saved in the output folder directly and not under subfolders")
        self.commandlinecheckbox.stateChanged.connect(lambda:self.stepcheck(self.commandlinecheckbox))
        self.commandlinecheckbox.setChecked(False)
        self.commandlinestat=False        
        self.CmdLineTextBox = QtWidgets.QPlainTextEdit()
        self.gpucallers="2"
        self.numcallers="1"
        self.chunksperrunner="48"
        self.devicename="auto"
        if self.flowcelltype=="MinION":
            self.guppycommand='\"'+self.guppypath+'\" --gpu_runners_per_device '+self.gpucallers+' --chunks_per_runner '+self.chunksperrunner+' --num_callers '+self.numcallers+' -i \"ignoreinputfolder\" -s \"ignorethis\" --config \"'+self.configpath + '\" --input_file_list inlistfast5.txt --device \"'+self.devicename+'\" --disable_qscore_filtering --read_batch_size 4000'
        else:
            self.guppycommand='\"'+self.guppypath+'\" --gpu_runners_per_device '+self.gpucallers+' --chunks_per_runner '+self.chunksperrunner+' --num_callers '+self.numcallers+' -i \"ignoreinputfolder\" -s \"ignorethis\" --config \"'+self.configpath + '\" --input_file_list inlistfast5.txt --device \"'+self.devicename+'\" --disable_qscore_filtering --read_batch_size 1000'            
        self.CmdLineTextBox.setPlainText(self.guppycommand)
        self.CmdLineTextBox.setDisabled(True)
     #   self.CmdLineTextBox.textChanged.connect(self.updateguppycommand)
        self.demrunbutton=QtWidgets.QPushButton("Run",self.d)
        self.demrunbutton.clicked.connect(self.setdemoutpath)

        layout.addWidget(basecalllabel,0,0,1,1)
        layout.addWidget(TypeLabel,1,0,1,1)
        layout.addWidget(self.flowcelltypebox,1,2,1,1)
        layout.addWidget(GuppyPathLabel,2,0,1,1)
        layout.addWidget(self.guppypathbutton,2,2,1,1)
        layout.addWidget(self.guppypathtext,2,3,1,2)
        layout.addWidget(ConfigFileLabel,3,0,1,1)
        layout.addWidget(self.configpathbutton,3,2,1,1)
        layout.addWidget(self.configpathtext,3,3,1,2)
        layout.addWidget(DeviceLabel,4,0,1,1)
        layout.addWidget(self.DeviceTextBox,4,2,1,1)
        layout.addWidget(NumCallersLabel,5,0,1,1)
        layout.addWidget(self.NumCallersTextBox,5,2,1,1)
        layout.addWidget(GPURunnersPerLabel,6,0,1,1)
        layout.addWidget(self.GPURunnerTextBox,6,2,1,1)
        layout.addWidget(ChunksPerRunnerLabel,7,0,1,1)
        layout.addWidget(self.ChunksPerRunnerTextBox,7,2,1,1)
        layout.addWidget(ChunksPerRunnerLabel,7,0,1,1)
        layout.addWidget(self.commandlinecheckbox,8,0,1,4)
        layout.addWidget(self.CmdLineTextBox,9,0,2,4)
        layout.addWidget(self.demrunbutton,10,4)
        self.livebasecalldialog.setLayout(layout)
        self.livebasecalldialog.exec_()
    def livedoradobasecallingparams(self):
     #   print ("Dorado")
  #      self.stoplivebarcodingbtn=QtWidgets.QPushButton("Stop real time barcoding")
     #   self.grid_right.addWidget(self.stoplivebarcodingbtn,1,0,alignment=QtCore.Qt.AlignRight)
  #      self.stoplivebarcodingbtn.adjustSize()
  #      self.stoplivebarcodingbtn.clicked.connect(self.stoplivebarcodingwarning)

        self.livebasecallsoftwaredialog.close()
        if "doradoconfig.txt" in os.listdir("config"):
            with open(os.path.join("config","doradoconfig.txt")) as infile:
                l=infile.readlines()
                if len(l)==2:
                    self.doradopath=l[0].strip().split("=")[1]
                    self.doradoconfigpath=l[1].strip().split("=")[1]
                else:
                    self.doradopath=''
                    self.doradoconfigpath=''
        else:
            self.doradopath=''
            self.doradoconfigpath=''
        self.livedoradobasecalldialog=QtWidgets.QDialog()
        layout=QtWidgets.QGridLayout()
        layout.setSpacing(20)
        basecalllabel=QtWidgets.QLabel('<b>Basecalling</b>', self.livedoradobasecalldialog)
        DoradoPathLabel=QtWidgets.QLabel('Path to Dorado', self.livedoradobasecalldialog)       
        Chunklabel=QtWidgets.QLabel('Chunk size', self.livedoradobasecalldialog)
        OverlapLabel=QtWidgets.QLabel('Overlap', self.livedoradobasecalldialog)
        DoradoConfigFileLabel=QtWidgets.QLabel('Path to Dorado configfile', self.livedoradobasecalldialog)
        DeviceLabel=QtWidgets.QLabel('GPU device', self.livedoradobasecalldialog)

        self.doradopathbutton=QtWidgets.QPushButton("...", self.livedoradobasecalldialog)
        self.doradopathbutton.clicked.connect(self.setdoradopath)
        self.doradoconfigpathbutton=QtWidgets.QPushButton("...", self.livedoradobasecalldialog)
        self.doradoconfigpathbutton.clicked.connect(self.setdoradoconfigpath)

        self.doradopathtext=QtWidgets.QLabel(self.doradopath, self.livedoradobasecalldialog)
        self.doradoconfigpathtext=QtWidgets.QLabel(self.doradoconfigpath, self.livedoradobasecalldialog)



        self.ChunkTextBox = QtWidgets.QLineEdit()
        self.ChunkTextBox.setText("10000")
        self.OverlapTextBox = QtWidgets.QLineEdit()
        self.OverlapTextBox.setText("500")
        self.DeviceTextBox2 = QtWidgets.QLineEdit()
        self.DeviceTextBox2.setText("all")

        self.ChunkTextBox.textChanged.connect(self.updatechunks)
        self.ChunkTextBox.textChanged.connect(self.updateoverlap)
        self.DeviceTextBox2.textChanged.connect(self.updatedevicename2)
        self.commandlinecheckbox=QtWidgets.QCheckBox("Do you instead want to specify your own command? <br> Note that the fastq files must be saved in the output folder directly and not under subfolders")
        self.commandlinecheckbox.stateChanged.connect(lambda:self.stepcheck(self.commandlinecheckbox))
        self.commandlinecheckbox.setChecked(False)
        self.commandlinestat=False        
        self.CmdLineTextBox2 = QtWidgets.QPlainTextEdit()

        self.chunks="10000"
        self.overlap="500"
        self.devicename="cuda:all"
        self.doradocommand='\"'+self.doradopath+'\" basecaller --emit-fastq --device \"'+self.devicename+'\" --chunksize '+self.chunks+' --overlap '+self.overlap+' \"'+self.doradoconfigpath + '\"  \"ignoreinputfolder\"'
        self.CmdLineTextBox2.setPlainText(self.doradocommand)
        self.CmdLineTextBox2.setDisabled(True)
     #   self.CmdLineTextBox.textChanged.connect(self.updateguppycommand)
        self.demrunbutton=QtWidgets.QPushButton("Run",self.d)
        self.demrunbutton.clicked.connect(self.setdemoutpath)

        layout.addWidget(basecalllabel,0,0,1,1)
        layout.addWidget(DoradoPathLabel,2,0,1,1)
        layout.addWidget(self.doradopathbutton,2,2,1,1)
        layout.addWidget(self.doradopathtext,2,3,1,2)
        layout.addWidget(DoradoConfigFileLabel,3,0,1,1)
        layout.addWidget(self.doradoconfigpathbutton,3,2,1,1)
        layout.addWidget(self.doradoconfigpathtext,3,3,1,2)
        layout.addWidget(DeviceLabel,4,0,1,1)
        layout.addWidget(self.DeviceTextBox2,4,2,1,1)
        layout.addWidget(Chunklabel,5,0,1,1)
        layout.addWidget(self.ChunkTextBox,5,2,1,1)
        layout.addWidget(OverlapLabel,6,0,1,1)
        layout.addWidget(self.OverlapTextBox,6,2,1,1)
        layout.addWidget(self.commandlinecheckbox,7,0,1,4)
        layout.addWidget(self.CmdLineTextBox2,8,0,2,4)
        layout.addWidget(self.demrunbutton,9,4)
        self.livedoradobasecalldialog.setLayout(layout)
        self.livedoradobasecalldialog.exec_()
    def updatechunks(self):
        self.chunks=self.ChunkTextBox.text()
        self.updatedoradocommand()
    def updateoverlap(self):
        self.overlap=self.OverlapTextBox.text()
        self.updatedoradocommand()
    def updatedevicename2(self):
        self.devicename=self.DeviceTextBox2.text()
        self.updatedoradocommand()
    def updatedevicename(self):
        self.devicename=self.DeviceTextBox.text()
        self.updateguppycommand()
    def updatenumcallers(self):
        self.numcallers=self.NumCallersTextBox.text()
        self.updateguppycommand()
    def updategpucallers(self):
        self.gpucallers=self.GPURunnerTextBox.text()
        self.updateguppycommand()
    def updatechunksperrunner(self):
        self.chunksperrunner=self.ChunksPerRunnerTextBox.text()
        self.updateguppycommand()
    def updateguppycommand(self):
     #   print "updatingguppy",self.flowcelltype
        if self.flowcelltype=="MinION":
            self.guppycommand='\"'+self.guppypath+'\" --gpu_runners_per_device '+self.gpucallers+' --chunks_per_runner '+self.chunksperrunner+' --num_callers '+self.numcallers+' -i \"ignoreinputfolder\" -s \"ignorethis\" --config \"'+self.configpath + '\" --input_file_list inlistfast5.txt --device \"'+self.devicename+'\" --disable_qscore_filtering --read_batch_size 4000'
        else:
            self.guppycommand='\"'+self.guppypath+'\" --gpu_runners_per_device '+self.gpucallers+' --chunks_per_runner '+self.chunksperrunner+' --num_callers '+self.numcallers+' -i \"ignoreinputfolder\" -s \"ignorethis\" --config \"'+self.configpath + '\" --input_file_list inlistfast5.txt --device \"'+self.devicename+'\" --disable_qscore_filtering --read_batch_size 1000'
        self.CmdLineTextBox.setPlainText(self.guppycommand)
    def updatedoradocommand(self):
        self.doradocommand='\"'+self.doradopath+'\" basecaller --emit-fastq --device \"'+self.devicename+'\" --chunksize '+self.chunks+' --overlap '+self.overlap+' \"'+self.doradoconfigpath + '\"  \"ignoreinputfolder\"'
        self.CmdLineTextBox2.setPlainText(self.doradocommand)
    def setdoradopath(self):
        self.doradopath = QtWidgets.QFileDialog.getOpenFileName(self, "Select Dorado executable")
        self.doradopath=str(self.doradopath[0])
        self.doradopathtext.setText(self.doradopath)
        self.updatedoradocommand()

    def setdoradoconfigpath(self):
        self.doradoconfigpath = QtWidgets.QFileDialog.getExistingDirectory(self, "Select config file")
        self.doradoconfigpath=str(self.doradoconfigpath)
        self.doradoconfigpathtext.setText(self.doradoconfigpath)
        self.updatedoradocommand()
            
    def setguppypath(self):
        self.guppypath = QtWidgets.QFileDialog.getOpenFileName(self, "Select guppy executable")
        self.guppypath=str(self.guppypath[0])
        self.guppypathtext.setText(self.guppypath)

    def setconfigpath(self):
        self.configpath = QtWidgets.QFileDialog.getOpenFileName(self, "Select config file")
        self.configpath=str(self.configpath[0])
        self.configpathtext.setText(self.configpath)
    def getparams(self):
        self.conventionalstack.setCurrentIndex(1)
        self.flag=False

        try:
            self.infastq
            self.demfile
            self.flag=1
        except AttributeError:
            try:
                self.indir
                                 
                self.flag=2
            except AttributeError:
                try:
                    self.forfixfname
                    self.flag=3
                except AttributeError:
                    self.showerrordialog("You forgot at least one input")

        if self.flag!=False:
            self.d=QtWidgets.QDialog()
            layout=QtWidgets.QGridLayout()
            layout.setSpacing(10)
            demlabel=QtWidgets.QLabel('<b>Demultiplexing Settings</b>',self.d)
            ConsensusByLengthlabel=QtWidgets.QLabel('<b>Consensus by length settings</b>',self.d)
            ConsensusBySimlabel=QtWidgets.QLabel('<b>Consensus by similarity settings</b>',self.d)
            Consensusgeneralsettings=QtWidgets.QLabel('<b>General settings for consensus calling</b>',self.d)
            Geneticcodelabel=QtWidgets.QLabel('Genetic Code: ',self.d)
            self.gencodebox=QtWidgets.QComboBox()
            self.gencodebox.addItem("1. The Standard Code")
            self.gencodebox.addItem("2. The Vertebrate Mitochondrial Code")
            self.gencodebox.addItem("3. The Yeast Mitochondrial Code")
            self.gencodebox.addItem("4. The Mold, Protozoan, and Coelenterate Mitochondrial Code")
            self.gencodebox.addItem("5. The Invertebrate Mitochondrial Code")
            self.gencodebox.addItem("6. The Ciliate, Dasycladacean and Hexamita Nuclear Code")
            self.gencodebox.addItem("9. The Echinoderm and Flatworm Mitochondrial Code")
            self.gencodebox.addItem("10. The Euplotid Nuclear Code")
            self.gencodebox.addItem("11. The Bacterial, Archaeal and Plant Plastid Code")
            self.gencodebox.addItem("12. The Alternative Yeast Nuclear Code")
            self.gencodebox.addItem("13. The Ascidian Mitochondrial Code")
            self.gencodebox.addItem("14. The Alternative Flatworm Mitochondrial Code")
            self.gencodebox.addItem("16. Chlorophycean Mitochondrial Code")
            self.gencodebox.addItem("21. Trematode Mitochondrial Code")
            self.gencodebox.addItem("22. Scenedesmus obliquus Mitochondrial Code")
            self.gencodebox.addItem("23. Thraustochytrium Mitochondrial Code")
            self.gencodebox.addItem("24. Rhabdopleuridae Mitochondrial Code")
            self.gencodebox.addItem("25. Candidate Division SR1 and Gracilibacteria Code")
            self.gencodebox.addItem("26. Pachysolen tannophilus Nuclear Code")
            self.gencodebox.addItem("27. Karyorelict Nuclear Code")
            self.gencodebox.addItem("28. Condylostoma Nuclear Code")
            self.gencodebox.addItem("29. Mesodinium Nuclear Code")
            self.gencodebox.addItem("30. Peritrich Nuclear Code")
            self.gencodebox.addItem("31. Blastocrithidia Nuclear Code")
            self.gencodebox.addItem("33. Cephalodiscidae Mitochondrial UAA-Tyr Code")
            self.gencodebox.activated[str].connect(self.gencodechoice)
            self.gencode=5
            self.gencodebox.setCurrentIndex(4)
            MinLengthLabel = QtWidgets.QLabel('Minimum Length: ',self.d)
            ExpLengthLabel = QtWidgets.QLabel('Length of barcode: ',self.d)
            randomsubsetLabel = QtWidgets.QLabel('Coverage used: ',self.d)
            n90percsubsetLabel = QtWidgets.QLabel('Coverage used: ',self.d)
        #   transmode=QtWidgets.QLabel('Check translation?: ',self.d)
            demlenlab=QtWidgets.QLabel('Window to define product length (length of barcode +/-): ',self.d)
            primermismatchlab=QtWidgets.QLabel('Number of mismatches for primer search: ',self.d)
            primersearchlab=QtWidgets.QLabel('Window for primer and tag search: ',self.d)
            postdemlenlab=QtWidgets.QLabel('Maximum deviation of read length from barcode length: ',self.d)
            consensesfreqfixedlabel=QtWidgets.QLabel('Main consensus calling frequency: ',self.d)
            consensesfreqrangelabel=QtWidgets.QLabel('Range of frequencies to assess: ',self.d)
            consensesfreqsteplabel=QtWidgets.QLabel('to be examined at step size of ',self.d)
            self.MinLengthTextBox = QtWidgets.QLineEdit()
            self.MinLengthTextBox.setText("658")
            self.ExpLengthTextBox = QtWidgets.QLineEdit()
            self.ExpLengthTextBox.setText("658")
            self.PrimerSearchTextBox = QtWidgets.QLineEdit()
            self.PrimerSearchTextBox.setText("100")
            self.PostdemLengthTextBox = QtWidgets.QLineEdit()
            self.PostdemLengthTextBox.setText("50")
            self.DemLengthTextBox = QtWidgets.QLineEdit()
            self.DemLengthTextBox.setText("100")
            tagmismatchlabel = QtWidgets.QLabel('Tag Mismatch: ',self.d)
            self.combobox1 =  QtWidgets.QComboBox()
            self.combobox1.addItem('0')
            self.combobox1.addItem('1')
            self.combobox1.addItem('2')
            self.combobox1.setCurrentText('2')
            self.combobox1.currentTextChanged.connect(self.tagmismatchchanged)
            self.PrimersearchMismatchTextBox = QtWidgets.QLineEdit()
            self.PrimersearchMismatchTextBox.setText("10")
            self.randomsubsetnTextBox = QtWidgets.QLineEdit()
            self.randomsubsetnTextBox.setText("25,50,100,200,500")
            self.n90percsubsetnTextBox = QtWidgets.QLineEdit()
            self.n90percsubsetnTextBox.setText("100")
            self.consensesfreqfixedTextbox= QtWidgets.QLineEdit()
            self.consensesfreqrangeTextbox= QtWidgets.QLineEdit()
            self.consensesfreqstepTextbox= QtWidgets.QLineEdit()
            self.consensesfreqfixedTextbox.setText("0.3")
            self.consensesfreqrangeTextbox.setText("0.2,0.5")
            self.consensesfreqstepTextbox.setText("0.05")

            self.tagmm=2


            self.minlen=658
            self.explen=658
            self.postdemlen=50
            self.demlen=100
            self.primermismatch=10
            self.primersearchlen=100
            self.consensesfreqfixed=0.3
            self.consensesfreqrange=[0.2,0.5]
            self.consensesfreqstep=0.05
            regex = QtCore.QRegExp("^[0-9]*$")
            validator = QtGui.QRegExpValidator(regex,self.MinLengthTextBox)
            #self.transstat="Yes"
            #self.trans1=QtWidgets.QRadioButton("Yes ",self.d)
            #self.trans2=QtWidgets.QRadioButton("No ",self.d)
            #self.trans1.toggled.connect(self.settrans1)
            #self.trans2.toggled.connect(self.settrans2)
            self.demrunbutton=QtWidgets.QPushButton("Run",self.d)
            self.demrunbutton.clicked.connect(self.setdemoutpath)
            modelabel=QtWidgets.QLabel("Steps to run")
            self.consensuslencheckbox = QtWidgets.QCheckBox("Consensus by length")
            self.consensus90checkbox=QtWidgets.QCheckBox("Consensus by similarity")
            self.fixcheckbox=QtWidgets.QCheckBox("Consensus by barcode comparisons")

            self.consensuslencheckbox.stateChanged.connect(lambda:self.stepcheck(self.consensuslencheckbox))
            self.consensus90checkbox.stateChanged.connect(lambda:self.stepcheck(self.consensus90checkbox))
            self.fixcheckbox.stateChanged.connect(lambda:self.stepcheck(self.fixcheckbox))

            if self.flag==2:
                #self.consensuslencheckbox.setDisabled(True)
                #self.consensus90checkbox.setDisabled(True)
                #self.fixcheckbox.setDisabled(True)
                self.MinLengthTextBox.setDisabled(True)
                self.DemLengthTextBox.setDisabled(True)
                self.PrimerSearchTextBox.setDisabled(True)

            self.consensuslencheckbox.setChecked(True)
            self.consensus90checkbox.setChecked(True)
            self.fixcheckbox.setChecked(True)

            self.consensuslenstat=True
            self.consensus90stat=True
            self.fixstat=True

            if self.flag==3:
                self.MinLengthTextBox.setDisabled(True)
                self.DemLengthTextBox.setDisabled(True)
                self.consensuslencheckbox.setChecked(False)
                self.consensus90checkbox.setChecked(False)
                self.fixcheckbox.setChecked(True)
                self.consensuslencheckbox.setDisabled(True)
                self.consensus90checkbox.setDisabled(True)
                self.fixcheckbox.setDisabled(True)              
                self.randomsubsetnTextBox.setDisabled(True)
                self.n90percsubsetnTextBox.setDisabled(True)
                self.PostdemLengthTextBox.setDisabled(True)
                self.consensuslenstat=False
                self.consensus90stat=False
                self.fixstat=True
                self.consensesfreqfixedTextbox.setDisabled(True)
                self.consensesfreqrangeTextbox.setDisabled(True)
                self.consensesfreqstepTextbox.setDisabled(True)
                self.PrimerSearchTextBox.setDisabled(True)
            #self.trans1.setChecked(True)
            layout.addWidget(demlabel,0,0,1,1)
            layout.addWidget(MinLengthLabel,1,0,1,2)
            layout.addWidget(self.MinLengthTextBox,1,2,1,1)
            layout.addWidget(ExpLengthLabel,2,0,1,2)
            layout.addWidget(self.ExpLengthTextBox,2,2,1,1)
            layout.addWidget(demlenlab,3,0,1,2)
            layout.addWidget(self.DemLengthTextBox,3,2,1,1)

            layout.addWidget(primersearchlab,4,0,1,1)
            layout.addWidget(self.PrimerSearchTextBox,4,2,1,1)
            layout.addWidget(primermismatchlab,4,3,1,1)
            layout.addWidget(self.PrimersearchMismatchTextBox,4,4,1,1)        
            layout.addWidget(tagmismatchlabel,5,0,1,2) 
            layout.addWidget(self.combobox1,5,2,1,1)  
            layout.addWidget(ConsensusByLengthlabel,6,0,1,1)
            layout.addWidget(randomsubsetLabel,7,0,1,2)
            layout.addWidget(self.randomsubsetnTextBox,7,2,1,1)
            layout.addWidget(postdemlenlab,8,0,1,2)
            layout.addWidget(self.PostdemLengthTextBox,8,2,1,1)    
            layout.addWidget(ConsensusBySimlabel,9,0,1,1)          
            layout.addWidget(n90percsubsetLabel,10,0,1,2)
            layout.addWidget(self.n90percsubsetnTextBox,10,2,1,1)

            layout.addWidget(Consensusgeneralsettings,11,0,1,1)     

            layout.addWidget(consensesfreqfixedlabel,12,0,1,2)
            layout.addWidget(self.consensesfreqfixedTextbox,12,2)
            layout.addWidget(consensesfreqrangelabel,13,0,1,2)
            layout.addWidget(self.consensesfreqrangeTextbox,13,2)
            layout.addWidget(consensesfreqsteplabel,14,0,1,2)
            layout.addWidget(self.consensesfreqstepTextbox,14,2)
            layout.addWidget(Geneticcodelabel,15,0,1,2)
            layout.addWidget(self.gencodebox,15,2,1,2)
            layout.addWidget(modelabel,16,0,1,1)
            layout.addWidget(self.consensuslencheckbox,16,2,1,1)
            layout.addWidget(self.consensus90checkbox,16,3,1,1)
            layout.addWidget(self.fixcheckbox,16,4,1,1)


            layout.addWidget(self.demrunbutton,17,4)
            self.d.setLayout(layout)
            self.d.exec_()
    def tagmismatchchanged(self,b):
        self.tagmm=int(b)
    def subsetornot(self):
        if self.subsetmaxcheckbox.isChecked() == True:
            self.subsetmax=True
        else:
            self.subsetmax=False
    def stepcheck(self,b):
        if b.text() == "Consensus by length":
            if b.isChecked() == True:
                self.consensuslenstat=True
                self.consensus90checkbox.setDisabled(False)
                self.fixcheckbox.setDisabled(False)
            else:
                self.consensuslenstat=False
                if self.flag==1:
                    self.consensus90checkbox.setChecked(False)
                    self.fixcheckbox.setChecked(False)
                    self.consensus90stat=False
                    self.fixstat=False
                    self.consensus90checkbox.setDisabled(True)
                    self.fixcheckbox.setDisabled(True)
        if b.text() == "Consensus by similarity":
            if b.isChecked() == True:
                self.consensus90stat=True
            else:
                self.consensus90stat=False  
        if b.text() == "Consensus by barcode comparisons":
            if b.isChecked() == True:
                self.fixstat=True
            else:
                self.fixstat=False
        if b.text()=="Do you want to basecall as well as do barcode calling?":
            if b.isChecked() == True:
                self.basecallingstat=True
                self.liveornot='basecall'
            else:
                self.basecallingstat=False
        if b.text()=="Do you want to transfer files from Mk1C live?":
            if b.isChecked() == True:
                self.remotetransferstat=True
                self.HostTextBox.setDisabled(False)
                self.UsernameTextBox.setDisabled(False)
                self.PasswordTextBox.setDisabled(False)
                self.RemotepathTextbox.setDisabled(False)
            else:
                self.remotetransferstat=False
                self.HostTextBox.setText("")
                self.UsernameTextBox.setText("")
                self.PasswordTextBox.setText("")
                self.RemotepathTextbox.setText("")
        if b.text()=="Do you instead want to specify your own command? <br> Note that the fastq files must be saved in the output folder directly and not under subfolders":
            if b.isChecked() == True:
                self.commandlinestat=True
                self.CmdLineTextBox.setDisabled(False)
            else:
                self.CmdLineTextBox.setPlainText(self.guppycommand)
                self.CmdLineTextBox.setDisabled(True)
    def getreffasta(self):
        self.remreferencefasta = QtWidgets.QFileDialog.getOpenFileName(self, "Select your reference fasta")
    def gencodechoice(self,text):
        self.gencode=int(text.split(".")[0])
    def flowcelltypechoice(self,text):
        self.flowcelltype=str(text.encode(encoding="utf-8"))
      #  print self.flowcelltype
        self.updateguppycommand()
    def setbasecallertype(self):
    #   self.basecallertype=str(self.basecallertype,"utf-8")
      #  print (self.basecallertype)
        if self.basecallertype=='Guppy':
         #   print("g Guppy")
            self.livebasecallingparams()
        elif self.basecallertype=='Dorado':
         #   print ("d Dorado")
            self.livedoradobasecallingparams()
    def basecallertypechoice(self,text):
        self.basecallertype=str(text)
       # print("choice",self.basecallertype)
    def settrans1(self):
        self.transstat="Yes"
    def settrans2(self):
        self.transstat="No"
    def setdemultiplex(self):
        self.mb.close()
        self.demultiplex()
    def demultiplex(self):
        self.inputmode=1
        self.myprepdemultiplex=prepdemultiplex(self.demfile,self.infastq,os.path.join(self.demoutpath,"1_demultiplexing"),self.minlen,self.explen,self.demlen,self.logfile,self.tagmm)
        self.myprepdemultiplex.notifyMessage.connect(self.addlog)
        self.myprepdemultiplex.taskFinished.connect(self.setdemultiplexrun)
        self.runprepdemultiplex()
    def runprepdemultiplex(self):
        self.myprepdemultiplex.start()
    def addlog(self,i):
        self.statuswidget.logstatus.textCursor().insertHtml(i)
    def addloglive(self,i):
    #    print "addloglive triggered",i
        self.totalseqs=i  
        timediff=datetime.datetime.fromtimestamp(int(time.time()))-self.timestamp
        minutes=timediff.total_seconds()/60 
        self.livereadgraphXm=np.append(self.livereadgraphXm,float(minutes)) 
        self.livereadgraphYm=np.append(self.livereadgraphYm,float(i))
        self.livereadgraph.setData(self.livereadgraphXm,self.livereadgraphYm)
        self.livereadgraph.setPos(0,0)
    #       QtGui.QApplication.processEvents()
   #     print i
    def addloglivedem(self,i,j):
     #   print "addloglivedem triggered",i,j
        oldsampleids=deepcopy(self.sampleids)
        self.sampleids=self.myrundemultiplexlive.samplecounts

        #print self.sampleids
        self.ndemultiplexedused=i
        self.ndemultiplexed=j
        timediff=datetime.datetime.fromtimestamp(int(time.time()))-self.timestamp
        minutes=timediff.total_seconds()/60 
        self.livedemgraphXm=np.append(self.livedemgraphXm,float(minutes)) 
        self.livedemgraphYm=np.append(self.livedemgraphYm,float(j)) 
        self.livedemultiplexedgraph.setData(self.livedemgraphXm,self.livedemgraphYm)
        self.livedemultiplexedgraph.setPos(0,0)
        self.livedemusedgraphXm=np.append(self.livedemusedgraphXm,float(minutes)) 
        self.livedemusedgraphYm=np.append(self.livedemusedgraphYm,float(i))
        self.livedemultiplexedusedgraph.setData(self.livedemusedgraphXm,self.livedemusedgraphYm)
        self.livedemultiplexedusedgraph.setPos(0,0) 
     #      QtGui.QApplication.processEvents()  
     #   print "dem",i  
     #   print "demused",j
    def setdemultiplexrun(self):
        self.nseqspasslen=self.myprepdemultiplex.nseqspasslen
        self.sampleids=self.myprepdemultiplex.sampleids
        tagdict=self.myprepdemultiplex.tagdict
        muttags_fr=self.myprepdemultiplex.muttags_fr
        sampledict=self.myprepdemultiplex.sampledict
        typedict=self.myprepdemultiplex.typedict
        primerfset=self.myprepdemultiplex.primerfset
        primerrset=self.myprepdemultiplex.primerrset
        taglen=self.myprepdemultiplex.taglen
        maxid=self.myprepdemultiplex.maxid
        self.lastbitn=self.myprepdemultiplex.lastbitn
        self.totalseqs=self.myprepdemultiplex.totalseqs
        self.nseqsfordemultiplexing=self.myprepdemultiplex.nseqsfordemultiplexing
        self.logfile.write(str(round(time.time())) +": Split your input into parts of 20,000 sequences\n")
        basename=os.path.basename(self.infastq)
        self.statuswidget.pbarTitle.setText('Demultiplexing')
        self.statuswidget.logstatus.textCursor().insertHtml('<br><p style="color:blue;"><b>Phase 1: Demultiplexing</b></p> This step will find the primers and use the adjacent tags to assign the reads to specimen-specific bins with unique tag combinations. The primers and tags are ignored for further processing. <br>Demultiplexed reads are stored in '+ str(os.path.join(self.demoutpath,"demultiplexed"))+".")
        #Builds dictionary of sequences from reformated file
        prefix1=basename+"_reformat_out_1pdt_p"
        prefix2=basename+"_reformat_out_2pdt_p"
    #   print prefix1,prefix2
        typelist=[]
        partlist1=fnmatch.filter(os.listdir(os.path.join(self.demoutpath,"1_demultiplexing")), prefix1+"*")
        partlist2=fnmatch.filter(os.listdir(os.path.join(self.demoutpath,"1_demultiplexing")), prefix2+"*")
        for each in partlist1:
            typelist.append((each,self.primersearchlen))
        for each in partlist2:
            typelist.append((each,self.primersearchlen*2))
                
        partlist=partlist1+partlist2
        def chunker_list(seq, size):
            return [seq[i::size] for i in range(size)]
        nparts=chunker_list(partlist,4)
    #   print "nparts",nparts
        nparts=[[x,os.path.join(self.demoutpath,"1_demultiplexing"),tagdict,muttags_fr,sampledict,typedict,primerfset,primerrset,taglen,i,maxid,self.lastbitn,typelist,self.primermismatch] for i,x in enumerate(nparts)]
        self.resultlist=[]
        def addresult(result):
            self.resultlist.append(result)
        n=0
        self.queue = multiprocessing.Queue()
        self.pool = multiprocessing.Pool(processes=4, initializer=pool_init1, initargs=(self.queue,))
        for i,j in enumerate(nparts):
            self.pool.apply_async(func=rundemultiplex, args=(j,),  callback=addresult)
            os.mkdir(os.path.join(self.demoutpath,"1_demultiplexing",str(i)))
        self.pool.close()
        self.statuswidget.pbar.setMaximum(self.nseqsfordemultiplexing)
        self.statuswidget.pbar.setFormat("%v"+"/"+str(self.nseqsfordemultiplexing))
        self.timer = QtCore.QTimer()
        self.timer.timeout.connect(self.updateProgressdemultiplex)
        self.timer.start(1)
        self.sumprogress=[0]*4
    def updateProgressdemultiplex(self):
        if len(self.resultlist)!=4:
            if self.queue.empty(): return
            num_row, progress,labeltext = self.queue.get()
            self.processlabel.setText(labeltext)
            self.sumprogress[num_row]=progress
        #   print self.sumprogress,num_row,sum(self.sumprogress),self.nseqsfordemultiplexing
            self.statuswidget.pbar.setValue(int(sum(self.sumprogress)))
            self.updateProgressdemultiplex()
        else:
            self.timer.stop()
            self.statuswidget.pbar.setValue(self.nseqsfordemultiplexing)
            self.mymergedatasets=mergedemfiles(self.sampleids,os.path.join(self.demoutpath,"1_demultiplexing"),os.path.join(self.demoutpath,"demultiplexed"))
            self.mymergedatasets.notifyProgress.connect(self.updatepbar)
            self.mymergedatasets.taskFinished.connect(self.printdone)
            self.statuswidget.pbarTitle.setText("Merging your demultiplexed files")
            self.statuswidget.pbar.setMaximum(len(self.sampleids))
            self.statuswidget.pbar.setFormat("%v"+"/"+str(len(self.sampleids)))
            self.runmergedemsets()
    def runmergedemsets(self):
        self.mymergedatasets.start()
    def updatepbar(self,i):
        self.statuswidget.pbar.setValue(i)
    def printdone(self):
        self.logfile.write(str(round(time.time())) +": Demultiplexing is complete\n")
    #   self.statuswidget.logstatus.textCursor().insertHtml("<br>Please wait while we calculate the statistics.")
        with open("runtime",'w') as runtimefile:
            runtimefile.write(str(round(time.time() - self.start_time, 2)))
    #   partlist1=fnmatch.filter(os.listdir(os.path.join(self.demoutpath,"1_demultiplexing")), "*all.fa")
    #   for each in partlist1:
    #       shutil.move(os.path.join(self.demoutpath,"1_demultiplexing",each), os.path.join(self.demoutpath,"demultiplexed",each))
        dirlist=os.listdir(os.path.join(self.demoutpath,"demultiplexed"))
        self.ndemultiplexed=0
        self.nsampledemultiplexed=0
        self.nsampledemultiplexed5=0
        for each in dirlist:
            with open(os.path.join(self.demoutpath,"demultiplexed",each)) as infile:
                l=infile.readlines()
                self.sampleids[each.split("_all.fa")[0]]=len(l)/2
                self.ndemultiplexed+=len(l)/2
                if len(l)/2>=5:
                    self.nsampledemultiplexed5+=1
                self.nsampledemultiplexed+=1
                
        self.logfile.write(str(round(time.time())) +": Demultiplexed reads are stored in "+ str(os.path.join(self.demoutpath,"demultiplexed"))+". \nNumber of reads demultiplexed = "+ str("{:,}".format(self.ndemultiplexed)) +" ("+str('{0:.2f}'.format(float(self.ndemultiplexed)/float(self.nseqsfordemultiplexing)*100))+"%).")
        self.statuswidget.logstatus.textCursor().insertHtml("<br>Number of reads demultiplexed = "+ str("{:,}".format(self.ndemultiplexed)) +" ("+str('{0:.2f}'.format(float(self.ndemultiplexed)/float(self.nseqsfordemultiplexing)*100))+"%).<br>Number of specimen bins with at least 1 read ="+str(self.nsampledemultiplexed)+".<br>Number of specimen bins with at least 5 reads="+str(self.nsampledemultiplexed5))
        demsheet=self.wb.add_worksheet("1. Demultiplexing")
        headers=["SpecimenID","Number of sequences demultiplexed"]
        c=0
        while c<len(headers):
            demsheet.write(0,c,headers[c])
            c+=1
        for i,j in enumerate(self.sampleids.keys()):
            demsheet.write(i+1,0,j)
            demsheet.write(i+1,1,self.sampleids[j])
        self.inlistforconsensus=fnmatch.filter(os.listdir(os.path.join(self.demoutpath,"demultiplexed")),"*")
        if self.consensuslenstat==True:
            self.statuswidget.logstatus.textCursor().insertHtml('<br><b><p style="color:blue;">Phase 2a: Consensus by Length</b></p>Reads are sorted by deviation from the length of the amplicon. The reads with the best length match are selected for alignment whereby the maximum number of reads to consider was specified by the user during setup. The reads are aligned and a consensus barcode is called. It is only accepted as a candidate barcode if it has the correct length, is translatable, and free of ambiguous bases. If the iterative mode was chosen during setup, this process is repeated at increasingly higher coverage for those bins that fail to yield a barcode at lower coverages')
            self.printreadfasta(0)
                                                                    
        else:
            self.statuswidget.logstatus.textCursor().insertHtml("Please find the results in "+os.path.join(self.demoutpath,"demultiplexed") +" folder.")

    def printreadfasta(self,v):
     #   print self.selectlens
        if self.consensuslenstat==True:
            self.nsampledem=len(self.inlistforconsensus)
                                              
            self.statuswidget.pbar.setMaximum(self.nsampledem)
            self.statuswidget.pbar.setFormat("%v"+"/"+str(self.nsampledem))     
            os.mkdir(os.path.join(self.demoutpath,"2a_ConsensusByLength","demultiplexed_"+str(self.selectlens[self.selectlenscounter])))
            os.mkdir(os.path.join(self.demoutpath,"2a_ConsensusByLength","demultiplexed_"+str(self.selectlens[self.selectlenscounter])+"_mafft"))
            self.statuswidget.pbarTitle.setText('Consensus by length, coverage '+str(self.selectlens[self.selectlenscounter]))
            
            self.statuswidget.layout.addWidget(self.statuswidget.pbar,5,1,1,1)
            def chunker_list(seq, size):
                return [seq[i::size] for i in range(size)]
            nparts=chunker_list(self.inlistforconsensus,1)
                        
            self.counter=len(self.inlistforconsensus)
            if v==0:
                nparts=[[x,self.demoutpath,"demultiplexed",os.path.join("2a_ConsensusByLength","demultiplexed_"+str(self.selectlens[self.selectlenscounter])),self.selectlens[self.selectlenscounter],"consensus_"+str(self.selectlens[self.selectlenscounter])+"_barcodes.fa",i,self.explen,v,self.postdemlen,"consensus_by_length",self.consensesfreqfixed,self.consensesfreqrange,self.consensesfreqstep,self.gencode] for i,x in enumerate(nparts)]
            elif v==1:
                nparts=[[x,self.demoutpath,self.indir,os.path.join("2a_ConsensusByLength","demultiplexed_"+str(self.selectlens[self.selectlenscounter])),self.selectlens[self.selectlenscounter],"consensus_"+str(self.selectlens[self.selectlenscounter])+"_barcodes.fa",i,self.explen,v,self.postdemlen,"consensus_by_length",self.consensesfreqfixed,self.consensesfreqrange,self.consensesfreqstep,self.gencode] for i,x in enumerate(nparts)]
            self.myconsensus1=runconsensusparts(nparts[0])
            self.myconsensus1.notifyProgress.connect(self.updatecon1)                          
            self.myconsensus1.taskFinished.connect(self.updateconsensus)
            self.runcon1()
        else:
            self.updateconsensus()
    def runcon1(self):
        self.myconsensus1.start()
    def updatecon1(self,i):
        self.statuswidget.pbar.setValue(i)
    def wait(self):
        self.demoutpath             
    def updateconsensus(self):
                    
        timetokill=False
        self.corlist=[]
        if self.consensuslenstat==True:
            self.statuswidget.pbar.setValue(self.nsampledem)
            self.resultlist=[[self.myconsensus1.transcheck,self.myconsensus1.conseqs,self.myconsensus1.flags,self.myconsensus1.coverages]]                                                 
            if self.selectlenscounter==0:
                self.sampleids=self.myconsensus1.sampleids
            self.inlistforconsensus=[]
            with open(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensus_"+str(self.selectlens[self.selectlenscounter])+"_barcodes.fa"),'a') as outfile:
                with open(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensus"+str(self.selectlens[self.selectlenscounter])+"prederr_barcodes.fa"),'a') as outfile3:
                    for each in self.resultlist:
                        for k in each[0].keys():
                            self.con200trans[k]=each[0][k]
                        for k in each[1].keys():
                            if len(each[1][k])!=0:
                                self.con200barcodes[k]=each[1][k]
                                self.con200length[k]=len(each[1][k])
                        for k in each[3].keys():
                            self.con200cov[k]=each[3][k]
                        for k in each[2].keys():
                            self.con200flags[k]=each[2][k]
                        #    print k
                            if each[2][k]==True:
                                outfile.write(">"+k+"_all.fa"+";"+str(len(each[1][k]))+";"+str(each[3][k])+'\n'+each[1][k]+'\n')
                            #   outfile2.write(">"+k+"_all.fa"+";"+str(len(each[1][k]))+";"+str(each[3][k])+'\n'+each[1][k]+'\n')
                                
                            else:
                                if len(each[1][k])!=0:
             
                                    outfile.write(">"+k+"_all.fa"+";"+str(len(each[1][k]))+";"+str(each[3][k])+'\n'+each[1][k]+'\n')
                                                                                                 
                                    outfile3.write(">"+k+"_all.fa"+";"+str(len(each[1][k]))+";"+str(each[3][k])+'\n'+each[1][k]+'\n')
                                    if self.sampleids[k]>=self.selectlens[self.selectlenscounter]:
                                    #   print k,self.sampleids[k],self.selectlens[self.selectlenscounter]
                                        self.inlistforconsensus.append(k+"_all.fa")
            self.selectlenscounter+=1
            if len(self.inlistforconsensus)>0:
                if self.selectlenscounter!=len(self.selectlens):
                    self.printreadfasta(self.inputmode-1)
                                 
                else:
                    with open(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensus_all_step1.fa"),'a') as outfile:
                                           
                        with open(os.path.join(self.demoutpath,"barcodesets","temps","consensusgood_temp.fa"),'a') as outfile2:
                            with open(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensus_all_prederr_barcodes.fa"),'a') as outfile3:
                                for k in self.con200barcodes.keys():
                                    outfile.write(">"+k+"_all.fa"+";"+str(self.con200length[k])+";"+str(self.con200cov[k])+'\n'+self.con200barcodes[k]+'\n')
                                    if self.con200flags[k]==True:
                                        outfile2.write(">"+k+"_all.fa"+";"+str(self.con200length[k])+";"+str(self.con200cov[k])+'\n'+self.con200barcodes[k]+'\n')
                                        self.corlist.append(k+"_all.fa")
                                        self.ngoodbarcodescounter+=1
                                    else:
                                        outfile3.write(">"+k+"_all.fa"+";"+str(self.con200length[k])+";"+str(self.con200cov[k])+'\n'+self.con200barcodes[k]+'\n')

                    self.msacheck1("consensus_all_step1.fa")
            else:
                with open(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensus_all_step1.fa"),'a') as outfile:
                    with open(os.path.join(self.demoutpath,"barcodesets","temps","consensusgood_temp.fa"),'a') as outfile2:
                        with open(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensus_all_prederr_barcodes.fa"),'a') as outfile3:
                            for k in self.con200barcodes.keys():
                                outfile.write(">"+k+"_all.fa"+";"+str(self.con200length[k])+";"+str(self.con200cov[k])+'\n'+self.con200barcodes[k]+'\n')
                                if self.con200flags[k]==True:
                                    outfile2.write(">"+k+"_all.fa"+";"+str(self.con200length[k])+";"+str(self.con200cov[k])+'\n'+self.con200barcodes[k]+'\n')
                                    self.corlist.append(k+"_all.fa")
                                    self.ngoodbarcodescounter+=1
                                else:
                                    outfile3.write(">"+k+"_all.fa"+";"+str(self.con200length[k])+";"+str(self.con200cov[k])+'\n'+self.con200barcodes[k]+'\n')   
                self.msacheck1("consensus_all_step1.fa")
                            
        else:
            transdict={}
            conseqs={}
            coverages={}
            flags={}
            if self.consensus90stat==False:
                fname=self.forfixfname
            else:
                fname=str(self.forconsensus90fname[0])
            with open(fname) as consensusfile:
                l=consensusfile.readlines()
                for i,j in enumerate(l):
                    if ">" in j:
                        try:
                            conseqs[j.split(";")[0].split("_all.fa")[0][1:]]=l[i+1].strip().upper()
                            coverages[j.split(";")[0].split("_all.fa")[0][1:]]=int(j.split(";")[2])
                            if len(l[i+1].strip().upper())==self.explen:
                                if l[i+1].strip().upper().count("N")==0:
                                    if self.translate_corframe(l[i+1].strip().upper(),self.gencode)=="1":
                                        flags[j.split(";")[0].split("_all.fa")[0][1:]]=True
                                    else:
                                        flags[j.split(";")[0].split("_all.fa")[0][1:]]=False
                                else:
                                    flags[j.split(";")[0].split("_all.fa")[0][1:]]=False
                            else:
                                flags[j.split(";")[0].split("_all.fa")[0][1:]]=False
                           
                            transdict[j.split(";")[0].split("_all.fa")[0][1:]]=self.translate_corframe(l[i+1].strip().upper(),self.gencode)
                        except IndexError:
                            self.showwarningdialog("Your FASTA file looks wrong. It must be generated by the ONTbarcoder pipeline")
                            self.sumstack.setCurrentIndex(0)
                            timetokill=True
                            break                           
            if timetokill==True:
                self.wait()         
            else:
                self.resultlist=[[transdict,conseqs,flags,coverages]]
                if self.flag!=3:
                        
                    self.sampleids=self.mycounterdemreads.counter
                    self.dirdict=self.mycounterdemreads.dirdict
                                                                     
                for each in self.resultlist:
                    for k in each[0].keys():
                        self.con200trans[k]=each[0][k]
                    for k in each[1].keys():
                        if len(each[1][k])!=0:
                            self.con200barcodes[k]=each[1][k]
                            self.con200length[k]=len(each[1][k])
                    for k in each[3].keys():
                        self.con200cov[k]=each[3][k]
                    for k in each[2].keys():
                        self.con200flags[k]=each[2][k]
                with open(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensus_all_step1.fa"),'a') as outfile:
                    with open(os.path.join(self.demoutpath,"barcodesets","temps","consensusgood_temp.fa"),'a') as outfile2:
                        with open(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensus_all_prederr_barcodes.fa"),'a') as outfile3:
                            for k in self.con200barcodes.keys():
                                outfile.write(">"+k+"_all.fa"+";"+str(self.con200length[k])+";"+str(self.con200cov[k])+'\n'+self.con200barcodes[k]+'\n')
                                if self.con200flags[k]==True:
                                    outfile2.write(">"+k+"_all.fa"+";"+str(self.con200length[k])+";"+str(self.con200cov[k])+'\n'+self.con200barcodes[k]+'\n')
                                                                    
                                    self.corlist.append(k+"_all.fa")
                                    self.ngoodbarcodescounter+=1
                                else:
                                    outfile3.write(">"+k+"_all.fa"+";"+str(self.con200length[k])+";"+str(self.con200cov[k])+'\n'+self.con200barcodes[k]+'\n')   
                self.msacheck1(fname)
    def msacheck1(self,fname):
        n90percsubsetn=int(str(self.n90percsubsetnTextBox.text()))
        if self.consensuslenstat==True:
            #self.forconsensus90fname="consensus_all_step1.fa"
            self.logfile.write(str(round(time.time())) +': Consensus calling based on sequences closest to length reads is complete\n\nConsensus sequences are stored in '+ str(os.path.join(self.demoutpath,"barcodesets","consensus","consensus_all_step1.fa"))+".")
            self.statuswidget.logstatus.textCursor().insertHtml('<br><br>Consensus sequences are stored in '+ str(os.path.join(self.demoutpath,"barcodesets","consensus","consensus_all_step1.fa"))+".")
            self.statuswidget.pbarTitle.setText('Identifying correct and incorrect barcodes based on MSA: ')
            self.statuswidget.logstatus.textCursor().insertHtml("<br><br>Application of the 4th barcode quality criterion to the candidate barcodes. They are aligned, but only those are accepted as barcodes that do not cause indels in the MSA. All bins that did not yield an accepted barcode are passed to Phase 2b: Consensus by Similarity.")
        #   self.statuswidget.logstatus.textCursor().insertHtml('Please wait while an MSA is constructed')
            
        else:
            #self.forconsensus90fname=str(self.forconsensus90fname[0])
            self.statuswidget.logstatus.textCursor().insertHtml('<br>Consensus sequences have already been provided: ' +fname+".")
            self.statuswidget.pbarTitle.setText('Identifying correct and incorrect barcodes based on MSA: ')            
        try:
            self.indir
            if self.consensus90stat==True:
                self.mycheckmsa=MSAcheck(self.demoutpath,self.explen,fname,1,"90perc",1,self.indir,"consensusgood_temp.fa","consensus_all_prederr_barcodes.fa",n90percsubsetn,"consensus_by_length",self.ngoodbarcodescounter,self.corlist,"consensusgood",self.dirdict)
            else:
                self.mycheckmsa=MSAcheck(self.demoutpath,self.explen,fname,0,"90perc",1,self.indir,"consensusgood_temp.fa","consensus_all_prederr_barcodes.fa",n90percsubsetn,"consensus_by_length",self.ngoodbarcodescounter,self.corlist,"consensusgood",self.dirdict)
        except AttributeError:
            if self.consensus90stat==True:
                self.mycheckmsa=MSAcheck(self.demoutpath,self.explen,fname,1,"90perc",0,"","consensusgood_temp.fa","consensus_all_prederr_barcodes.fa",n90percsubsetn,"consensus_by_length",self.ngoodbarcodescounter,self.corlist,"consensusgood",{})
            else:
                self.mycheckmsa=MSAcheck(self.demoutpath,self.explen,fname,0,"90perc",0,"","consensusgood_temp.fa","consensus_all_prederr_barcodes.fa",n90percsubsetn,"consensus_by_length",self.ngoodbarcodescounter,self.corlist,"consensusgood",{})
        if self.consensus90stat==True:
            self.mycheckmsa.taskFinished.connect(self.callsecondconsensus)
        else:
            self.mycheckmsa.taskFinished.connect(self.mafft200setup)
        self.mycheckmsa.notifyProgress2.connect(self.updatemsabar11)
        self.mycheckmsa.notifyProgress1.connect(self.updatemsabar12)
        self.mycheckmsa.notifyProgress3.connect(self.updatemsa1label)
        self.mycheckmsa.notifyProgress4.connect(self.updatemsabarlabel)
        self.runcheckmsa1()
    def updatemsabar12(self,i):
        self.statuswidget.pbar.setMaximum(i[1])
        self.statuswidget.pbar.setFormat("%v"+"/"+str(i[1]))
        self.statuswidget.pbar.setValue(i[0])
    def updatemsabar11(self,i):
    #   print i
        self.statuswidget.pbar.setMaximum(i[1])
        self.statuswidget.pbar.setFormat("%v"+"/"+str(i[1]))
        self.statuswidget.pbar.setValue(i[0])
    def updatemsabarlabel(self):
        self.logfile.write(str(round(time.time())) + ": Reading the MSA to split the barcodes post length, translation and internal gap check. \nResults are saved in "+str(os.path.join(self.demoutpath,"barcodesets","predgood","consensus_predgood_barcodes.fa")) + " \nand " +str(os.path.join(self.demoutpath,"barcodesets","pred_erroneous","consensus_prederr_barcodes.fa"))+'\n')
        self.statuswidget.logstatus.textCursor().insertHtml("<br><br>Results are saved in "+str(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensus_predgood_barcodes.fa")) + " and " +str(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensus_prederr_barcodes.fa")))
        self.statuswidget.pbarTitle.setText("Identifying sequences similar to initial consensus (cutoff 90%)")
    #   self.statuswidget.logstatus.textCursor().insertHtml("Here the reads are mapped to consensus in previous steps to identify \nbest quality reads which are based on those similar to first consensus.")
    def updatemsa1label(self,i):
    #   print "MSAdone"
    #   self.statuswidget.logstatus.textCursor().insertHtml("Done")
        self.statuswidget.pbarTitle.setText("Parsing MSA")
    #   self.statuswidget.logstatus.textCursor().insertHtml("<br>Reading the MSA to split the barcodes post length, translation and internal gap check<br>")
    def runcheckmsa1(self):
        self.mycheckmsa.start()

    def callsecondconsensus(self,i):
        self.con200goodn=i
        if self.con200goodn>0:
            self.logfile.write(str(round(time.time())) + ": Reading the MSA to split the barcodes post length, translation and internal gap check. \nResults are saved in "+str(os.path.join(self.demoutpath,"barcodesets","predgood","consensus_predgood_barcodes.fa")) + " \nand " +str(os.path.join(self.demoutpath,"barcodesets","pred_erroneous","consensus_prederr_barcodes.fa"))+'.\nNumber of good barcodes identified in this step='+str(self.con200goodn)+"\n")
            self.statuswidget.logstatus.textCursor().insertHtml("<br><br><b>Number of good barcodes identified in this step="+str(self.con200goodn)+"</b>")
        else:
            self.logfile.write(str(round(time.time())) + ": No good consensus barcode found in first round, trying now with reads closest to the first consensus\n")
            self.statuswidget.logstatus.textCursor().insertHtml("<br>No good consensus barcode found in first round.")
        demsheet=self.wb.add_worksheet("2a Consensus by length")

        headers=["SpecimenID","Number of sequences demultiplexed","stage","length","barcode","translation check"]
        c=0
        while c<len(headers):
            demsheet.write(0,c,headers[c])
            c+=1
        for i,j in enumerate(self.sampleids.keys()):
            demsheet.write(i+1,0,j)
            demsheet.write(i+1,1,self.sampleids[j])
            demsheet.write(i+1,2,"Consensus by length")
            try: 
                demsheet.write(i+1,3,self.con200length[j])
                demsheet.write(i+1,4,self.con200barcodes[j])
                demsheet.write(i+1,5,self.con200trans[j])
            except KeyError:
                demsheet.write(i+1,3,"NA")
                demsheet.write(i+1,4,"NA")
                demsheet.write(i+1,5,"NA")  
        self.n90perc={}
        
        if "90perc" in os.listdir(os.path.join(self.demoutpath,"2b_ConsensusBySimilarity")):
            dirlist=os.listdir(os.path.join(self.demoutpath,"2b_ConsensusBySimilarity","90perc"))
            
            for each in dirlist:
                with open(os.path.join(self.demoutpath,"2b_ConsensusBySimilarity","90perc",each)) as infile:
                    l=infile.readlines()
                    self.n90perc[each.split("_all.fa")[0]]=len(l)/2

            self.statuswidget.pbarTitle.setText("Consensus by Similarity, coverage: "+ str(self.n90percsubsetn))
            self.statuswidget.logstatus.textCursor().insertHtml('<br><b><p style="color:blue;">Phase 2b: Consensus by Similarity</p></b>Those read bins that failed to yield a QC-compliant barcode in Phase 2a are subjected to further analysis. Each bin has a preliminary barcode that is assumed to summarize the main signal in the bin. The preliminary barcode is used to identify and eliminate aberrant reads from the bin that have a similarity of less than 90% to the preliminary barcode. Based on the remaining reads, ONTbarcoder then builds a new alignment using the maximum number of reads specified by the user with the highest similarity scores. Consensus calling and quality control follows the same procedures as under Phase 2a.')
    #       print "sec cons"
            partlist=fnmatch.filter(os.listdir(os.path.join(self.demoutpath,"2b_ConsensusBySimilarity","90perc")),"*")
            self.nsampledem=len(partlist)
        #   os.mkdir(os.path.join(self.demoutpath,"2b_ConsensusBySimilarity","90perc_200"))
            os.mkdir(os.path.join(self.demoutpath,"2b_ConsensusBySimilarity","90perc_mafft"))
            self.statuswidget.pbar.setMaximum(self.nsampledem)
            self.statuswidget.pbar.setFormat("%v"+"/"+str(self.nsampledem))     
            def chunker_list(seq, size):
                return [seq[i::size] for i in range(size)]
        #   self.nsampledem=len(partlist)
            nparts=chunker_list(partlist,1)
             
            self.counter=len(partlist)
            nparts=[[x,os.path.join(self.demoutpath),os.path.join("2b_ConsensusBySimilarity","90perc"),os.path.join("2b_ConsensusBySimilarity","90perc"),0,"90perc_barcodes.fa",i,self.explen,0,self.postdemlen,"consensus_by_similarity",self.consensesfreqfixed,self.consensesfreqrange,self.consensesfreqstep,self.gencode] for i,x in enumerate(nparts)]
            self.myconsensus2=runconsensusparts(nparts[0])
            self.myconsensus2.notifyProgress.connect(self.updatecon2)
            self.myconsensus2.taskFinished.connect(self.updateconsensus2)
            self.runcon2()
        else: 
            self.mafft200setup(0)
    def runcon2(self):
        self.myconsensus2.start()
    def updatecon2(self,i):
        self.statuswidget.pbar.setValue(i)
    def updateconsensus2(self):
           
        self.n90trans={}
        self.n90length={}
        self.n90barcodes={}
        self.n90cov={}
        runmsastat=True
        self.corlist=[]
        if self.consensus90stat==True:
            if "90perc" in os.listdir(os.path.join(self.demoutpath,"2b_ConsensusBySimilarity")):
                self.statuswidget.pbar.setValue(self.nsampledem)
                self.resultlist=[[self.myconsensus2.transcheck,self.myconsensus2.conseqs,self.myconsensus2.flags,self.myconsensus2.coverages]]
                with open(os.path.join(self.demoutpath,"barcodesets","consensus_by_similarity","90perc_barcodes.fa"),'a') as outfile:
                    with open(os.path.join(self.demoutpath,"barcodesets","temps","consensusgood_temp.fa"),'a') as outfile2:
                        with open(os.path.join(self.demoutpath,"barcodesets","consensus_by_similarity","90perc_prederr_barcodes.fa"),'a') as outfile3:
                            for each in self.resultlist:
                                for k in each[0].keys():
                                    self.n90trans[k]=each[0][k]
                                for k in each[1].keys():
                                    if len(each[1][k])!=0:
                                        self.n90barcodes[k]=each[1][k]
                                        self.n90length[k]=len(each[1][k])
                                for k in each[3].keys():
                                    self.n90cov[k]=each[3][k]
                                for k in each[2].keys():
                                    if each[2][k]==True:
                                        outfile.write(">"+k+"_all.fa"+";"+str(len(each[1][k]))+";"+str(each[3][k])+'\n'+each[1][k]+'\n')
                                        outfile2.write(">"+k+"_all.fa"+";"+str(len(each[1][k]))+";"+str(each[3][k])+'\n'+each[1][k]+'\n')
                                        self.ngoodbarcodescounter+=1
                                        self.corlist.append(k+"_all.fa")
                                    else:
                                        if len(each[1][k])!=0:
                                            outfile.write(">"+k+"_all.fa"+";"+str(len(each[1][k]))+";"+str(each[3][k])+'\n'+each[1][k]+'\n')
                                            outfile3.write(">"+k+"_all.fa"+";"+str(len(each[1][k]))+";"+str(each[3][k])+'\n'+each[1][k]+'\n')
                self.msacheck2()
                                                                                                    
                     
                                         
            else:
                self.mafft200setup(0)
        else:
            self.mafft200setup(0)

    def msacheck2(self):
        self.logfile.write(str(round(time.time())) +": Second consensus calling completed\n")
        self.statuswidget.pbarTitle.setText("Identifying correct and incorrect barcodes based on MSA")
        self.statuswidget.logstatus.textCursor().insertHtml("<br><br>Consensus sequences are stored in "+ str(os.path.join(self.demoutpath,"barcodesets","consensus_by_similarity","90perc_barcodes.fa"))+".")
        self.statuswidget.logstatus.textCursor().insertHtml("<br><br>Application of the 4th barcode quality criterion to the candidate barcodes. They are aligned, but only those are accepted as barcodes that do not cause indels in the MSA. All bins that did not yield an accepted barcode are passed to Phase 3: Barcode fixing through Consensus by Barcode Comparisons, if enabled")
                       
        self.mycheckmsa2=MSAcheck(self.demoutpath,self.explen,"90perc_barcodes.fa",2,"90perc",0,"","consensusgood_temp.fa","90perc_prederr_barcodes.fa",0,"consensus_by_similarity",self.ngoodbarcodescounter,self.corlist,"90perc",{})
        self.mycheckmsa2.taskFinished.connect(self.mafft200setup)
        self.mycheckmsa2.notifyProgress2.connect(self.updatemsabar21)
        self.mycheckmsa2.notifyProgress3.connect(self.updatemsa2label)
        self.runcheckmsa2()
    def updatemsabar21(self,i):
        self.statuswidget.pbar.setMaximum(i[1])
        self.statuswidget.pbar.setFormat("%v"+"/"+str(i[1]))
        self.statuswidget.pbar.setValue(i[0])
    def updatemsa2label(self,i):
    #   self.statuswidget.logstatus.textCursor().insertHtml("Done")
        self.statuswidget.pbarTitle.setText("Parsing MSA:")
        #self.statuswidget.logstatus.textCursor().insertHtml("<br><br>Reading the MSA to split the barcodes post length, translation and internal gap check")
    def runcheckmsa2(self):
        self.mycheckmsa2.start()
    def mafft200setup(self,i):
        if self.consensus90stat==False:
            self.con200goodn=i

        def builddict_sequences(infile):
            seqdict={}
            with open(infile) as inseqs:
                l=inseqs.readlines()
                for i,j in enumerate(l):
                    if ">" in j:
                        seqdict[j.strip().replace(">","")]=l[i+1].strip().replace("-","").upper()
            return seqdict
        con90done=True
        if self.consensus90stat==True:
            if "90perc" in os.listdir(os.path.join(self.demoutpath,"2b_ConsensusBySimilarity")):
                self.n90goodn=i
                self.logfile.write(str(round(time.time())) +": Reading the MSA to split the barcodes post length, translation and internal gap check. \nResults are saved in "+str(os.path.join(self.demoutpath,"barcodesets","consensus_by_similarity","90perc_predgood_barcodes.fa")) + " \nand " +str(os.path.join(self.demoutpath,"barcodesets","consensus_by_similarity","90perc_prederr_barcodes.fa"))+"\nNumber of good barcodes based on 90% matching="+str(self.n90goodn)+"\n")
                self.statuswidget.logstatus.textCursor().insertHtml("<br><br>Results are saved in "+str(os.path.join(self.demoutpath,"barcodesets","consensus_by_similarity","90perc_predgood_barcodes.fa")) + " <br>and " +str(os.path.join(self.demoutpath,"barcodesets","consensus_by_similarity","90perc_prederr_barcodes.fa"))+"<br><br><b>Number of good barcodes identified in this step="+str(self.n90goodn)+"</b><br>")
                demsheet=self.wb.add_worksheet("2b Consensus by similarity")
                headers=["SpecimenID","Coverage for closest sequences(90% cutoff)","stage","length","barcode","translation check"]
                c=0
                while c<len(headers):
                    demsheet.write(0,c,headers[c])
                    c+=1
                for i,j in enumerate(self.sampleids.keys()):
                    demsheet.write(i+1,0,j)
                    try:
                        demsheet.write(i+1,1,self.n90perc[j])
                        demsheet.write(i+1,2,"Consensus by similarity")
                    except KeyError:
                        demsheet.write(i+1,1,"NA")
                        demsheet.write(i+1,2,"NA")
                    try:
                        demsheet.write(i+1,3,self.n90length[j])
                    except KeyError:
                        demsheet.write(i+1,3,"NA")
                    try:
                        demsheet.write(i+1,4,self.n90barcodes[j])
                    except KeyError:
                        demsheet.write(i+1,4,"NA")
                    try:
                        demsheet.write(i+1,5,self.n90trans[j])
                    except KeyError:
                        demsheet.write(i+1,5,"NA")      
                rset,pset,rgoodset,pgoodset=builddict_sequences(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensus_all_step1.fa")),builddict_sequences(os.path.join(self.demoutpath,"barcodesets","consensus_by_similarity","90perc_barcodes.fa")),builddict_sequences(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensusgood_predgood_barcodes.fa")),builddict_sequences(os.path.join(self.demoutpath,"barcodesets","consensus_by_similarity","90perc_predgood_barcodes.fa"))
                gooddict={}
                self.i=0
                with open(os.path.join(self.demoutpath,"barcodesets","temps","consensus_90perc_predgood_combined_barcodes.fa"),'w') as outfile:
                    donelist=[]
                    for k in pgoodset.keys():
                        outfile.write(">"+k+'\n'+pgoodset[k]+'\n')
                        gooddict[k.split(";")[0]]=pgoodset[k]
                        donelist.append(k.split(";")[0])
                    for k in rgoodset.keys():
                        if k.split(";")[0] not in donelist:
                            outfile.write(">"+k+'\n'+rgoodset[k]+'\n')
                            gooddict[k.split(";")[0]]=rgoodset[k]
                psetids={}
                for k in pset.keys():
                    id=k.split(';')[0]
                    psetids[id]=k
                    if id not in gooddict.keys():
                        self.n90errn+=1
                
                with open(os.path.join(self.demoutpath,"barcodesets","temps","consensus_90perc_prederr_combined_barcodes.fa"),'w') as outfile:
                    for k in rset.keys():
                        id=k.split(';')[0]
                        if id not in gooddict.keys():
                            self.con200errn+=1
                            if id not in psetids.keys():
                                outfile.write(">"+k+";200random"+'\n'+rset[k]+'\n')
                                
                            else:
                                outfile.write(">"+psetids[id]+";90perc"+'\n'+pset[psetids[id]]+'\n')
            else:
                con90done=False
        else:
            con90done=False
        if con90done==False:
            self.statuswidget.logstatus.textCursor().insertHtml("<br><br>Consensus by similarity not performed.<br>")
            rset,rgoodset=builddict_sequences(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensus_all_step1.fa")),builddict_sequences(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensusgood_predgood_barcodes.fa"))
            gooddict={}
            self.i=0
            with open(os.path.join(self.demoutpath,"barcodesets","temps","consensus_90perc_predgood_combined_barcodes.fa"),'w') as outfile:
                donelist=[]
                for k in rgoodset.keys():
                    if k.split(";")[0] not in donelist:
                        outfile.write(">"+k+'\n'+rgoodset[k]+'\n')
                        gooddict[k.split(";")[0]]=rgoodset[k]

            with open(os.path.join(self.demoutpath,"barcodesets","temps","consensus_90perc_prederr_combined_barcodes.fa"),'w') as outfile:
                for k in rset.keys():
                    id=k.split(';')[0]
                    if id not in gooddict.keys():
                        self.con200errn+=1
                        outfile.write(">"+k+";200random"+'\n'+rset[k]+'\n')
        if self.fixstat==True:
            self.fixbarcodessetup()
        else:
            self.printoutputs()
    def fixbarcodessetup(self):
        def translate_corframe(seq,gencode):
            translatedset=[]
            each=seq
            corframe=get_cor_frame(each.replace("-",""),gencode)
            if corframe==1:
                translation=Seq(each.replace('-','')).translate(table=gencode,to_stop=True).__str__()
                explen=len(each.replace('-',''))
            if corframe==2:
                translation=Seq(each[1:].replace('-','')).translate(table=gencode,to_stop=True).__str__()
                explen=len(each[1:].replace('-',''))
            if corframe==3:
                translation=Seq(each[2:].replace('-','')).translate(table=gencode,to_stop=True).__str__()
                explen=len(each[2:].replace('-',''))
            if corframe==4:
                translation=Seq(each.replace('-','')).reverse_complement().translate(table=gencode,to_stop=True).__str__()
                explen=len(Seq(each.replace('-','')).__str__())
            if corframe==5:
                translation=Seq(each[:-1].replace('-','')).reverse_complement().translate(table=gencode,to_stop=True).__str__()
                explen=len(Seq(each[:-1].replace('-','')).__str__())
            if corframe==6:
                translation=Seq(each[:-2].replace('-','')).reverse_complement().translate(table=gencode,to_stop=True).__str__()
                explen=len(Seq(each[:-2].replace('-','')).__str__())
            if len(translation)<int(explen/3):
                return "0"
            elif len(translation)==int(explen/3):
                return "1"
                                    
        def get_cor_frame(seq,gencode):
            seqset=[Seq(seq),Seq(seq[1:]),Seq(seq[2:]),Seq(seq).reverse_complement(),Seq(seq[:-1]).reverse_complement(),Seq(seq[:-2]).reverse_complement()]
            aminoset=[]
            for each in seqset:
                a=each.translate(table=gencode,to_stop=True)
                aminoset.append(a.__str__())
            maxlen,corframe=0,0
            for i,each in enumerate(aminoset):
                if len(each)>maxlen:
                    maxlen=len(each)
                    corframe=i+1
            return corframe
                        
        self.statuswidget.pbarTitle.setText("Consensus by barcode comparisons")
        self.statuswidget.logstatus.textCursor().insertHtml('<br><b><p style="color:blue;">Phase 3: Barcode fixing through Consensus by Barcode Comparisons</p></b><br>Preliminary barcodes that failed the QC criteria applied during Phase 2a and 2b are fixed.  20 QC-compliant barcodes are found that have the highest similarity to each of the preliminary barcodes that failed QC. Each preliminary barcode is aligned to its 20 best-matching QC-compliant barcodes and indels in the preliminary barcode are identified and fixed according to the criteria described in the manuscript.')
        self.demsheet=self.wb.add_worksheet("3.Final barcodes")

        headers=["SpecimenID","Number of sequences demultiplexed","Number used for generating barcodes","stage","type","length","barcode","translation check","#ambiguities"]
        c=0
        while c<len(headers):
            self.demsheet.write(0,c,headers[c])
            c+=1
        def builddict_sequences(infile):
            seqdict={}
            with open(infile) as inseqs:
                l=inseqs.readlines()
                for i,j in enumerate(l):
                    if ">" in j:
                        seqdict[j.strip().replace(">","")]=l[i+1].strip().replace("-","").upper()
                                
            return seqdict
        try:
            musclergoodset,musclepgoodset=builddict_sequences(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensusgood_predgood_barcodes.fa")),builddict_sequences(os.path.join(self.demoutpath,"barcodesets","consensus_by_similarity","90perc_predgood_barcodes.fa"))
            gooddict={}
            self.i=0
            with open(os.path.join(self.demoutpath,"barcodesets","Final_predgood_combined_barcodes.fa"),'w') as outfile:
                with open(os.path.join(self.demoutpath,"barcodesets","Final_all_combined_barcodes.fa"),'w') as outfile2:
                    for k in musclergoodset.keys():
                        outfile.write(">"+k.split(";estgaps=")[0]+";ambs="+str(self.con200barcodes[k.split("_all.fa")[0]].count("N"))+";estgaps="+k.split(";estgaps=")[1]+'\n'+musclergoodset[k]+'\n')
                        outfile2.write(">"+k.split(";estgaps=")[0]+";ambs="+str(self.con200barcodes[k.split("_all.fa")[0]].count("N"))+";estgaps="+k.split(";estgaps=")[1]+'\n'+musclergoodset[k]+'\n')
                        gooddict[k.split(";")[0]]=musclergoodset[k]
                        self.demsheet.write(self.i+1,0,k.split("_all.fa")[0])
                        self.demsheet.write(self.i+1,1,self.sampleids[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,2,self.con200cov[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,3,"Consensus by length")
                        self.demsheet.write(self.i+1,4,"correct")
                        self.demsheet.write(self.i+1,5,self.con200length[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,6,self.con200barcodes[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,7,self.con200trans[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,8,self.con200barcodes[k.split("_all.fa")[0]].count("N"))
                        self.i+=1
                    for k in musclepgoodset.keys():
                        outfile.write(">"+k.split(";estgaps=")[0]+";ambs="+str(self.n90barcodes[k.split("_all.fa")[0]].count("N"))+";estgaps="+k.split(";estgaps=")[1]+'\n'+musclepgoodset[k]+'\n')
                        outfile2.write(">"+k.split(";estgaps=")[0]+";ambs="+str(self.n90barcodes[k.split("_all.fa")[0]].count("N"))+";estgaps="+k.split(";estgaps=")[1]+'\n'+musclepgoodset[k]+'\n')
                        gooddict[k.split(";")[0]]=musclepgoodset[k]
                        self.demsheet.write(self.i+1,0,k.split("_all.fa")[0])
                        self.demsheet.write(self.i+1,1,self.sampleids[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,2,self.n90cov[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,3,"Consensus by similarity")
                        self.demsheet.write(self.i+1,4,"correct")
                        self.demsheet.write(self.i+1,5,self.n90length[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,6,self.n90barcodes[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,7,self.n90trans[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,8,self.n90barcodes[k.split("_all.fa")[0]].count("N"))
                        self.i+=1
        except IOError:
            musclergoodset=builddict_sequences(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensusgood_predgood_barcodes.fa"))
            gooddict={}
            self.i=0
            with open(os.path.join(self.demoutpath,"barcodesets","Final_predgood_combined_barcodes.fa"),'w') as outfile:
                with open(os.path.join(self.demoutpath,"barcodesets","Final_all_combined_barcodes.fa"),'w') as outfile2:
                    for k in musclergoodset.keys():
                        outfile.write(">"+k.split(";estgaps=")[0]+";ambs="+str(self.con200barcodes[k.split("_all.fa")[0]].count("N"))+";estgaps="+k.split(";estgaps=")[1]+'\n'+musclergoodset[k]+'\n')
                        outfile2.write(">"+k.split(";estgaps=")[0]+";ambs="+str(self.con200barcodes[k.split("_all.fa")[0]].count("N"))+";estgaps="+k.split(";estgaps=")[1]+'\n'+musclergoodset[k]+'\n')
                        gooddict[k.split(";")[0]]=musclergoodset[k]
                        self.demsheet.write(self.i+1,0,k.split("_all.fa")[0])
                        try:
                            self.demsheet.write(self.i+1,1,self.sampleids[k.split("_all.fa")[0]])
                        except AttributeError:
                            pass
                        self.demsheet.write(self.i+1,2,self.con200cov[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,3,"Consensus by length")
                        self.demsheet.write(self.i+1,4,"correct")
                        self.demsheet.write(self.i+1,5,self.con200length[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,6,self.con200barcodes[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,7,self.con200trans[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,8,self.con200barcodes[k.split("_all.fa")[0]].count("N"))
                        self.i+=1

        self.seqdict,seqdict2={},{}
        refseqdict,refseqdict2={},{}
        self.hapiddict={}
        self.errbarcodeset={}
        with open(os.path.join(self.demoutpath,"barcodesets","temps","consensus_90perc_prederr_combined_barcodes.fa")) as infile:
            l=infile.readlines()
            for i,j in enumerate(l):
                if ">" in j:
                    samplid=j[1:].split(";")[0]
                    self.errbarcodeset[samplid]=j+l[i+1]
                    if samplid not in gooddict.keys():
                        try:
                            seqdict2[l[i+1].strip()].append(j.strip()[1:])
                            self.hapiddict[j.strip()[1:]]=seqdict2[l[i+1].strip()][0]
                        except KeyError:
                            seqdict2[l[i+1].strip()]=[j.strip()[1:]]
                            self.hapiddict[j.strip()[1:]]=seqdict2[l[i+1].strip()][0]
        with open(os.path.join(self.demoutpath,"barcodesets","Final_predgood_combined_barcodes.fa")) as infile:
            l=infile.readlines()
            for i,j in enumerate(l):
                if ">" in j:
                    try:
                        refseqdict2[l[i+1].strip()].append(j.strip()[1:])
                    except KeyError:
                        refseqdict2[l[i+1].strip()]=[j.strip()[1:]]
        for each in seqdict2.keys():
            m=seqdict2[each]
            m.sort()
            m=["tofix-"+m[0]]+m
            self.seqdict[each]=m
        for each in refseqdict2.keys():
            m=refseqdict2[each]
            m.sort()
            refseqdict[each]=m
        def chunker_list(seq, size):
            return [seq[i::size] for i in range(size)]
        partlist=list(self.seqdict.keys())
        self.nsamplefix=len(partlist)
        self.statuswidget.pbar.setMaximum(self.nsamplefix)
        self.statuswidget.pbar.setFormat("%v"+"/"+str(self.nsamplefix))
        nparts=chunker_list(partlist,1)
        nparts=[[x,self.seqdict,refseqdict,os.path.join(self.demoutpath,"3_ConsensusByBarcodeComparison"),i] for i,x in enumerate(nparts)]
        self.myfix=runtoptwenty(nparts[0])
        self.myfix.notifyProgress.connect(self.updatefix)
        self.myfix.taskFinished.connect(self.updatefixbarcodes)
        self.runfix()
    def runfix(self):
        self.myfix.start()
    def updatefix(self,i):
        self.statuswidget.pbar.setValue(i)
    def updatefixbarcodes(self):
        self.statuswidget.pbar.setValue(self.nsamplefix)
        self.fixbarcodes()
    def change_ext_gaps(self,sequence):
        bps_base=['A','T','G','C','N']
        start_pos, end_pos = 0, 0
        for i,bp in enumerate(sequence):
            if bp in bps_base:
                start_pos = i - 1
                break
        for i,bp in enumerate(sequence[:: - 1]):
            if bp in bps_base:
                end_pos = len(sequence) - i
                break
        new_sequence = "?" * (start_pos + 1) + sequence[start_pos + 1 : end_pos] + "?" * (len(sequence) - end_pos)
        return new_sequence
    def fixbarcodes(self):
        self.logfile.write(str(round(time.time())) +": Alignments for fixing barcodes is complete\n")
        def consensus(indict,perc_thresh):
            poslist=[]
            n=0
            while n<len(list(indict.values())[0]):
                newlist=[]
                for i,each in enumerate(indict.keys()):
                    try:
                        newlist.append(indict[each][n])
                    except IndexError:
                    #   print each,indict[each],indict.values()[0],len(indict[each]),len(indict.values()[0])
                        break
                poslist.append(newlist)
                n+=1
            sequence=[]
            countpos=1
            for character in poslist:
                charcounter=Counter(character)
                baseset={}
                for k,v in charcounter.items():
                    percbp=float(v)/float(len(character))
                    if percbp>perc_thresh:
                        baseset[k]=v
                if len(baseset)==0:
                    bp='N'
                if len(baseset)==1:
                    bp=list(baseset.keys())[0]
                if len(baseset)>1:
                    bp="N"
                sequence.append(bp)
                countpos+=1
            return ''.join(sequence)

        def callconsensus(i,perc_thresh):
            with open(i) as infile:
                l=infile.readlines()
                seqdict,poslist={},[]
                for i,j in enumerate(l):
                    if ">" in j:
                        poslist.append(i)
                for i,j in enumerate(poslist):
                    ambcounts=0
                    k01=l[j].strip().split('>')[1]
                    if i!=len(poslist)-1:
                        k3=l[j+1:poslist[i+1]]
                    if i==len(poslist)-1:
                        k3=l[j+1:]
                    k4=''.join(k3).replace('\n','')
                    if "tofix-" not in k01:
                        seqdict[k01]=k4
                    else:
                        refseq=k4
                conseq=consensus(seqdict,perc_thresh)
                return conseq,seqdict,refseq
        def translate_corframe(seq,gencode):
            translatedset=[]
            each=seq
            corframe=get_cor_frame(each.replace("-",""),gencode)
            if corframe==1:
                translation=Seq(each.replace('-','')).translate(table=gencode,to_stop=True).__str__()
                explen=len(each.replace('-',''))
            if corframe==2:
                translation=Seq(each[1:].replace('-','')).translate(table=gencode,to_stop=True).__str__()
                explen=len(each[1:].replace('-',''))
            if corframe==3:
                translation=Seq(each[2:].replace('-','')).translate(table=gencode,to_stop=True).__str__()
                explen=len(each[2:].replace('-',''))
            if corframe==4:
                translation=Seq(each.replace('-','')).reverse_complement().translate(table=gencode,to_stop=True).__str__()
                explen=len(Seq(each.replace('-','')).__str__())
            if corframe==5:
                translation=Seq(each[:-1].replace('-','')).reverse_complement().translate(table=gencode,to_stop=True).__str__()
                explen=len(Seq(each[:-1].replace('-','')).__str__())
            if corframe==6:
                translation=Seq(each[:-2].replace('-','')).reverse_complement().translate(table=gencode,to_stop=True).__str__()
                explen=len(Seq(each[:-2].replace('-','')).__str__())
            if len(translation)<int(explen/3):
                return "0"
            elif len(translation)==int(explen/3):
                return "1"

        def get_cor_frame(seq,gencode):
            seqset=[Seq(seq),Seq(seq[1:]),Seq(seq[2:]),Seq(seq).reverse_complement(),Seq(seq[:-1]).reverse_complement(),Seq(seq[:-2]).reverse_complement()]
            aminoset=[]
            for each in seqset:
                a=each.translate(table=gencode,to_stop=True)
                aminoset.append(a.__str__())
            maxlen,corframe=0,0
            for i,each in enumerate(aminoset):
                if len(each)>maxlen:
                    maxlen=len(each)
                    corframe=i+1
            return corframe
        dirlist=fnmatch.filter(os.listdir(os.path.join(self.demoutpath,"3_ConsensusByBarcodeComparison")),"*aln.fa")
        pairs={}
        with open(os.path.join(self.demoutpath,"barcodesets","fixing","fixedbarcodes.fa"),'w') as gfile:
            for each in dirlist:
                try:
                    conseq,seqdict,refseq=callconsensus(os.path.join(self.demoutpath,"3_ConsensusByBarcodeComparison",each),0.5)
                    seqdictgood={}
                    seqdictbad={}
                    errcount=0
                    newseq=''
            #       print refseq
                    refseq=self.change_ext_gaps(refseq.upper())
                    conseq=conseq.upper()
                    for i,j in enumerate(refseq):
                        
                        if j=="-":
                            if conseq[i]!="-":
                                errcount+=1
                                newseq+="N"
                                
                        elif j!="?":
                            if conseq[i]=="-":
                                errcount+=1
                            else:
                                newseq+=j
                #   print newseq
                    newseq=newseq.upper()
                    if translate_corframe(newseq,self.gencode)=="1":
                        for barcode in self.seqdict[refseq.replace("-","").replace("?","")][1:]:
                            self.demsheet.write(self.i+1,0,barcode.split("_all.fa")[0])
                            try:
                                self.demsheet.write(self.i+1,1,self.sampleids[barcode.split("_all.fa")[0]])
                            except AttributeError:
                                pass
                            
                            try:
                                self.demsheet.write(self.i+1,2,self.n90cov[barcode.split("_all.fa")[0]])
                                self.demsheet.write(self.i+1,3,"Consensus by similarity, fixed indel")
                                gfile.write(">"+barcode.split(";")[0]+";"+str(len(newseq))+";"+str(self.n90cov[barcode.split("_all.fa")[0]])+";ambs="+str(newseq.count("N"))+";estgaps="+str(errcount)+'\n'+newseq+'\n')
                            except:
                                self.demsheet.write(self.i+1,2,self.con200cov[barcode.split("_all.fa")[0]])
                                self.demsheet.write(self.i+1,3,"Consensus by length, fixed indel")
                                gfile.write(">"+barcode.split(";")[0]+";"+str(len(newseq))+";"+str(self.con200cov[barcode.split("_all.fa")[0]])+";ambs="+str(newseq.count("N"))+";estgaps="+str(errcount)+'\n'+newseq+'\n')
                            self.demsheet.write(self.i+1,4,"removed "+str(errcount)+" indels")
                            self.demsheet.write(self.i+1,5,str(len(newseq)))
                            self.demsheet.write(self.i+1,6,newseq)
                            self.demsheet.write(self.i+1,7,"1")
                            self.demsheet.write(self.i+1,8,newseq.count("N"))
                            self.i+=1
                except IndexError:
                    pass
            ambiguity_codes=[("R", "A"), ("R", "G"),("M", "A"), ("M", "C"),("S", "C"), ("S", "G"),("Y", "C"), ("Y", "T"),("K", "G"), ("K", "T"),("W", "A"), ("W", "T"),("V", "A"), ("V", "C"),("V", "G"),("H", "A"),("H", "C"),("H", "T"),("D", "A"), ("D", "G"),("D", "T"), ("B", "C"),("B", "G"), ("B", "T"),("N", "A"), ("N", "G"), ("N", "C"), ("N", "T")]
        self.wb.close()
        self.logfile.write(str(round(time.time())) +": Fixed barcodes are generated\n")
        with open(os.path.join(self.demoutpath,"barcodesets","fixing","fixedbarcodes.fa")) as infile2:
            with open(os.path.join(self.demoutpath,"barcodesets","Final_all_combined_barcodes.fa"),'a') as outfile:
                l2=infile2.readlines()
                for each in l2:
                    outfile.write(each)
                    if ">" in each:
                        self.nfixed+=1
        self.statuswidget.pbarTitle.setText("Printing outputs")
    #   self.statuswidget.logstatus.textCursor().insertHtml("Please wait while the output files are printed")
        self.printoutputs1()            
    def printoutputs(self):
        self.statuswidget.pbarTitle.setText("Printing outputs")
        self.statuswidget.logstatus.textCursor().insertHtml("Please wait while the output files are printed")
        self.demsheet=self.wb.add_worksheet("6.Final barcodes")

        headers=["SpecimenID","Number of sequences demultiplexed","Number used for generating barcodes","stage","type","length","barcode","translation check","#ambiguities"]
        c=0
        while c<len(headers):
            self.demsheet.write(0,c,headers[c])
            c+=1
        def builddict_sequences(infile):
            seqdict={}
            with open(infile) as inseqs:
                l=inseqs.readlines()
                for i,j in enumerate(l):
                    if ">" in j:
                        seqdict[j.strip().replace(">","")]=l[i+1].strip().replace("-","").upper()
            return seqdict
        try:
            musclergoodset,musclepgoodset=builddict_sequences(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensusgood_predgood_barcodes.fa")),builddict_sequences(os.path.join(self.demoutpath,"barcodesets","consensus_by_similarity","90perc_predgood_barcodes.fa"))
            gooddict={}
            self.i=0
            with open(os.path.join(self.demoutpath,"barcodesets","Final_predgood_combined_barcodes.fa"),'w') as outfile:
                with open(os.path.join(self.demoutpath,"barcodesets","Final_all_combined_barcodes.fa"),'w') as outfile2:
                    for k in musclergoodset.keys():
                        outfile.write(">"+k.split(";estgaps=")[0]+";ambs="+str(self.con200barcodes[k.split("_all.fa")[0]].count("N"))+";estgaps="+k.split(";estgaps=")[1]+'\n'+musclergoodset[k]+'\n')
                        outfile2.write(">"+k.split(";estgaps=")[0]+";ambs="+str(self.con200barcodes[k.split("_all.fa")[0]].count("N"))+";estgaps="+k.split(";estgaps=")[1]+'\n'+musclergoodset[k]+'\n')
                        gooddict[k.split(";")[0]]=musclergoodset[k]
                        self.demsheet.write(self.i+1,0,k.split("_all.fa")[0])
                    #   try:
                        self.demsheet.write(self.i+1,1,self.sampleids[k.split("_all.fa")[0]])
                    #   except AttributeError:
                    #       pass
                        
                        self.demsheet.write(self.i+1,2,self.con200cov[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,3,"Consensus by length")
                        self.demsheet.write(self.i+1,4,"correct")
                        self.demsheet.write(self.i+1,5,self.con200length[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,6,self.con200barcodes[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,7,self.con200trans[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,8,self.con200barcodes[k.split("_all.fa")[0]].count("N"))
                        self.i+=1
                    for k in musclepgoodset.keys():
                        outfile.write(">"+k.split(";estgaps=")[0]+";ambs="+str(self.n90barcodes[k.split("_all.fa")[0]].count("N"))+";estgaps="+k.split(";estgaps=")[1]+'\n'+musclepgoodset[k]+'\n')
                        outfile2.write(">"+k.split(";estgaps=")[0]+";ambs="+str(self.n90barcodes[k.split("_all.fa")[0]].count("N"))+";estgaps="+k.split(";estgaps=")[1]+'\n'+musclepgoodset[k]+'\n')
                        gooddict[k.split(";")[0]]=musclepgoodset[k]
                        self.demsheet.write(self.i+1,0,k.split("_all.fa")[0])
                    #   try:
                        self.demsheet.write(self.i+1,1,self.sampleids[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,2,self.n90cov[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,3,"Consensus by similarity")
                        self.demsheet.write(self.i+1,4,"correct")
                        self.demsheet.write(self.i+1,5,self.n90length[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,6,self.n90barcodes[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,7,self.n90trans[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,8,self.n90barcodes[k.split("_all.fa")[0]].count("N"))
                        self.i+=1
        except IOError:
            musclergoodset=builddict_sequences(os.path.join(self.demoutpath,"barcodesets","consensus_by_length","consensusgood_predgood_barcodes.fa"))
            gooddict={}
            self.i=0
            with open(os.path.join(self.demoutpath,"barcodesets","Final_predgood_combined_barcodes.fa"),'w') as outfile:
                with open(os.path.join(self.demoutpath,"barcodesets","Final_all_combined_barcodes.fa"),'w') as outfile2:
                    for k in musclergoodset.keys():
                        outfile.write(">"+k.split(";estgaps=")[0]+";ambs="+str(self.con200barcodes[k.split("_all.fa")[0]].count("N"))+";estgaps="+k.split(";estgaps=")[1]+'\n'+musclergoodset[k]+'\n')
                        outfile2.write(">"+k.split(";estgaps=")[0]+";ambs="+str(self.con200barcodes[k.split("_all.fa")[0]].count("N"))+";estgaps="+k.split(";estgaps=")[1]+'\n'+musclergoodset[k]+'\n')
                        gooddict[k.split(";")[0]]=musclergoodset[k]
                        self.demsheet.write(self.i+1,0,k.split("_all.fa")[0])
                        try:
                            self.demsheet.write(self.i+1,1,self.sampleids[k.split("_all.fa")[0]])
                        except AttributeError:
                            pass
                        self.demsheet.write(self.i+1,2,self.con200cov[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,3,"Consensus by length")
                        self.demsheet.write(self.i+1,4,"correct")
                        self.demsheet.write(self.i+1,5,self.con200length[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,6,self.con200barcodes[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,7,self.con200trans[k.split("_all.fa")[0]])
                        self.demsheet.write(self.i+1,8,self.con200barcodes[k.split("_all.fa")[0]].count("N"))
                        self.i+=1
        self.printoutputs1()
    def printoutputs1(self):
        self.nfinal=0
        self.errbarcodeset={}
        with open(os.path.join(self.demoutpath,"barcodesets","temps","consensus_90perc_prederr_combined_barcodes.fa")) as infile:
            l=infile.readlines()
            for i,j in enumerate(l):
                if ">" in j:
                    samplid=j[1:].split(";")[0]
                    self.errbarcodeset[samplid]=j+l[i+1]
        os.mkdir(os.path.join(self.demoutpath,"Main_barcode_results"))
        self.nsinfinalbarcodes,self.nfilteredbarcodes,self.nperfectbarcodes,self.n1to5errbarcodes,self.n6to10errbarcodes,self.n11to15errbarcodes,self.nover16errbarcodes=0,0,0,0,0,0,0
        perfectlist,filteredlist,n1to5errlist,n6to10errlist,n11to15errlist,nover16list,alllist=[],[],[],[],[],[],[]
        with open(os.path.join(self.demoutpath,"barcodesets","Final_all_combined_barcodes.fa")) as infile1:
            with open(os.path.join(self.demoutpath,"Main_barcode_results","QC_Compliant_barcodes_noamb_noerr.fa"),'w') as outfile1:
                with open(os.path.join(self.demoutpath,"Main_barcode_results","Filtered_barcodes_1percamb_upto5err.fa"),'w') as outfilefiltered:
                    with open(os.path.join(self.demoutpath,"Main_barcode_results","Allbarcodes.fa"),'w') as outfilefinal:
                        with open(os.path.join(self.demoutpath,"Main_barcode_results","Fixed_barcodes_1to5err.fa"),'w') as outfile2:
                            with open(os.path.join(self.demoutpath,"Main_barcode_results","Fixed_barcodes_6to10err.fa"),'w') as outfile3:
                                with open(os.path.join(self.demoutpath,"Main_barcode_results","Fixed_barcodes_11to15err.fa"),'w') as outfile4:
                                    with open(os.path.join(self.demoutpath,"Main_barcode_results","Fixed_barcodes_16to20err.fa"),'w') as outfile5:
                                        l=infile1.readlines()
                                          
                        
                                        for i,j in enumerate(l):
                                            if ">" in j:
                                                self.nfinal+=1
                                                self.nsinfinalbarcodes+=l[i+1].strip().count("N")
                                                outfilefinal.write(j+l[i+1])
                                                alllist.append(j.split(";")[0][1:])
                                                if l[i+1].count("N")==0:
                                                    if int(j.split("estgaps=")[1].strip())==0:
                                                        outfile1.write(j+l[i+1])
                                                        self.nperfectbarcodes+=1
                                                        perfectlist.append(j.split(";")[0][1:])
                                                if l[i+1].count("N")<=self.explen*0.01:
                                                    if int(j.split("estgaps=")[1].strip())<=5:
                                                        outfilefiltered.write(j+l[i+1])
                                                        self.nfilteredbarcodes+=1
                                                        if j.split(";")[0][1:] not in perfectlist:
                                                            filteredlist.append(j.split(";")[0][1:])
                                                if int(j.split("estgaps=")[1].strip())>0:
                                                    if int(j.split("estgaps=")[1].strip())<=5:
                                                        outfile2.write(j+l[i+1])
                                                        self.n1to5errbarcodes+=1
                                                        n1to5errlist.append(j.split(";")[0][1:])
                                                if int(j.split("estgaps=")[1].strip())>5:
                                                    if int(j.split("estgaps=")[1].strip())<=10:
                                                        outfile3.write(j+l[i+1])
                                                        self.n6to10errbarcodes+=1
                                                        n6to10errlist.append(j.split(";")[0][1:])
                                                if int(j.split("estgaps=")[1].strip())>10:
                                                    if int(j.split("estgaps=")[1].strip())<=15:
                                                        outfile4.write(j+l[i+1])
                                                        self.n11to15errbarcodes+=1
                                                        n11to15errlist.append(j.split(";")[0][1:])
                                                if int(j.split("estgaps=")[1].strip())>15:
                                                #   if int(j.split("estgaps=")[1].strip())<=20:
                                                    outfile5.write(j+l[i+1])
                                                    self.nover16errbarcodes+=1
                                                    nover16list.append(j.split(";")[0][1:])
        errlist=[]
        with open(os.path.join(self.demoutpath,"Main_barcode_results","Remaining.fa"),'w') as outfile1:
            for each in self.errbarcodeset.keys():
                if each not in alllist:
                    outfile1.write(self.errbarcodeset[each])
                    errlist.append(each)
                    self.nerr+=1
        self.logfile.write(str(round(time.time())) +": The pipeline is completed. "+str(self.nfinal)+ " consensus barcodes have been built.\nPlease check "+os.path.join(self.demoutpath,"Main_barcode_results")+ " for the final output")
        os.mkdir(os.path.join(self.demoutpath,"Main_barcode_results","QC_Compliant"))
        os.mkdir(os.path.join(self.demoutpath,"Main_barcode_results","Filtered"))
        os.mkdir(os.path.join(self.demoutpath,"Main_barcode_results","1to5errors"))
        os.mkdir(os.path.join(self.demoutpath,"Main_barcode_results","6to10errors"))
        os.mkdir(os.path.join(self.demoutpath,"Main_barcode_results","11to15errors"))
        os.mkdir(os.path.join(self.demoutpath,"Main_barcode_results","Over16errors"))
        os.mkdir(os.path.join(self.demoutpath,"Main_barcode_results","Remaining"))
        outdirlist=[os.path.join(self.demoutpath,"Main_barcode_results","QC_Compliant"),os.path.join(self.demoutpath,"Main_barcode_results","Filtered"),os.path.join(self.demoutpath,"Main_barcode_results","1to5errors"),os.path.join(self.demoutpath,"Main_barcode_results","6to10errors"),os.path.join(self.demoutpath,"Main_barcode_results","11to15errors"),os.path.join(self.demoutpath,"Main_barcode_results","Over16errors"),os.path.join(self.demoutpath,"Main_barcode_results","Remaining")]
        if self.inputmode==1:
            self.mycopyfiles=copyfiles(os.path.join(self.demoutpath,"demultiplexed"),outdirlist,[perfectlist,filteredlist,n1to5errlist,n6to10errlist,n11to15errlist,nover16list,errlist],os.path.join(self.demoutpath),{},1)
            self.mycopyfiles.taskFinished.connect(self.printfinalout)
            self.runcopy()
        elif self.inputmode==2:
            self.mycopyfiles=copyfiles(os.path.join(self.indir),outdirlist,[perfectlist,filteredlist,n1to5errlist,n6to10errlist,n11to15errlist,nover16list,errlist],os.path.join(self.demoutpath),self.dirdict,2)
            self.mycopyfiles.taskFinished.connect(self.printfinalout)
            self.runcopy()
        else:
            self.printfinalout()
    def runcopy(self):
        self.mycopyfiles.start()
    def printfinalout(self):
        self.outputtable=QtWidgets.QTableWidget()
        
        self.outputtable.setColumnCount(2)
        self.outputtable.setMaximumWidth(self.size().width())
        self.outputtable.setSizePolicy(QtWidgets.QSizePolicy.Minimum,QtWidgets.QSizePolicy.Minimum)
        self.outputtable.setColumnWidth(0,int((self.outputtable.size().width())*0.75))
        self.outputtable.setColumnWidth(1,int((self.outputtable.size().width())*0.25))
        if self.inputmode==2:
            self.indir
            self.outputtable.setRowCount(14)
            self.outputtable.setItem(0,0,QtWidgets.QTableWidgetItem("Number of good barcodes obtained after first alignment of up to 200 reads"))
            self.outputtable.setItem(1,0,QtWidgets.QTableWidgetItem("Number of erroneous barcodes obtained after first alignment of up to 200 reads"))
            self.outputtable.setItem(2,0,QtWidgets.QTableWidgetItem("Number of good barcodes obtained after aligning similar reads"))
            self.outputtable.setItem(3,0,QtWidgets.QTableWidgetItem("Number of erroneous barcodes obtained after aligning similar reads"))
            self.outputtable.setItem(4,0,QtWidgets.QTableWidgetItem("Number of barcodes fixed"))
            self.outputtable.setItem(5,0,QtWidgets.QTableWidgetItem("Final number of barcodes of expected length"))
            self.outputtable.setItem(6,0,QtWidgets.QTableWidgetItem("Final number of barcodes that cannot be fixed"))
            self.outputtable.setItem(7,0,QtWidgets.QTableWidgetItem("Number of Ns in final barcodes"))
            self.outputtable.setItem(8,0,QtWidgets.QTableWidgetItem("Number of filtered barcodes"))
            self.outputtable.setItem(9,0,QtWidgets.QTableWidgetItem("Number of QC_Compliant barcodes"))
            self.outputtable.setItem(10,0,QtWidgets.QTableWidgetItem("Number of barcodes with 1-5 errors"))
            self.outputtable.setItem(11,0,QtWidgets.QTableWidgetItem("Number of barcodes with 6-10 errors"))
            self.outputtable.setItem(12,0,QtWidgets.QTableWidgetItem("Number of barcodes with 11-15 errors"))
            self.outputtable.setItem(13,0,QtWidgets.QTableWidgetItem("Number of barcodes with over 15 errors"))
            self.outputtable.setItem(0,1,QtWidgets.QTableWidgetItem(str(self.con200goodn)))
            self.outputtable.setItem(1,1,QtWidgets.QTableWidgetItem(str(self.con200errn)))
            self.outputtable.setItem(2,1,QtWidgets.QTableWidgetItem(str(self.n90goodn)))
            self.outputtable.setItem(3,1,QtWidgets.QTableWidgetItem(str(self.n90errn)))
            self.outputtable.setItem(4,1,QtWidgets.QTableWidgetItem(str(self.nfixed)))
            self.outputtable.setItem(5,1,QtWidgets.QTableWidgetItem(str(self.nfinal)))
            self.outputtable.setItem(6,1,QtWidgets.QTableWidgetItem(str(self.nerr)))
            self.outputtable.setItem(7,1,QtWidgets.QTableWidgetItem(str(self.nsinfinalbarcodes)))
            self.outputtable.setItem(8,1,QtWidgets.QTableWidgetItem(str(self.nfilteredbarcodes)))
            self.outputtable.setItem(9,1,QtWidgets.QTableWidgetItem(str(self.nperfectbarcodes)))
            self.outputtable.setItem(10,1,QtWidgets.QTableWidgetItem(str(self.n1to5errbarcodes)))
            self.outputtable.setItem(11,1,QtWidgets.QTableWidgetItem(str(self.n6to10errbarcodes)))
            self.outputtable.setItem(12,1,QtWidgets.QTableWidgetItem(str(self.n11to15errbarcodes)))
            self.outputtable.setItem(13,1,QtWidgets.QTableWidgetItem(str(self.nover16errbarcodes)))
        if self.inputmode==1:
            self.outputtable.setRowCount(19)
            self.outputtable.setItem(0,0,QtWidgets.QTableWidgetItem("Number of reads in file"))
            self.outputtable.setItem(1,0,QtWidgets.QTableWidgetItem("Number of reads passing length filter"))
            self.outputtable.setItem(2,0,QtWidgets.QTableWidgetItem("Number of reads used for demultiplexing"))
            self.outputtable.setItem(3,0,QtWidgets.QTableWidgetItem("Number of samples in demultiplexing file"))
            self.outputtable.setItem(4,0,QtWidgets.QTableWidgetItem("Number of samples with >=5X coverage"))
            self.outputtable.setItem(5,0,QtWidgets.QTableWidgetItem("Number of good barcodes obtained after first alignment of up to 200 reads"))
            self.outputtable.setItem(6,0,QtWidgets.QTableWidgetItem("Number of erroneous barcodes obtained after first alignment of up to 200 reads"))
            self.outputtable.setItem(7,0,QtWidgets.QTableWidgetItem("Number of good barcodes obtained after aligning similar reads"))
            self.outputtable.setItem(8,0,QtWidgets.QTableWidgetItem("Number of erroneous barcodes obtained after aligning similar reads"))
            self.outputtable.setItem(9,0,QtWidgets.QTableWidgetItem("Number of barcodes fixed"))
            self.outputtable.setItem(10,0,QtWidgets.QTableWidgetItem("Final number of barcodes of expected length"))
            self.outputtable.setItem(11,0,QtWidgets.QTableWidgetItem("Final number of barcodes that cannot be fixed"))
            self.outputtable.setItem(12,0,QtWidgets.QTableWidgetItem("Number of Ns in final barcodes"))
            self.outputtable.setItem(13,0,QtWidgets.QTableWidgetItem("Number of filtered barcodes"))
            self.outputtable.setItem(14,0,QtWidgets.QTableWidgetItem("Number of QC_Compliant barcodes"))
            self.outputtable.setItem(15,0,QtWidgets.QTableWidgetItem("Number of barcodes with 1-5 errors"))
            self.outputtable.setItem(16,0,QtWidgets.QTableWidgetItem("Number of barcodes with 6-10 errors"))
            self.outputtable.setItem(17,0,QtWidgets.QTableWidgetItem("Number of barcodes with 11-15 errors"))
            self.outputtable.setItem(18,0,QtWidgets.QTableWidgetItem("Number of barcodes with over 15 errors"))
            self.outputtable.setItem(0,1,QtWidgets.QTableWidgetItem(str(self.totalseqs)))
            self.outputtable.setItem(1,1,QtWidgets.QTableWidgetItem(str(self.nseqspasslen)))
            self.outputtable.setItem(2,1,QtWidgets.QTableWidgetItem(str(self.nseqsfordemultiplexing)))
            self.outputtable.setItem(3,1,QtWidgets.QTableWidgetItem(str(len(self.sampleids))))
            self.outputtable.setItem(4,1,QtWidgets.QTableWidgetItem(str(self.nsampledemultiplexed5)))
            self.outputtable.setItem(5,1,QtWidgets.QTableWidgetItem(str(self.con200goodn)))
            self.outputtable.setItem(6,1,QtWidgets.QTableWidgetItem(str(self.con200errn)))
            self.outputtable.setItem(7,1,QtWidgets.QTableWidgetItem(str(self.n90goodn)))
            self.outputtable.setItem(8,1,QtWidgets.QTableWidgetItem(str(self.n90errn)))
            self.outputtable.setItem(9,1,QtWidgets.QTableWidgetItem(str(self.nfixed)))
            self.outputtable.setItem(10,1,QtWidgets.QTableWidgetItem(str(self.nfinal)))
            self.outputtable.setItem(11,1,QtWidgets.QTableWidgetItem(str(self.nerr)))
            self.outputtable.setItem(12,1,QtWidgets.QTableWidgetItem(str(self.nsinfinalbarcodes)))
            self.outputtable.setItem(13,1,QtWidgets.QTableWidgetItem(str(self.nfilteredbarcodes)))
            self.outputtable.setItem(14,1,QtWidgets.QTableWidgetItem(str(self.nperfectbarcodes)))
            self.outputtable.setItem(15,1,QtWidgets.QTableWidgetItem(str(self.n1to5errbarcodes)))
            self.outputtable.setItem(16,1,QtWidgets.QTableWidgetItem(str(self.n6to10errbarcodes)))
            self.outputtable.setItem(17,1,QtWidgets.QTableWidgetItem(str(self.n11to15errbarcodes)))
            self.outputtable.setItem(18,1,QtWidgets.QTableWidgetItem(str(self.nover16errbarcodes)))
        elif self.inputmode==3:
            self.outputtable.setRowCount(12)
            self.outputtable.setItem(0,0,QtWidgets.QTableWidgetItem("Number of good barcodes "))
            self.outputtable.setItem(1,0,QtWidgets.QTableWidgetItem("Number of erroneous barcodes "))
            self.outputtable.setItem(2,0,QtWidgets.QTableWidgetItem("Number of barcodes fixed"))
            self.outputtable.setItem(3,0,QtWidgets.QTableWidgetItem("Final number of barcodes of expected length"))
            self.outputtable.setItem(4,0,QtWidgets.QTableWidgetItem("Final number of barcodes that cannot be fixed"))
            self.outputtable.setItem(5,0,QtWidgets.QTableWidgetItem("Number of Ns in final barcodes"))
            self.outputtable.setItem(6,0,QtWidgets.QTableWidgetItem("Number of filtered barcodes"))
            self.outputtable.setItem(7,0,QtWidgets.QTableWidgetItem("Number of QC_Compliant barcodes"))
            self.outputtable.setItem(8,0,QtWidgets.QTableWidgetItem("Number of barcodes with 1-5 errors"))
            self.outputtable.setItem(9,0,QtWidgets.QTableWidgetItem("Number of barcodes with 6-10 errors"))
            self.outputtable.setItem(10,0,QtWidgets.QTableWidgetItem("Number of barcodes with 11-15 errors"))
            self.outputtable.setItem(11,0,QtWidgets.QTableWidgetItem("Number of barcodes with over 15 errors"))
            self.outputtable.setItem(0,1,QtWidgets.QTableWidgetItem(str(self.con200goodn)))
            self.outputtable.setItem(1,1,QtWidgets.QTableWidgetItem(str(self.con200errn)))
            self.outputtable.setItem(2,1,QtWidgets.QTableWidgetItem(str(self.nfixed)))
            self.outputtable.setItem(3,1,QtWidgets.QTableWidgetItem(str(self.nfinal)))
            self.outputtable.setItem(4,1,QtWidgets.QTableWidgetItem(str(self.nerr)))
            self.outputtable.setItem(5,1,QtWidgets.QTableWidgetItem(str(self.nsinfinalbarcodes)))
            self.outputtable.setItem(6,1,QtWidgets.QTableWidgetItem(str(self.nfilteredbarcodes)))
            self.outputtable.setItem(7,1,QtWidgets.QTableWidgetItem(str(self.nperfectbarcodes)))
            self.outputtable.setItem(8,1,QtWidgets.QTableWidgetItem(str(self.n1to5errbarcodes)))
            self.outputtable.setItem(9,1,QtWidgets.QTableWidgetItem(str(self.n6to10errbarcodes)))
            self.outputtable.setItem(10,1,QtWidgets.QTableWidgetItem(str(self.n11to15errbarcodes)))
            self.outputtable.setItem(11,1,QtWidgets.QTableWidgetItem(str(self.nover16errbarcodes)))
        with open(os.path.join(self.demoutpath,"Outputtable.csv"),'w') as tabfile:
            if self.inputmode==2:
                self.indir
                tabfile.write("\nNumber of samples in demultiplexing file,"+str(len(self.sampleids)))
                tabfile.write("\nNumber of good barcodes obtained after first alignment of up to 200 reads,"+str(self.con200goodn))
                tabfile.write("\nNumber of erroneous barcodes obtained after first alignment of up to 200 reads,"+str(self.con200errn))
                tabfile.write("\nNumber of good barcodes obtained after aligning similar reads,"+str(self.n90goodn))
                tabfile.write("\nNumber of erroneous barcodes obtained after aligning similar reads,"+str(self.n90errn))
                tabfile.write("\nNumber of barcodes fixed,"+str(self.nfixed))
                tabfile.write("\nFinal number of barcodes,"+str(self.nfinal))
                tabfile.write("\nFinal number of barcodes that cannot be fixed,"+str(self.nerr))
                tabfile.write("\nNumber of Ns in final barcodes,"+str(self.nsinfinalbarcodes))
                tabfile.write("\nNumber of filtered barcodes,"+str(self.nfilteredbarcodes))
                tabfile.write("\nNumber of QC_Compliant barcodes,"+str(self.nperfectbarcodes))
                tabfile.write("\nNumber of barcodes with 1-5 errors,"+str(self.n1to5errbarcodes))
                tabfile.write("\nNumber of barcodes with 6-10 errors,"+str(self.n6to10errbarcodes))
                tabfile.write("\nNumber of barcodes with 11-15 errors,"+str(self.n11to15errbarcodes))
                tabfile.write("\nNumber of barcodes with over 15 errors,"+str(self.nover16errbarcodes))
            elif self.inputmode==1:
                tabfile.write("Number of reads in file,"+str(self.totalseqs))
                tabfile.write("\nNumber of reads passing length filter,"+str(self.nseqspasslen))
                tabfile.write("\nNumber of reads used for demultiplexing,"+str(self.nseqsfordemultiplexing))
                tabfile.write("\nNumber of samples in demultiplexing file,"+str(len(self.sampleids)))
                tabfile.write("\nNumber of samples with >=5X coverage,"+str(self.nsampledemultiplexed5))
                tabfile.write("\nNumber of good barcodes obtained after first alignment of up to 200 reads,"+str(self.con200goodn))
                tabfile.write("\nNumber of erroneous barcodes obtained after first alignment of up to 200 reads,"+str(self.con200errn))
                tabfile.write("\nNumber of good barcodes obtained after aligning similar reads,"+str(self.n90goodn))
                tabfile.write("\nNumber of erroneous barcodes obtained after aligning similar reads,"+str(self.n90errn))
                tabfile.write("\nNumber of barcodes fixed,"+str(self.nfixed))
                tabfile.write("\nFinal number of barcodes,"+str(self.nfinal))
                tabfile.write("\nFinal number of barcodes that cannot be fixed,"+str(self.nerr))
                tabfile.write("\nNumber of Ns in final barcodes,"+str(self.nsinfinalbarcodes))
                tabfile.write("\nNumber of filtered barcodes,"+str(self.nfilteredbarcodes))
                tabfile.write("\nNumber of QC_Compliant barcodes,"+str(self.nperfectbarcodes))
                tabfile.write("\nNumber of barcodes with 1-5 errors,"+str(self.n1to5errbarcodes))
                tabfile.write("\nNumber of barcodes with 6-10 errors,"+str(self.n6to10errbarcodes))
                tabfile.write("\nNumber of barcodes with 11-15 errors,"+str(self.n11to15errbarcodes))
                tabfile.write("\nNumber of barcodes with over 15 errors,"+str(self.nover16errbarcodes))
        self.outputtable.installEventFilter(self)
        self.statuswidget.logstatus.textCursor().insertHtml('<br><p style="color:blue;"><b>Final results</b></p>Results are saved in '+str(os.path.join(self.demoutpath,"Main_barcode_results"))+", "+str(os.path.join(self.demoutpath,"Outputtable.tsv"))+" and "+str(os.path.join(self.demoutpath,"runsummary.xlsx"))+"<br><br>See 'Results of Barcode calling' tab for the final output")
        self.showwarningdialog("The pipeline is completed. Please check "+os.path.join(self.demoutpath,"Main_barcode_results")+ " and the 'Results of Barcode calling' tab for the final output ")
        self.statuswidget.logstatus.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse) 
                                                                                                                                                                                                                                                   
        self.Tab.addTab(self.outputtable,"Results of Barcode calling")
        self.statuswidget.pbarTitle.setText("Please check 'Results of Barcode calling' tab for summary")
    #   self.statuswidget.layout.addWidget(self.outputtable,2,0,16,5)
        self.logfile.close()
    def messagebox(self,string):
        mb=QtWidgets.QMessageBox()
        mb.setIcon(QtWidgets.QMessageBox.Information)
        mb.setText(string)
        mb.setStandardButtons(QtWidgets.QMessageBox.Ok)
        mb.setWindowTitle("ONTBarcoder")
        mb.setStyleSheet("QMessageBox { messagebox-text-interaction-flags: 5; }")
        mb.exec_()

    def messagebox3(self,string):
        self.mb=QtWidgets.QMessageBox()
        self.mb.setText(string)
        self.mb.setIcon(QtWidgets.QMessageBox.Question)
        selectdirectory=QtWidgets.QPushButton('Select Directory')
        selectdirectory.clicked.connect(self.setdemoutpath2)
        defaultdirectory=QtWidgets.QPushButton('Go with default')
        defaultdirectory.clicked.connect(self.setdemoutpath3)
        self.mb.addButton(selectdirectory, QtWidgets.QMessageBox.YesRole)
        self.mb.addButton(defaultdirectory, QtWidgets.QMessageBox.NoRole)
        self.mb.setWindowTitle("ONTBarcoder")
        self.mb.exec_()
    def messagebox2(self,string):
        self.mb=QtWidgets.QMessageBox()
        self.mb.setText(string)
        self.mb.setIcon(QtWidgets.QMessageBox.Question)
        selectdirectory=QtWidgets.QPushButton('Select Directory')
        selectdirectory.clicked.connect(self.setcompoutpath2)
        defaultdirectory=QtWidgets.QPushButton('Go with default')
        defaultdirectory.clicked.connect(self.setcompoutpath3)
        self.mb.addButton(selectdirectory, QtWidgets.QMessageBox.YesRole)
        self.mb.addButton(defaultdirectory, QtWidgets.QMessageBox.NoRole)
        self.mb.setWindowTitle("ONTBarcoder")
        self.mb.exec_()

    def showerrordialog(self,string):
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Critical)
        msg.setText(string)
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        retval = msg.exec_()
                                                                            
                                                                          
                                   
                                         
                                        
                               
                                   
                                                                                                                      
                                                                                
                                                                              
                                       
             
                                        
                               
                                    
                                                                                                                                                            
                                                                                
                                                                              
                                                    

        
    def showwarningdialog(self,string):
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)
        msg.setText(string)
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        retval = msg.exec_()
    def showrefdalog(self):
        self.d=QtWidgets.QDialog()
        layout=QtWidgets.QGridLayout()
        layout.setSpacing(20)
        self.d.setLayout(layout)
        label=QtWidgets.QLabel("Are any of these files a reference that you would like to compare to?<br> If nothing is selected, all samples will be compared to each other.")
        label.setFont(QtGui.QFont("Times", 12, QtGui.QFont.Bold))
        layout.addWidget(label,0,0,1,2)
        r=1
        self.comprefstatdict=False
        self.blist=[]
        self.comprefiddict={}
        for i,each in enumerate(self.compfilelist):
            self.comprefiddict[each]=i
            b=QtWidgets.QRadioButton(each)
            b.setChecked(False)
            b.toggled.connect(lambda: self.changestate(i))
            self.blist.append(b)
            layout.addWidget(b,r,0,1,2)
            r+=1
        self.okbtn=QtWidgets.QPushButton("Done")
        self.okbtn.clicked.connect(self.close)
        layout.addWidget(self.okbtn,r+1,1)
                                                            
        self.d.show()
    def changestate(self,i):
        source = self.sender()
        if self.d.sender().isChecked()==True:
            self.comprefstatdict=self.d.sender().text()
        else:
            self.comprefstatdict=False
    #   print self.comprefstatdict
    def close(self):
        self.setcomparisons()
    def setcomparisons(self):
        self.d.close()
    #   print self.comprefstatdict
        self.compstacklayout.addWidget(self.compbar,1,0,1,1)
        self.tempdir=tempfile.mkdtemp()
        self.seqdictlist={}
        self.seqnamelist={}
        filelist=[]
        for fname in self.compfilelist:
            basename=os.path.basename(fname)
            filelist.append(basename)
            minionbarcodes={}
            minionbarcodesnames={}
            count=0
            with open(fname) as minionfile:
                l=minionfile.readlines()
                seqnameor,seqnamenew='',''
                seq=''  
                for i,each in enumerate(l):
                    if i!=len(l)-1:
                        if ">" in each:
                            count+=1
                            if seqnamenew!='':
                                try:
                                    minionbarcodes[seqnamenew].append(seq)
                                    minionbarcodesnames[seqnamenew].append(seqnameor)
                                except KeyError:
                                    minionbarcodes[seqnamenew]=[seq]
                                    minionbarcodesnames[seqnamenew]=[seqnameor]
                                seq=''
                            seqnameor=each
                            seqnamenew=each.strip()[1:].split("_")[0].split("-")[0].split(";")[0]
                            
                        else:
                            seq+=each.strip()
                    else:
                        seq+=each.strip()
                        try:
                            minionbarcodes[seqnamenew].append(seq)
                            minionbarcodesnames[seqnamenew].append(seqnameor)
                        except KeyError:
                            minionbarcodes[seqnamenew]=[seq]
                            minionbarcodesnames[seqnamenew]=[seqnameor]
                if self.comprefstatdict!=False:
                #   print fname,os.path.basename(self.comprefstatdict)
                    if basename==os.path.basename(self.comprefstatdict):
                        self.refseqcount=count
        #   print minionbarcodes
            with open(os.path.join(self.tempdir,basename)+'_nodups','w') as barcodes313file:
                for k in minionbarcodes.keys():
                    minissue=100000
                    seqname=''
                    seq=''
                    for i,mk in enumerate(minionbarcodes[k]):
                        if "ambs=" in minionbarcodesnames[k][i]:
                            if "estgaps=" in minionbarcodesnames[k][i]:
                                totalissues=int(minionbarcodesnames[k][i].split(";")[3].split("=")[1])+int(minionbarcodesnames[k][i].split(";")[4].split("=")[1].strip())
                                if totalissues<minissue:
                                    seq=mk.strip()+'\n'
                                    seqname=minionbarcodesnames[k][i]
                            else:
                                seq=mk.strip()+'\n'
                                seqname=minionbarcodesnames[k][i]
                        else:
                            seq=mk.strip()+'\n'
                            seqname=minionbarcodesnames[k][i]
                    barcodes313file.write(seqname+seq)
    #   self.refseqcount=0
        for fname in self.compfilelist:

            basename=os.path.basename(fname)
            with open(os.path.join(self.tempdir,basename)+'_nodups') as referencefile:
                l=referencefile.readlines()
                for i,each in enumerate(l):
                    if ">" in each:
                        count+=1
                        try:
                            self.seqdictlist[each.replace("-","_").strip()[1:].split("_")[0].split(";")[0]][basename]=l[i+1].strip().upper()
                            self.seqnamelist[each.replace("-","_").strip()[1:].split("_")[0].split(";")[0]][basename]=">"+basename+"_"+each[1:]
                        except KeyError:
                            self.seqdictlist[each.replace("-","_").strip()[1:].split("_")[0].split(";")[0]]={basename:l[i+1].strip().upper()}
                            self.seqnamelist[each.replace("-","_").strip()[1:].split("_")[0].split(";")[0]]={basename:">"+basename+"_"+each[1:]}

                    
        self.compbar.setMaximum(len(self.seqdictlist))
        self.compbar.setFormat("%v"+"/"+str(len(self.seqdictlist)))
        if len(self.compfilelist)==2:
            self.compmode=2
            self.myruncomparisons=runtwocomparisons(self.seqdictlist,self.seqnamelist,self.compoutpath)
            self.myruncomparisons.notifyProgress.connect(self.updatecompbar)
            self.myruncomparisons.taskFinished.connect(self.printcompdone)
            self.runmycomparisons()
        elif self.comprefstatdict==False:
            if len(self.compfilelist)>2:
                self.compmode=3
            #   print "multisample"
                self.myruncomparisons=runmulticomparisons(self.seqdictlist,self.seqnamelist,self.compoutpath,filelist)
                self.myruncomparisons.notifyProgress.connect(self.updatecompbar)
                self.myruncomparisons.taskFinished.connect(self.printcompdone)
                self.runmycomparisons()
        else:
            if len(self.compfilelist)>2:
                self.compmode=4
            #   print "pairwisemode"
                self.myruncomparisons=runpaircomparisons(self.seqdictlist,self.seqnamelist,self.compoutpath,filelist,os.path.basename(self.comprefstatdict))
                self.myruncomparisons.notifyProgress.connect(self.updatecompbar)
                self.myruncomparisons.taskFinished.connect(self.printcompdone)
                self.runmycomparisons()             
    def updatecompbar(self,i):
        self.compbar.setValue(i)
    def runmycomparisons(self):
        self.comptablabel.setText("<center>Running the comparisons of barcode sets</center>")
        self.myruncomparisons.start()

    def printcompdone(self,i):
        outlist=i
    #   print "done"
        if self.compmode==2:
            strdictlens=""
            for each in outlist[3].keys():
                strdictlens+="Number of barcodes in "+ each+" only: "+str(outlist[3][each])+"<br>"
            self.comptablabel.setText("Comparisons are done<br>Number of Identical Barcodes: "+str(outlist[0])+"<br>Number of Compatible barcodes: "+str(outlist[1])+"<br>Number of incompatiblebarcodes: "+str(outlist[2])+"<br>"+strdictlens +"<br><br>Your outputs are saved in "+str(self.compoutpath)+".<br><br><b> You can drag another set of files to compare a different set of barcodes")
            self.comptablabel.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse)
        if self.compmode==3:
            strdictlens=""
        #   for each in outlist[2].keys():
        #       strdictlens+="Number of barcodes in "+ each+" only: "+str(outlist[2][each])+"<br>"
            self.comptablabel.setText("Comparisons are done<br>Number of Identical/compatible barcodes across all sets: "+str(outlist[0])+"<br>Number of specimens with at least one incompatible barcode: "+str(outlist[1])+"<br>"+strdictlens+"<br><br>Your outputs are saved in "+str(self.compoutpath)+".<br><br><b> You can drag another set of files to compare a different set of barcodes")
            self.comptablabel.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse)
        if self.compmode==4:
            strdictlens=""
            strdictlens2=''
            for each in outlist[1].keys():
                if each!=os.path.basename(self.comprefstatdict).split(".")[0]:
                    strdictlens+="<b>Results of "+each+"</b><br>Total number of barcodes: "+ each+" "+str(outlist[1][each][4])+"<br>Number compared: "+str(outlist[1][each][3])+"<br>Number identical: "+str(outlist[1][each][0])+"<br> Number compatible: "+str(outlist[1][each][1])+"<br> Number incompatible: "+str(outlist[1][each][2])+"<br><br>"
                    strdictlens2+="Results of "+each+"\nTotal number of barcodes: "+ each+" "+str(outlist[1][each][4])+"\nNumber compared: "+str(outlist[1][each][3])+"\nNumber identical: "+str(outlist[1][each][0])+"\nNumber compatible: "+str(outlist[1][each][1])+"\nNumber incompatible: "+str(outlist[1][each][2])+"\n\n"
            self.comptablabel.setText("Comparisons are done<br>Number of reference barcodes: "+str(self.refseqcount)+"<br>"+strdictlens+"<br><br>Your outputs are saved in "+str(self.compoutpath)+".<br><br><b> You can drag another set of files to compare a different set of barcodes")
            self.comptablabel.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse)
            with open(os.path.join(self.compoutpath,"resultssummary.txt"),'w') as outfile:
                outfile.write("Comparisons are done\nNumber of reference barcodes: "+str(self.refseqcount)+"\n"+strdictlens2)

    def translate_corframe(self,seq,gencode):
        translatedset=[]
        each=seq
        corframe=self.get_cor_frame(each.replace("-",""),gencode)
        if corframe==1:
            translation=Seq(each.replace('-','')).translate(table=gencode,to_stop=True).__str__()
            explen=len(each.replace('-',''))
        if corframe==2:
            translation=Seq(each[1:].replace('-','')).translate(table=gencode,to_stop=True).__str__()
            explen=len(each[1:].replace('-',''))
        if corframe==3:
            translation=Seq(each[2:].replace('-','')).translate(table=gencode,to_stop=True).__str__()
            explen=len(each[2:].replace('-',''))
        if corframe==4:
            translation=Seq(each.replace('-','')).reverse_complement().translate(table=gencode,to_stop=True).__str__()
            explen=len(Seq(each.replace('-','')).__str__())
        if corframe==5:
            translation=Seq(each[:-1].replace('-','')).reverse_complement().translate(table=gencode,to_stop=True).__str__()
            explen=len(Seq(each[:-1].replace('-','')).__str__())
        if corframe==6:
            translation=Seq(each[:-2].replace('-','')).reverse_complement().translate(table=gencode,to_stop=True).__str__()
            explen=len(Seq(each[:-2].replace('-','')).__str__())
        if len(translation)<int(explen/3):
            return "0"
        elif len(translation)==int(explen/3):
            return "1"

    def get_cor_frame(self,seq,gencode):
        seqset=[Seq(seq),Seq(seq[1:]),Seq(seq[2:]),Seq(seq).reverse_complement(),Seq(seq[:-1]).reverse_complement(),Seq(seq[:-2]).reverse_complement()]
        aminoset=[]
        for each in seqset:
            a=each.translate(table=gencode,to_stop=True)
            aminoset.append(a.__str__())
        maxlen,corframe=0,0
        for i,each in enumerate(aminoset):
            if len(each)>maxlen:
                maxlen=len(each)
                corframe=i+1
        return corframe
    def changeEvent(self, event):
        return super(OptWindow, self).changeEvent(event)
    def eventFilter(self, source, event):
        if (event.type() == QtCore.QEvent.KeyPress and
            event.matches(QtGui.QKeySequence.Copy)):
            self.copySelection()
            return True
        return super(OptWindow, self).eventFilter(source, event)
    def copySelection(self):
        selectedcells = self.outputtable.selectedIndexes()
        if selectedcells:
            rows = sorted(index.row() for index in selectedcells)
            columns = sorted(index.column() for index in selectedcells)
            rowcount = rows[-1] - rows[0] + 1
            colcount = columns[-1] - columns[0] + 1
            table = [[''] * colcount for _ in range(rowcount)]
            for index in selectedcells:
                row = index.row() - rows[0]
                column = index.column() - columns[0]
                table[row][column] = index.data()
            stream = io.BytesIO()
            csv.writer(stream, delimiter='\t').writerows(table)
            QtWidgets.qApp.clipboard().setText(stream.getvalue())
        return
if __name__=='__main__':
    multiprocessing.freeze_support()
    app=QtWidgets.QApplication(sys.argv)
    with open("stylesheet.qss",'r') as fh:
        text=fh.read()
    app.setStyleSheet(text)
    ex=OptWindow()
    sys.exit(app.exec_())

client.close()
